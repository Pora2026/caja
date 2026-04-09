import os
import sys
import json
from dotenv import load_dotenv
load_dotenv()  # lee .env local si existe
from functools import wraps
from io import BytesIO
from datetime import date, datetime, timedelta, UTC
from typing import Optional, Tuple

from services.backup_service import perform_backup

from flask import (
    Flask, request, redirect, url_for, render_template_string,
    send_file, send_from_directory, session, abort
)

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash


import hashlib

from openpyxl import load_workbook, Workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ==============================
# FIX STATIC + APP + DB
# ==============================
def resource_path(relative_path: str) -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(sys._MEIPASS, relative_path)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)


STATIC_DIR = resource_path("static")

app = Flask(
    __name__,
    static_folder=STATIC_DIR,
    static_url_path="/static"
)
#  favicon (evita error al pedir /favicon.ico)
@app.route("/favicon.ico")
def favicon():
    return send_from_directory(os.path.join(app.static_folder, "img"), "favicon.ico")

# ==============================
# DB + SECRET (local .env / Render / exe)
# ==============================

# 1) SECRET_KEY desde env (Render o .env local)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-local-caja")

# 2) Si existe DATABASE_URL => usarla (Render / .env)
db_url = os.getenv("DATABASE_URL")

if db_url:
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)

    #  Si es SQLite, convertir a ruta absoluta segura
    if db_url.lower().startswith("sqlite:///"):
        rel_path = db_url.replace("sqlite:///", "", 1)
        base_dir = os.path.dirname(os.path.abspath(__file__))
        abs_path = os.path.join(base_dir, rel_path)
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)

        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{abs_path}"
        DB_PATH = abs_path
    else:
        app.config["SQLALCHEMY_DATABASE_URI"] = db_url
        DB_PATH = db_url
else:
    # 3) Si NO hay DATABASE_URL:
    if getattr(sys, "frozen", False):
        def get_programdata_dir(app_name: str) -> str:
            base = os.environ.get("PROGRAMDATA", r"C:\ProgramData")
            path = os.path.join(base, app_name)
            os.makedirs(path, exist_ok=True)
            return path

        DATA_DIR = get_programdata_dir("PORA")
        INSTANCE_DIR = os.path.join(DATA_DIR, "instance")
        os.makedirs(INSTANCE_DIR, exist_ok=True)
        DB_FILE = os.path.join(INSTANCE_DIR, "caja.db")

        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_FILE}"
        DB_PATH = DB_FILE
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        instance_dir = os.path.join(base_dir, "instance")
        os.makedirs(instance_dir, exist_ok=True)
        DB_FILE = os.path.join(instance_dir, "caja.db")

        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_FILE}"
        DB_PATH = DB_FILE

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

#  Forzar schema por defecto en Postgres (evita "no schema selected")
uri = (app.config.get("SQLALCHEMY_DATABASE_URI") or "").lower()
if uri.startswith("postgresql://"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "connect_args": {"options": "-c search_path=caja"}
    }
    
db = SQLAlchemy(app)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GDRIVE_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbx_XqC6hSrA-zlzjRhVfDv93CGjMO5YhPJuzO3lQ5Djtzbln2zWONzcJIQ8R7KHyJsV/exec"


def backup_caja_payload():
    try:
        shifts = Shift.query.all()
        expenses = CashExpense.query.all()
        closes = ShiftClose.query.all()

        return {
            "exported_at": datetime.now(UTC).isoformat(),
            "shifts": [s.id for s in shifts],
            "expenses": [e.id for e in expenses],
            "closes": [c.id for c in closes],
        }
    except Exception as e:
        print("Backup payload error:", e)
        return {
            "exported_at": datetime.now(UTC).isoformat(),
            "error": str(e),
            "shifts": [],
            "expenses": [],
            "closes": [],
        }


def backup_caja_local_y_drive():
    return perform_backup(
        payload=backup_caja_payload(),
        base_dir=BASE_DIR,
        webhook_url=GDRIVE_WEBHOOK_URL,
        prefix="caja",
    )


# ==============================
# CONFIG
# ==============================

TURNS = [("MORNING", "Manana"), ("AFTERNOON", "Tarde")]
TURN_NAMES = dict(TURNS)

WEEKDAYS_ES = ["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]

def weekday_es(d: date) -> str:
    try:
        return WEEKDAYS_ES[d.weekday()]
    except Exception:
        return ""

CATEGORIES = [
    "Insumos urgentes",
    "Delivery / Cadete (efectivo)",
    "Mantenimiento / Varios",
    "Imprevistos",
    "Otros (requiere nota)",
]

EMPLOYEES = ["Paula", "Pato", "Lautaro", "Sofia", "Matias"]
ADMINS = ["Bernardo", "Ximena"]

MAX_CONSUMOS = 5

# Asistencia - jornada fija y grupos
JORNADA_MIN = 9 * 60 + 10  # 09:10 => 550 min

GROUP_A = {
    "morning_in": "07:55",
    "morning_out": "12:30",
    "afternoon_in": "15:55",
    "afternoon_out": "20:30",
}
GROUP_B = {
    "morning_in": "07:55",
    "morning_out": "13:00",
    "afternoon_in": "16:55",
    "afternoon_out": "21:00",
}

# Limites mensuales (pagos)
RP_LIMIT_MIN_PER_MONTH = 120                 # 2hs por mes (minutos)
SICK_LIMIT_MIN_PER_MONTH = 2 * JORNADA_MIN   # 2 dias por mes (minutos)

NOVELTY_ITEMS = [
    "",  # normal
    "Tardanza",
    "Inasistencia",
    "Razones particulares",
    "Enfermedad",
    "Curso",
    "Vacaciones",
    "Delivery",
    "Otros",
]

# ==============================
# MODELOS
# ==============================
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=True)
    role = db.Column(db.String(10), default="user")  # admin / user
    is_active = db.Column(db.Integer, default=1)

    # ===== Mobile PIN (PORA Mobile) =====
    # Guardamos hash (para validar) + fingerprint deterministico (para buscar y garantizar unicidad).
    mobile_pin_hash = db.Column(db.String(255), nullable=True)
    mobile_pin_fingerprint = db.Column(db.String(64), unique=True, nullable=True)
    mobile_pin_attempts = db.Column(db.Integer, default=0)
    mobile_pin_locked_until = db.Column(db.DateTime, nullable=True)


class Shift(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.Date, nullable=False)
    turn = db.Column(db.String(10), nullable=False)  # MORNING / AFTERNOON
    responsible = db.Column(db.String(50), nullable=False)

    opening_cash = db.Column(db.Integer, nullable=False)

    sales_cash = db.Column(db.Integer, default=0)   # en tu operacion: efectivo NETO (ya desconto gastos)
    sales_mp = db.Column(db.Integer, default=0)
    sales_pya = db.Column(db.Integer, default=0)
    sales_rappi = db.Column(db.Integer, default=0)
    sales_apps = db.Column(db.Integer, default=0)  # legacy

    # Draft persistente de la calculadora de delivery
    delivery_data_json = db.Column(db.Text)
    hour_shift = db.Column(db.String(10))
    hour_in = db.Column(db.String(5))
    hour_out = db.Column(db.String(5))

    status = db.Column(db.String(10), default="OPEN")  # OPEN / CLOSED
    closed_at = db.Column(db.DateTime)

    __table_args__ = (db.UniqueConstraint("day", "turn", name="uq_day_turn"),)

class CashExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey("shift.id"), nullable=False)
    category = db.Column(db.String(50), nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    note = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ShiftClose(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey("shift.id"), unique=True)

    withdrawn_cash = db.Column(db.Integer, nullable=False)  # regla import: retirado = sales_cash
    ending_calc = db.Column(db.Integer, nullable=False, default=0)
    ending_cash = db.Column(db.Integer, nullable=False)
    difference = db.Column(db.Integer, nullable=False)

    note = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    close_ok = db.Column(db.Integer, default=1)

    edited_by = db.Column(db.String(50))
    edited_at = db.Column(db.DateTime)
    edit_reason = db.Column(db.String(200))
    edit_count = db.Column(db.Integer, default=0)

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.Date, nullable=False)
    employee = db.Column(db.String(50), nullable=False)

    group_code = db.Column(db.String(1))       # A / B sugerido
    mode = db.Column(db.String(10), default="AUTO")  # AUTO / MANUAL

    morning_in = db.Column(db.String(5))
    morning_out = db.Column(db.String(5))
    afternoon_in = db.Column(db.String(5))
    afternoon_out = db.Column(db.String(5))

    novelty = db.Column(db.String(40))         # una sola columna
    novelty_minutes = db.Column(db.Integer)    # RP min; Enfermedad se guarda en min internos
    notes = db.Column(db.String(200))

    __table_args__ = (db.UniqueConstraint("day", "employee", name="uq_att_day_emp"),)

class AttendanceConsumption(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    attendance_id = db.Column(db.Integer, db.ForeignKey("attendance.id"), nullable=False)
    idx = db.Column(db.Integer, nullable=False)
    item = db.Column(db.String(120))
    amount = db.Column(db.Integer)
    __table_args__ = (db.UniqueConstraint("attendance_id", "idx", name="uq_attcons_att_idx"),)

class RotationConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    week0_map = db.Column(db.String(500))  # "Paula:A,Pato:B,..." (A/B)
    created_by = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Vacation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee = db.Column(db.String(50), nullable=False)
    start_day = db.Column(db.Date, nullable=False)
    end_day = db.Column(db.Date, nullable=False)
class CalendarDay(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.Date, unique=True, nullable=False)
    holiday_type = db.Column(db.String(20))  # "", "LABORABLE", "NO_LABORABLE"


class AdvanceRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    amount_requested = db.Column(db.Integer, nullable=False)
    requested_for_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.String(300))

    status = db.Column(db.String(12), default="PENDING")  # PENDING / APPROVED / REJECTED
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    decided_at = db.Column(db.DateTime)
    decided_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    admin_comment = db.Column(db.String(300))


# ==============================
# MIGRACIONES / SEED
# ==============================
def _table_cols(table: str):
    # PRAGMA solo existe en SQLite
    if not is_sqlite():
        return []
    with db.engine.begin() as conn:
        return [row[1] for row in conn.exec_driver_sql(f"PRAGMA table_info({table})").fetchall()]

def is_sqlite():
    uri = (app.config.get("SQLALCHEMY_DATABASE_URI") or "").lower()
    return uri.startswith("sqlite:")

def ensure_columns_shift():
    if is_sqlite():
        cols = _table_cols("shift")
        with db.engine.begin() as conn:
            if "sales_pya" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN sales_pya INTEGER DEFAULT 0")
            if "sales_rappi" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN sales_rappi INTEGER DEFAULT 0")
            if "sales_apps" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN sales_apps INTEGER DEFAULT 0")
            if "delivery_data_json" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN delivery_data_json TEXT")
            if "hour_shift" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN hour_shift TEXT")
            if "hour_in" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN hour_in TEXT")
            if "hour_out" not in cols:
                conn.exec_driver_sql("ALTER TABLE shift ADD COLUMN hour_out TEXT")
        return

    with db.engine.begin() as conn:
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS sales_pya INTEGER DEFAULT 0;')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS sales_rappi INTEGER DEFAULT 0;')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS sales_apps INTEGER DEFAULT 0;')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS delivery_data_json TEXT;')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS hour_shift VARCHAR(10);')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS hour_in VARCHAR(5);')
        conn.exec_driver_sql('ALTER TABLE shift ADD COLUMN IF NOT EXISTS hour_out VARCHAR(5);')

def ensure_columns_shift_close():
    cols = _table_cols("shift_close")
    with db.engine.connect() as conn:
        if "close_ok" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN close_ok INTEGER DEFAULT 1")
        if "edited_by" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN edited_by TEXT")
        if "edited_at" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN edited_at DATETIME")
        if "edit_reason" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN edit_reason TEXT")
        if "edit_count" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN edit_count INTEGER DEFAULT 0")
        if "ending_calc" not in cols:
            conn.exec_driver_sql("ALTER TABLE shift_close ADD COLUMN ending_calc INTEGER DEFAULT 0")
            
def ensure_columns_user_mobile():
    # SQLite
    if is_sqlite():
        cols = _table_cols("user")
        with db.engine.begin() as conn:
            if "mobile_pin_hash" not in cols:
                conn.exec_driver_sql("ALTER TABLE user ADD COLUMN mobile_pin_hash TEXT")
            if "mobile_pin_fingerprint" not in cols:
                conn.exec_driver_sql("ALTER TABLE user ADD COLUMN mobile_pin_fingerprint TEXT")
            if "mobile_pin_attempts" not in cols:
                conn.exec_driver_sql("ALTER TABLE user ADD COLUMN mobile_pin_attempts INTEGER DEFAULT 0")
            if "mobile_pin_locked_until" not in cols:
                conn.exec_driver_sql("ALTER TABLE user ADD COLUMN mobile_pin_locked_until DATETIME")
        return

    # Postgres (Render)
    with db.engine.begin() as conn:
        conn.exec_driver_sql('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS mobile_pin_hash VARCHAR(255);')
        conn.exec_driver_sql('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS mobile_pin_fingerprint VARCHAR(64);')
        conn.exec_driver_sql('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS mobile_pin_attempts INTEGER DEFAULT 0;')
        conn.exec_driver_sql('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS mobile_pin_locked_until TIMESTAMP;')
        conn.exec_driver_sql(
            'CREATE UNIQUE INDEX IF NOT EXISTS ux_user_mobile_pin_fingerprint '
            'ON "user"(mobile_pin_fingerprint) '
            'WHERE mobile_pin_fingerprint IS NOT NULL;'
        )

def ensure_columns_attendance():
    cols = _table_cols("attendance")
    with db.engine.connect() as conn:
        if "group_code" not in cols:
            conn.exec_driver_sql("ALTER TABLE attendance ADD COLUMN group_code TEXT")
        if "mode" not in cols:
            conn.exec_driver_sql("ALTER TABLE attendance ADD COLUMN mode TEXT DEFAULT 'AUTO'")
        if "novelty" not in cols:
            conn.exec_driver_sql("ALTER TABLE attendance ADD COLUMN novelty TEXT")
        if "novelty_minutes" not in cols:
            conn.exec_driver_sql("ALTER TABLE attendance ADD COLUMN novelty_minutes INTEGER")
        if "notes" not in cols:
            conn.exec_driver_sql("ALTER TABLE attendance ADD COLUMN notes TEXT")


def _is_postgres():
    uri = (app.config.get("SQLALCHEMY_DATABASE_URI") or "").lower()
    return uri.startswith("postgresql://")


def seed_users():
    """
    Seed idempotente (no rompe si corre mas de una vez).
    En Gunicorn con 1 worker no hay drama, pero esto lo deja robusto.
    """
    try:
        existing = {u[0] for u in User.query.with_entities(User.username).all()}
        for name in EMPLOYEES:
            if name not in existing:
                db.session.add(User(username=name, role="user"))
        for name in ADMINS:
            if name not in existing:
                db.session.add(User(username=name, role="admin"))
        db.session.commit()
        backup_caja_local_y_drive()
    except IntegrityError:
        db.session.rollback()

def init_db():
    with app.app_context():
        # Crear schema caja si no existe (para Postgres)
        if not is_sqlite():
            from sqlalchemy import text
            db.session.execute(text("CREATE SCHEMA IF NOT EXISTS caja"))
            db.session.commit()
    
        db.create_all()
        ensure_columns_user_mobile()
        ensure_columns_shift()
        if is_sqlite():
            ensure_columns_shift_close()
            ensure_columns_attendance()
        seed_users()
        
        backup_caja_local_y_drive()

# Ejecutar init al levantar (Render / gunicorn y local)
init_db()

# ==============================
# AUTH
# ==============================
def current_user():
    uid = session.get("user_id")
    return User.query.get(uid) if uid else None


def sync_advance_to_attendance(ar) -> None:
    """Al aprobar un adelanto, lo refleja como consumo en Asistencia (como se carga manualmente).
    - Dia: requested_for_date
    - Empleado: username del solicitante
    - Item: 'adelanto'
    - Amount: amount_requested
    """
    emp = (User.query.get(ar.user_id).username or "").strip()
    if not emp:
        return
    day_obj = ar.requested_for_date
    # Asegurar fila de asistencia
    row = Attendance.query.filter_by(day=day_obj, employee=emp).first()
    if not row:
        row = Attendance(day=day_obj, employee=emp, mode="AUTO")
        db.session.add(row)
        db.session.flush()
    # Buscar proximo idx disponible
    cons = AttendanceConsumption.query.filter_by(attendance_id=row.id).order_by(AttendanceConsumption.idx.asc()).all()
    used = {c.idx for c in cons}
    idx = 1
    while idx in used and idx <= MAX_CONSUMOS:
        idx += 1
    if idx > MAX_CONSUMOS:
        # si esta lleno, sobreescribimos el ultimo
        AttendanceConsumption.query.filter_by(attendance_id=row.id, idx=MAX_CONSUMOS).delete()
        idx = MAX_CONSUMOS
    db.session.add(AttendanceConsumption(attendance_id=row.id, idx=idx, item="adelanto", amount=int(ar.amount_requested or 0)))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u or u.role != "admin":
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


# ==============================
# PORA Mobile (PIN)
# ==============================

MOBILE_SESSION_KEY = "m_user_id"

def mobile_current_user():
    uid = session.get(MOBILE_SESSION_KEY)
    return User.query.get(uid) if uid else None

def mobile_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not mobile_current_user():
            return redirect(url_for("m_login"))
        return fn(*args, **kwargs)
    return wrapper

def _pin_fingerprint(pin: str) -> str:
    # Fingerprint deterministico (para busqueda y unicidad) usando SECRET_KEY como "pepper".
    # NO guarda el PIN en claro.
    secret = (app.config.get("SECRET_KEY") or "").encode("utf-8")
    p = (pin or "").strip().encode("utf-8")
    return hashlib.sha256(secret + b"::" + p).hexdigest()

def set_mobile_pin(user: User, pin: str):
    pin = (pin or "").strip()
    if not (pin.isdigit() and len(pin) == 4):
        raise ValueError("El PIN debe tener 4 digitos.")
    fp = _pin_fingerprint(pin)
    # Unicidad (no permitir duplicados)
    exists = User.query.filter(User.mobile_pin_fingerprint == fp, User.id != user.id).first()
    if exists:
        raise ValueError("PIN en uso. Proba otra combinacion.")
    user.mobile_pin_hash = generate_password_hash(pin)
    user.mobile_pin_fingerprint = fp
    user.mobile_pin_attempts = 0
    user.mobile_pin_locked_until = None

def check_mobile_pin(user: User, pin: str) -> bool:
    if not user or not user.mobile_pin_hash:
        return False
    return check_password_hash(user.mobile_pin_hash, (pin or "").strip())

def is_mobile_locked(user: User) -> bool:
    if not user:
        return False
    until = user.mobile_pin_locked_until
    return bool(until and until > datetime.utcnow())

# ==============================
# UTILS
# ==============================
def parse_day_param(d_str: str):
    if not d_str:
        dd = date.today()
        return dd, dd.isoformat()
    try:
        dd = date.fromisoformat(d_str)
        return dd, d_str
    except Exception:
        dd = date.today()
        return dd, dd.isoformat()

def parse_range_params(start_s: str, end_s: str):
    today = date.today()
    default_end = today
    default_start = today - timedelta(days=14)
    start = date.fromisoformat(start_s) if start_s else default_start
    end = date.fromisoformat(end_s) if end_s else default_end
    if start > end:
        start, end = end, start
    return start, end

def to_int(v):
    v = (v or "").strip()
    return int(v) if v != "" else 0

def safe_int(v):
    v = (v or "").strip()
    if v == "":
        return None
    try:
        return int(v)
    except:
        return None

def safe_float(v):
    v = (v or "").strip().replace(",", ".")
    if v == "":
        return None
    try:
        return float(v)
    except:
        return None

def fmt_money(n) -> str:
    if n is None:
        n = 0
    try:
        n = int(n)
    except:
        n = 0
    s = f"{n:,}".replace(",", ".")
    return f"$ {s}"

app.jinja_env.filters["money"] = fmt_money

def weekday_es(value) -> str:
    """Return weekday name in Spanish for a date/datetime/ISO-date string."""
    if value is None:
        return ""
    try:
        import datetime as _dt
        if isinstance(value, _dt.datetime):
            d = value.date()
        elif isinstance(value, _dt.date):
            d = value
        else:
            # accept 'YYYY-MM-DD' or 'YYYY-MM-DD ...'
            s = str(value).strip()
            s = s[:10]
            d = _dt.datetime.strptime(s, "%Y-%m-%d").date()
        names = ["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]
        return names[d.weekday()]
    except Exception:
        return ""
app.jinja_env.filters["weekday_es"] = weekday_es

def responsible_name(value) -> str:
    """Normalize responsible display; fallback to 'Bernardo' when missing or numeric."""
    s = "" if value is None else str(value).strip()
    if not s:
        return "Bernardo"
    # numeric-like?
    try:
        float(s.replace(",", "."))
        # if it parses to float and contains only digits/dot/comma, treat as invalid name
        if re.fullmatch(r"[0-9]+([\.,][0-9]+)?", s):
            return "Bernardo"
    except Exception:
        pass
    return s
app.jinja_env.filters["responsible_name"] = responsible_name


# Make helper available in Jinja templates
try:
    app.jinja_env.globals["responsible_name"] = responsible_name
except Exception:
    pass

def valid_time_str(t: str) -> bool:
    if not t:
        return True
    if len(t) != 5 or t[2] != ":":
        return False
    hh, mm = t.split(":")
    if not (hh.isdigit() and mm.isdigit()):
        return False
    h = int(hh); m = int(mm)
    return 0 <= h <= 23 and 0 <= m <= 59

def time_to_min(t: str) -> Optional[int]:
    if not t or not valid_time_str(t):
        return None
    hh, mm = map(int, t.split(":"))
    return hh * 60 + mm

def diff_minutes(t_in: str, t_out: str):
    mi = time_to_min(t_in)
    mo = time_to_min(t_out)
    if mi is None or mo is None:
        return None
    return mo - mi

DELIVERY_SHIFT_PRESETS = {
    "MORNING": {"hour_in": "08:55", "hour_out": "12:50"},
    "AFTERNOON": {"hour_in": "16:55", "hour_out": "20:50"},
}

def delivery_hours_decimal(hour_in: Optional[str], hour_out: Optional[str]) -> float:
    mins = diff_minutes(hour_in or "", hour_out or "")
    if mins is None or mins < 0:
        return 0.0
    return round(mins / 60.0, 2)


def build_delivery_payload(shift_row: Shift) -> dict:
    payload = {
        "rates": [1500, 2000, 2500, 3000, 2500],
        "qtys": [0, 0, 0, 0, 0],
        "consume_amount": 0,
        "consume_note": "",
        "hour_shift": shift_row.hour_shift or "MORNING",
        "hour_in": shift_row.hour_in or DELIVERY_SHIFT_PRESETS["MORNING"]["hour_in"],
        "hour_out": shift_row.hour_out or DELIVERY_SHIFT_PRESETS["MORNING"]["hour_out"],
    }
    if shift_row.delivery_data_json:
        try:
            raw = json.loads(shift_row.delivery_data_json)
            if isinstance(raw, dict):
                payload.update(raw)
        except Exception:
            pass

    hour_shift = (payload.get("hour_shift") or shift_row.hour_shift or "MORNING").strip().upper()
    if hour_shift not in DELIVERY_SHIFT_PRESETS:
        hour_shift = "MORNING"
    payload["hour_shift"] = hour_shift

    payload["hour_in"] = (payload.get("hour_in") or shift_row.hour_in or DELIVERY_SHIFT_PRESETS[hour_shift]["hour_in"])
    payload["hour_out"] = (payload.get("hour_out") or shift_row.hour_out or DELIVERY_SHIFT_PRESETS[hour_shift]["hour_out"])

    rates = payload.get("rates") if isinstance(payload.get("rates"), list) else None
    if not rates or len(rates) != 5:
        payload["rates"] = [1500, 2000, 2500, 3000, 2500]
    qtys = payload.get("qtys") if isinstance(payload.get("qtys"), list) else None
    if not qtys or len(qtys) != 5:
        payload["qtys"] = [0, 0, 0, 0, 0]

    try:
        payload["consume_amount"] = int(float(payload.get("consume_amount") or 0))
    except Exception:
        payload["consume_amount"] = 0
    payload["consume_note"] = str(payload.get("consume_note") or "")
    payload["qtys"][4] = delivery_hours_decimal(payload["hour_in"], payload["hour_out"])
    return payload

def fmt_minutes(m):
    if m is None:
        return "-"
    sign = "-" if m < 0 else ""
    m = abs(int(m))
    return f"{sign}{m//60:02d}:{m%60:02d}"

def emp_key(name: str) -> str:
    return (
        name.lower()
        .replace("a","a").replace("e","e").replace("i","i")
        .replace("o","o").replace("u","u")
        .replace(" ", "_")
    )
import re

def _parse_amount_to_int(s: str) -> int:
    # "$40.000" / "$40,000" / "$40000" -> 40000
    s = (s or "").strip()
    s = s.replace(".", "").replace(",", "")
    digits = re.findall(r"\d+", s)
    return int("".join(digits)) if digits else 0

def extract_consumptions_from_notes(notes: str, max_items: int = 5):
    """
    Detecta patrones tipo: 'lata $790', 'adelanto $40.000'
    Separadores permitidos entre items: '/', '|', ','
    Devuelve: [(item, amount_int), ...]
    """
    if not notes:
        return []
    txt = str(notes).strip()
    if "$" not in txt:
        return []

    parts = re.split(r"[\/\|,]+", txt)
    out = []
    for p in parts:
        p = p.strip()
        if "$" not in p:
            continue

        m = re.search(r"(?P<item>.*?)(\$)\s*(?P<amt>[\d\.\,]+)", p)
        if not m:
            continue

        item = (m.group("item") or "").strip() or "Consumo"
        amt = _parse_amount_to_int(m.group("amt"))
        if amt <= 0:
            continue

        out.append((item[:120], amt))
        if len(out) >= max_items:
            break

    return out
def extract_consumptions_and_clean_notes(notes: str, max_items: int = 5):
    """
    Devuelve (consumos, notes_limpias)
    - consumos: [(item, amount_int), ...]
    - notes_limpias: texto sin los $monto (mantiene lo descriptivo)
    """
    if notes is None:
        return [], None

    txt = str(notes).strip()
    if txt == "":
        return [], None

    cons = extract_consumptions_from_notes(txt, max_items=max_items)
    if not cons:
        return [], txt

    # limpiar: remover "$ 12.345" del texto pero dejar el item
    cleaned_parts = []
    parts = re.split(r"[\/\|,]+", txt)

    for p in parts:
        p = p.strip()
        if "$" in p:
            # borra el "$123" (con puntos/comas) y espacios alrededor
            p2 = re.sub(r"\$\s*[\d\.\,]+", "", p).strip()
            if p2:
                cleaned_parts.append(p2)
        else:
            if p:
                cleaned_parts.append(p)

    notes_clean = " / ".join(cleaned_parts).strip()
    return cons, (notes_clean if notes_clean else None)

# ==============================
# CAJA UTILS
# ==============================
def expenses_total(shift_id: int) -> int:
    return sum(e.amount for e in CashExpense.query.filter_by(shift_id=shift_id).all())


def cash_final_value(s: Shift) -> int:
    """Caja final (efectivo que queda en la caja) guardada en el cierre.
    - Para turnos OPEN (aun sin cierre), devuelve 0.
    """
    try:
        sid = getattr(s, "id", None)
        if not sid:
            return 0
        c = ShiftClose.query.filter_by(shift_id=sid).first()
        return int(c.ending_cash or 0) if c else 0
    except Exception:
        return 0

def cash_bruto(s: Shift, cash_final: Optional[int] = None) -> int:
    """Efectivo bruto = Retirado (efectivo total) + Caja final."""
    retirado = int(s.sales_cash or 0)
    if cash_final is None:
        cash_final = cash_final_value(s)
    return retirado + int(cash_final or 0)

def cash_neto(s: Shift) -> int:
    """Efectivo neto = Retirado - Caja inicial.
    Nota: puede ser negativo si Retirado < Caja inicial (ej: faltante / error de carga).
    """
    return int(s.sales_cash or 0) - int(s.opening_cash or 0)

def calc_ingreso_bruto(s: Shift, egresos: int, cash_final: Optional[int] = None) -> int:
    """Ingreso total (bruto) =
       (Efectivo bruto + MP + PedidosYa + Rappi) + Egresos
       donde Efectivo bruto = Retirado + Caja final.
    """
    return (
        cash_bruto(s, cash_final=cash_final) +
        int(s.sales_mp or 0) +
        int(getattr(s, "sales_pya", 0) or 0) +
        int(getattr(s, "sales_rappi", 0) or 0) +
        int(egresos or 0)
    )

def calc_ingreso_neto(s: Shift, egresos: Optional[int] = None, cash_final: Optional[int] = None) -> int:
    """Ingreso neto = Ingreso total (bruto) - Egresos total."""
    if egresos is None:
        try:
            egresos = expenses_total(int(s.id))
        except Exception:
            egresos = 0
    return calc_ingreso_bruto(s, int(egresos or 0), cash_final=cash_final) - int(egresos or 0)

def calc_ending_calc(cash_final: int, withdrawn: int) -> int:
    """Compatibilidad legacy.
    Antes: caja final teorica = efectivo bruto - retirado.
    Ahora: la Caja final se carga manualmente, por lo que la 'teorica' coincide con la real.
    """
    return int(cash_final or 0)


def prev_turn_of(day_obj: date, turn_code: str):
    if turn_code == "AFTERNOON":
        return day_obj, "MORNING"
    return day_obj - timedelta(days=1), "AFTERNOON"

def get_locked_opening_cash(day_obj: date, turn_code: str):
    pday, pturn = prev_turn_of(day_obj, turn_code)
    prev_shift = Shift.query.filter_by(day=pday, turn=pturn).first()
    if not prev_shift or prev_shift.status != "CLOSED":
        return None
    close_row = ShiftClose.query.filter_by(shift_id=prev_shift.id).first()
    if not close_row:
        return None

    ec = int(close_row.ending_calc or 0)
    if ec != 0:
        return ec
    er = int(close_row.ending_cash or 0)
    if er != 0:
        return er
    return 0

def can_edit_close(u: User, close_row: ShiftClose) -> bool:
    """Permite re-ediciones del cierre (admin y no-admin).

    Antes se limitaba a 1 edicion para no-admin. A pedido, se habilita
    edicion ilimitada y se conserva el registro de quien edito (edited_by/at).
    """
    if not u or not close_row:
        return False
    return True

# ==============================
# ASISTENCIA: GRUPOS / VACACIONES / CUPOS
# ==============================
def week_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def rotation_config_get() -> Optional[RotationConfig]:
    return RotationConfig.query.order_by(RotationConfig.id.desc()).first()

def parse_week0_map(s: str) -> dict:
    m = {}
    if not s:
        return m
    parts = [p.strip() for p in s.split(",") if p.strip()]
    for p in parts:
        if ":" in p:
            emp, g = p.split(":", 1)
            emp = emp.strip()
            g = g.strip().upper()
            if emp and g in ("A","B"):
                m[emp] = g
    return m

def make_week0_map_str(m: dict) -> str:
    out = []
    for emp in EMPLOYEES:
        g = (m.get(emp) or "A").upper()
        if g not in ("A","B"):
            g = "A"
        out.append(f"{emp}:{g}")
    return ",".join(out)

def group_for_employee_on_day(emp: str, d: date) -> str:
    cfg = rotation_config_get()
    if not cfg or not cfg.week0_map:
        return "A"
    week0 = parse_week0_map(cfg.week0_map)
    base = week0.get(emp, "A")
    base = base if base in ("A","B") else "A"
    monday = week_monday(d)
    monday0 = week_monday(cfg.created_at.date())
    weeks = (monday - monday0).days // 7
    if weeks % 2 == 0:
        return base
    return "B" if base == "A" else "A"

def expected_times_for_group(g: str, d: Optional[date]=None):
    # Monday-Friday: normal groups
    if d is not None and d.weekday() == 5:
        # Saturday: reduced schedule 05:05.
        # Group A works in the morning, Group B works in the afternoon.
        if g == "A":
            return {
                "morning_in": "07:55",
                "morning_out": "13:00",
                "afternoon_in": "",
                "afternoon_out": "",
            }
        return {
            "morning_in": "",
            "morning_out": "",
            "afternoon_in": "15:55",
            "afternoon_out": "21:00",
        }

    return GROUP_A if g == "A" else GROUP_B

def is_vacation(emp: str, d: date) -> bool:
    v = Vacation.query.filter(
        Vacation.employee == emp,
        Vacation.start_day <= d,
        Vacation.end_day >= d
    ).first()
    return bool(v)

def month_range(d: date) -> Tuple[date, date]:
    start = d.replace(day=1)
    if start.month == 12:
        end = date(start.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(start.year, start.month + 1, 1) - timedelta(days=1)
    return start, end
SATURDAY_MIN = 5 * 60 + 5  # 05:05 => 305 min

def jornada_min_for_day(d: date) -> int:
    # 0=lunes ... 5=sabado ... 6=domingo
    if d.weekday() == 5:
        return SATURDAY_MIN
    return JORNADA_MIN

def get_holiday_type(d: date) -> str:
    row = CalendarDay.query.filter_by(day=d).first()
    return (row.holiday_type or "") if row else ""

def set_holiday_type(d: date, htype: str):
    row = CalendarDay.query.filter_by(day=d).first()
    if not row:
        row = CalendarDay(day=d, holiday_type=htype or "")
        db.session.add(row)
    else:
        row.holiday_type = htype or ""

def monthly_used_minutes(emp: str, d: date, novelty_name: str) -> int:
    start, end = month_range(d)
    q = Attendance.query.filter(
        Attendance.employee == emp,
        Attendance.day >= start,
        Attendance.day <= end,
        Attendance.novelty == novelty_name
    ).all()
    return sum(int(a.novelty_minutes or 0) for a in q)
def compute_work_minutes_and_flags(a) -> dict:
    emp = a.employee
    d = a.day
    warnings = []

    g = a.group_code or group_for_employee_on_day(emp, d)
    exp = expected_times_for_group(g, d)  # OJO: tu funcion ya recibe (g,d)
    jornada_min = jornada_min_for_day(d)
    hday = (get_holiday_type(d) or "").upper().strip()  # "", "LABORABLE", "NO_LABORABLE"

    # Vacaciones (tabla Vacation o novedad)
    if a.novelty == "Vacaciones" or is_vacation(emp, d):
        return {
            "total_worked_min": jornada_min,
            "payable_min": jornada_min,
            "tardy_min": 0,
            "rp_unpaid_min": 0,
            "sick_unpaid_min": 0,
            "warnings": ["Vacaciones"],
            "group": g,
            "exp": exp,
        }

    # Inasistencia
    if a.novelty == "Inasistencia":
        return {
            "total_worked_min": 0,
            "payable_min": 0,
            "tardy_min": 0,
            "rp_unpaid_min": 0,
            "sick_unpaid_min": 0,
            "warnings": ["Inasistencia"],
            "group": g,
            "exp": exp,
        }

    # Delivery (pago aparte) - no computa horas
    if a.novelty == "Delivery":
        return {
            "total_worked_min": 0,
            "payable_min": 0,
            "tardy_min": 0,
            "rp_unpaid_min": 0,
            "sick_unpaid_min": 0,
            "warnings": ["Delivery (pago aparte)"],
            "group": g,
            "exp": exp,
        }

    # =========================
    # Calculo de minutos trabajados
    # =========================

    # Manana
    if a.morning_in or a.morning_out:
        mi = a.morning_in or exp.get("morning_in")
        mo = a.morning_out or exp.get("morning_out")
        mm = diff_minutes(mi, mo) if (mi and mo) else None
    else:
        mm = 0

    # Tarde
    if a.afternoon_in or a.afternoon_out:
        ai = a.afternoon_in or exp.get("afternoon_in")
        ao = a.afternoon_out or exp.get("afternoon_out")
        ma = diff_minutes(ai, ao) if (ai and ao) else None
    else:
        ma = 0

    total_worked = 0
    for v in (mm, ma):
        if v is None:
            continue
        if v > 0:
            total_worked += v

    # Si no cargo nada (ni manana ni tarde), asumimos jornada completa
    if not (a.morning_in or a.morning_out or a.afternoon_in or a.afternoon_out):
        total_worked = jornada_min

    # =========================
    # Tardanza (contra horario esperado)
    # =========================
    tardy_min = 0
    exp_in = time_to_min(exp.get("morning_in") or "")
    real_in = time_to_min(a.morning_in or (exp.get("morning_in") or ""))
    if exp_in is not None and real_in is not None:
        tardy_min = max(0, real_in - exp_in)

    payable = max(0, total_worked)

    # =========================
    # Reglas de cupos / descuentos
    # =========================
    rp_unpaid = 0
    sick_unpaid = 0

    if a.novelty == "Razones particulares":
        req = int(a.novelty_minutes or 0)
        used_excluding_today = max(0, monthly_used_minutes(emp, d, "Razones particulares") - req)
        remaining = max(0, RP_LIMIT_MIN_PER_MONTH - used_excluding_today)
        unpaid = max(0, req - remaining)
        rp_unpaid = unpaid
        if unpaid > 0:
            warnings.append(f"! Excede RP: {unpaid} min NO pagos")
        payable = max(0, payable - unpaid)

    if a.novelty == "Enfermedad":
        req = int(a.novelty_minutes or 0)
        used_excluding_today = max(0, monthly_used_minutes(emp, d, "Enfermedad") - req)
        remaining = max(0, SICK_LIMIT_MIN_PER_MONTH - used_excluding_today)
        unpaid = max(0, req - remaining)
        sick_unpaid = unpaid
        if unpaid > 0:
            warnings.append(f"! Excede Enfermedad: {unpaid} min NO pagos")
        payable = max(0, payable - unpaid)

    if a.novelty == "Curso":
        warnings.append("Curso")

    if a.novelty == "Tardanza" and tardy_min > 0:
        warnings.append(f"Tardanza: {tardy_min} min")

    # =========================
    # Validaciones "de gestion" (avisos)
    # =========================
    # Si faltan horas y no hay novedad -> warning
    if payable < jornada_min and not (a.novelty or "").strip():
        warnings.append("! Faltan horas: carga novedad")

    # Si sobran horas y no hay notas -> warning
    if payable > jornada_min and not (a.notes or "").strip():
        warnings.append("! Horas extra: justifica en notas")

    # Manana
    if a.morning_in or a.morning_out:
        mi = a.morning_in or (exp.get("morning_in") or "")
        mo = a.morning_out or (exp.get("morning_out") or "")
        mm = diff_minutes(mi, mo)
    else:
        mm = 0

    # Tarde
    if a.afternoon_in or a.afternoon_out:
        ai = a.afternoon_in or (exp.get("afternoon_in") or "")
        ao = a.afternoon_out or (exp.get("afternoon_out") or "")
        ma = diff_minutes(ai, ao)
    else:
        ma = 0

    # Sumar solo minutos positivos
    total_worked = 0
    for v in (mm, ma):
        if v is None:
            continue
        if v > 0:
            total_worked += v

    # Si no cargo nada en ningun bloque, asumimos jornada completa del dia
    if not (a.morning_in or a.morning_out or a.afternoon_in or a.afternoon_out):
        total_worked = jornada_min

    # Tardanza: contra el inicio esperado del bloque que corresponda
    exp_start = exp.get("morning_in") or exp.get("afternoon_in") or ""
    real_start = None
    if exp.get("morning_in"):
        real_start = a.morning_in or exp_start
    elif exp.get("afternoon_in"):
        real_start = a.afternoon_in or exp_start

    exp_in = time_to_min(exp_start)
    real_in = time_to_min(real_start or "")
    tardy_min = 0
    if exp_in is not None and real_in is not None:
        tardy_min = max(0, real_in - exp_in)

    payable = max(0, int(total_worked))

    rp_unpaid = 0
    sick_unpaid = 0

    if a.novelty == "Razones particulares":
        req = int(a.novelty_minutes or 0)
        used_excluding_today = max(0, monthly_used_minutes(emp, d, "Razones particulares") - req)
        remaining = max(0, RP_LIMIT_MIN_PER_MONTH - used_excluding_today)
        unpaid = max(0, req - remaining)
        rp_unpaid = unpaid
        if unpaid > 0:
            warnings.append(f"! Excede RP: {unpaid} min NO pagos")
        payable = max(0, payable - unpaid)

    if a.novelty == "Enfermedad":
        req = int(a.novelty_minutes or 0)
        used_excluding_today = max(0, monthly_used_minutes(emp, d, "Enfermedad") - req)
        remaining = max(0, SICK_LIMIT_MIN_PER_MONTH - used_excluding_today)
        unpaid = max(0, req - remaining)
        sick_unpaid = unpaid
        if unpaid > 0:
            warnings.append(f"! Excede Enfermedad: {unpaid} min NO pagos")
        payable = max(0, payable - unpaid)

    if a.novelty == "Curso":
        warnings.append("Curso")

    if a.novelty == "Tardanza" and tardy_min > 0:
        warnings.append(f"Tardanza: {tardy_min} min")

    # Reglas de feriado
    if hday == "LABORABLE":
        payable = int(payable) * 2
        warnings.append("Feriado laborable: paga x2")
    elif hday == "NO_LABORABLE":
        payable = int(jornada_min)
        warnings.append("Feriado no laborable: paga normal")

    # Reglas de control: menos horas => debe haber novedad
    if total_worked < jornada_min:
        if total_worked > 0 and not (a.novelty or "").strip():
            warnings.append("! Menos horas: carga una Novedad")

    # Mas horas => debe haber nota
    if total_worked > jornada_min:
        if not (a.notes or "").strip():
            warnings.append("! Horas extra: carga Nota")

    return {
        "total_worked_min": total_worked,
        "payable_min": payable,
        "tardy_min": tardy_min,
        "rp_unpaid_min": rp_unpaid,
        "sick_unpaid_min": sick_unpaid,
        "warnings": warnings,
        "group": g,
        "exp": exp,
    }

def consumptions_summary_for_attendance(att_id: int) -> Tuple[int, str]:
    cons = (
        AttendanceConsumption.query
        .filter_by(attendance_id=att_id)
        .order_by(AttendanceConsumption.idx.asc())
        .all()
    )
    total = sum(int(c.amount or 0) for c in cons)
    items = [c.item for c in cons if c.item]
    items_txt = ", ".join(items[:4])
    if len(items) > 4:
        items_txt += "..."
    return total, items_txt

# ==============================
# BASE CSS
# ==============================
BASE_CSS = """
<style>
  body{font-family:Arial, sans-serif; margin:24px auto; padding:0 12px; color:#111;}
  a{color:#5a2ca0; text-decoration:none;}
  a:hover{text-decoration:underline;}
  table{border-collapse:collapse; width:100%;}
  th,td{border:1px solid #ddd; padding:10px; font-size:14px; vertical-align:middle;}
  th{background:#f3f3f3; text-align:left;}

  /* Compact tables to show more rows (used in Resumen de cierres) */
  .tight th,.tight td{padding:6px 8px; font-size:13px;}
  .tight th{white-space:nowrap;}
  .nowrap{white-space:nowrap;}

  .muted{color:#666; font-size:12px;}
  .btn{display:inline-block; padding:8px 12px; border:1px solid #bbb; border-radius:10px; background:#f6f6f6; cursor:pointer; color:#111; text-decoration:none;}
  .btn:hover{background:#ececec;}
  .badge{padding:3px 10px; border-radius:999px; font-size:12px; display:inline-block; border:1px solid transparent;}
  .badge-ok{background:#c8f3d6; color:#0f5132; border-color:#85d9a1;}
  .badge-warn{background:#ffe2b8; color:#7a4b00; border-color:#ffbf66;}
  .badge-open{background:#ffe9a6; color:#5f4b00; border-color:#ffd255;}
  .badge-closed{background:#bfead3; color:#0f5132; border-color:#7ad2a4;}
  .box{border:1px solid #d0d0d0; border-radius:12px; padding:12px; background:#fafafa;}
  .row{display:flex; gap:14px; flex-wrap:wrap; align-items:center;}
  .top{display:flex; gap:16px; align-items:center; justify-content:space-between; flex-wrap:wrap;}
  .logo{display:flex; gap:12px; align-items:center;}
  .logo img{height:60px; border-radius:10px;}
  h1,h2,h3{margin:12px 0;}
  input,select,button,textarea{font-family:inherit;}
  .net-soft{background:#dff5e6;}
  .vac-row{background:#d9d9d9;}
  .warn-row{background:#ffd9c7;}
  .preset{color:#777 !important;}
  .preset::-webkit-datetime-edit{color:#777;}
  .preset:hover{color:#111 !important;}
  .preset:hover::-webkit-datetime-edit{color:#111;}
  .chip{display:inline-block; padding:2px 8px; border-radius:999px; font-size:12px; border:1px solid #bbb;}
  .chip-rp{background:#ffd7a6; border-color:#ffb866;}
  .chip-vac{background:#cfcfcf; border-color:#a9a9a9;}
  .chip-curso{background:#cfe7ff; border-color:#89bfff;}
</style>
"""

# ==============================
# AUTH TEMPLATES
# ==============================
LOGIN_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Login</title>
""" + BASE_CSS + """
<style>
  body{max-width:420px;}
  input,select,button{padding:10px;font-size:14px;width:100%;margin:6px 0}
  .card{border:1px solid #ddd;border-radius:12px;padding:16px;background:#fafafa}
  .err{color:#842029}
  .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body>
<div class="card">
  <h2 style="margin-top:0">Ingresar</h2>
  {% if err %}<p class="err"><b>{{err}}</b></p>{% endif %}
  <form method="post">
    <label>Usuario</label>
    <select name="username" required>
      {% for u in users %}
        <option value="{{u}}">{{u}}</option>
      {% endfor %}
    </select>
    <label>Contrasena</label>
    <input type="password" name="password" required>
    <button class="btn">Entrar</button>
  </form>
  <p class="muted">Primera vez: configura tu contrasena.</p>
  <p><a href="{{ url_for('setup_password') }}">Configurar contrasena</a></p>
</div>
</body>
</html>
"""

SETUP_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Configurar contrasena</title>
""" + BASE_CSS + """
<style>
  body{max-width:480px;}
  input,select,button{padding:10px;font-size:14px;width:100%;margin:6px 0}
  .card{border:1px solid #ddd;border-radius:12px;padding:16px;background:#fafafa}
  .err{color:#842029}
  .ok{color:#0f5132}
  .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body>
<div class="card">
  <h2 style="margin-top:0">Configurar contrasena</h2>
  {% if err %}<p class="err"><b>{{err}}</b></p>{% endif %}
  {% if msg %}<p class="ok"><b>{{msg}}</b></p>{% endif %}
  <form method="post">
    <label>Usuario</label>
    <select name="username" required>
      {% for u in users %}
        <option value="{{u}}">{{u}}</option>
      {% endfor %}
    </select>
    <label>Nueva contrasena</label>
    <input type="password" name="p1" required>
    <label>Confirmar contrasena</label>
    <input type="password" name="p2" required>
    <button class="btn">Guardar</button>
  </form>
  <p><a href="{{ url_for('login') }}">Volver a login</a></p>
</div>
</body>
</html>
"""

@app.route("/health")
def health():
    return "ok"

@app.route("/login", methods=["GET","POST"])
def login():
    users = [u.username for u in User.query.filter_by(is_active=1).order_by(User.username.asc()).all()]
    err = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        u = User.query.filter_by(username=username, is_active=1).first()
        if not u or not u.password_hash:
            err = "Usuario no configurado. Configura tu contrasena primero."
        elif not check_password_hash(u.password_hash, password):
            err = "Contrasena incorrecta."
        else:
            session["user_id"] = u.id
            nxt = request.args.get("next") or url_for("home")
            return redirect(nxt)
    return render_template_string(LOGIN_HTML, users=users, err=err)

@app.route("/setup", methods=["GET","POST"])
def setup_password():
    users = [u.username for u in User.query.filter_by(is_active=1).order_by(User.username.asc()).all()]
    err = None
    msg = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        p1 = request.form.get("p1") or ""
        p2 = request.form.get("p2") or ""
        if p1 != p2:
            err = "Las contrasenas no coinciden."
        elif len(p1) < 4:
            err = "Contrasena demasiado corta (minimo 4)."
        else:
            u = User.query.filter_by(username=username, is_active=1).first()
            if not u:
                err = "Usuario invalido."
            else:
                u.password_hash = generate_password_hash(p1)
                db.session.commit()
                backup_caja_local_y_drive()
                msg = "Contrasena guardada. Ya podes ingresar."
    return render_template_string(SETUP_HTML, users=users, err=err, msg=msg)

@app.route("/logout")
def logout():
    session.pop("user_id", None)
    return redirect(url_for("login"))

# ==============================
# PANEL
# ==============================
HOME_HTML = """

<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>PORA - Panel</title>
  """ + BASE_CSS + """
  <style>
    body{max-width:980px;}
    .brand{display:flex; gap:14px; align-items:center; flex-wrap:wrap;}
    .brand img{height:70px; border-radius:10px;}
    .cards{display:grid; grid-template-columns:repeat(3, minmax(220px, 1fr)); gap:14px; margin-top:18px;}
    @media (max-width: 820px){ .cards{grid-template-columns:1fr;} }
    a.card{border:1px solid #ddd; border-radius:14px; padding:16px; color:#111; background:#fafafa; display:block;}
    a.card:hover{background:#f0f0f0;}
    .t{font-size:18px; font-weight:bold; margin:0 0 6px;}
    .d{color:#555; margin:0; font-size:13px; line-height:1.35;}
    .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body>
  <div class="top">
    <div class="brand">
      <img src="{{ url_for('static', filename='img/pora_logo.png') }}" alt="PORA">
      <div>
        <h1 style="margin:0;">PORA - Panel</h1>
        <div class="muted">Usuario: <b>{{username}}</b> ({{role}})</div>
      </div>
    </div>
    <div><a class="btn" href="{{ url_for('logout') }}">Salir</a></div>
  </div>

  <div class="cards">
    <a class="card" href="{{ url_for('caja_index') }}">
      <p class="t">Control de Caja</p>
      <p class="d">Turnos, cierres, edicion, export e import.</p>
    </a>

    <a class="card" href="{{ url_for('asistencia') }}">
      <p class="t">Asistencia</p>
      <p class="d">Carga diaria + resumen + export + import.</p>
    </a>

    <a class="card" href="{{ url_for('stock') }}">
      <p class="t">Control de Stock</p>
      <p class="d">En construccion.</p>
    </a>
  
    <a class="card" href="{{ url_for('m_login') }}">
      <p class="t">PORA Mobile</p>
      <p class="d">Portal movil (PIN): adelantos, asistencia (futuro), stock (futuro).</p>
    </a>

    {% if role == 'admin' %}
    <a class="card" href="{{ url_for('admin_pin_mobile') }}">
      <p class="t">PIN Mobile</p>
      <p class="d">Asignar / cambiar PIN de acceso para el portal movil.</p>
    </a>
    {% endif %}

  </div>
</body>
</html>
"""

@app.route("/")
@login_required
def home():
    u = current_user()
    return render_template_string(HOME_HTML, username=u.username, role=u.role)

@app.route("/stock")
@login_required
def stock():
    return "<h2>Stock (en construccion)</h2><p><a href='/'>Volver</a></p>"


ADMIN_PIN_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>PIN Mobile</title>{{ base_css|safe }}</head>
<body style="max-width:900px;">
  <div class="top">
    <div>
      <h2 style="margin:0;">PIN Mobile (PORA Mobile)</h2>
      <div class="muted"><a href="{{ url_for('home') }}"><- Volver al menu</a></div>
    </div>
  </div>

  <div class="box" style="margin-top:10px;">
    <b>Reglas</b>
    <ul class="muted">
      <li>PIN de 4 digitos (solo numeros).</li>
      <li>No se permiten PIN duplicados.</li>
      <li>Si cambias un PIN, el anterior queda libre automaticamente.</li>
    </ul>
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <table style="margin-top:12px;">
    <tr><th>Usuario</th><th>Rol</th><th>PIN actual</th><th>Asignar / Cambiar PIN</th></tr>
    {% for u in users %}
      <tr>
        <td><b>{{u.username}}</b></td>
        <td>{{u.role}}</td>
        <td>{{ "Si" if u.mobile_pin_hash else "-" }}</td>
        <td>
          <form method="post" style="display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
            <input type="hidden" name="uid" value="{{u.id}}">
            <input type="text" name="pin" inputmode="numeric" pattern="\\d{4}" maxlength="4" placeholder="4 digitos" required style="width:120px; padding:8px;">
            <button class="btn">Guardar</button>
          </form>
        </td>
      </tr>
    {% endfor %}
  </table>
</body>
</html>
"""


@app.route("/admin/pin_mobile", methods=["GET","POST"])
@login_required
@admin_required
def admin_pin_mobile():
    err = None
    msg = None
    if request.method == "POST":
        try:
            uid = int(request.form.get("uid") or 0)
            pin = (request.form.get("pin") or "").strip()
            u = User.query.get(uid)
            if not u:
                raise ValueError("Usuario invalido.")
            set_mobile_pin(u, pin)
            db.session.commit()
            backup_caja_local_y_drive()
            msg = f"PIN actualizado para {u.username}."
        except Exception as ex:
            db.session.rollback()
            err = str(ex)

    users = User.query.filter_by(is_active=1).order_by(User.role.desc(), User.username.asc()).all()
    return render_template_string(ADMIN_PIN_HTML, base_css=BASE_CSS, users=users, err=err, msg=msg)


# ==============================
# PORA Mobile - UI (fase 1)
# ==============================

MOBILE_BASE_CSS = """
<style>
  body{font-family:Arial, sans-serif; margin:0; padding:0; background:#0f0f14; color:#fff; overflow-x:hidden;}
  .wrap{max-width:560px; margin:0 auto; padding:18px 14px 28px; box-sizing:border-box;}
  .card{background:#171722; border:1px solid rgba(255,255,255,0.08); border-radius:18px; padding:16px; box-shadow: 0 10px 30px rgba(0,0,0,0.25); max-width: 100%; box-sizing: border-box;overflow-x: hidden;}
  .title{font-size:26px; font-weight:800; margin:0 0 6px;}
  .sub{color:rgba(255,255,255,0.7); font-size:13px; margin:0 0 14px;}
  .err{background:rgba(255,90,90,0.18); border:1px solid rgba(255,90,90,0.35); padding:10px 12px; border-radius:14px; margin:10px 0;}
  .ok{background:rgba(90,255,140,0.16); border:1px solid rgba(90,255,140,0.35); padding:10px 12px; border-radius:14px; margin:10px 0;}
  .btn{display:block; width:100%; padding:14px 14px; border-radius:16px; border:1px solid rgba(255,255,255,0.12); background:#2a2a3a; color:#fff; font-weight:700; font-size:16px; cursor:pointer; text-align:center; text-decoration:none;}
  .btn:active{transform:scale(0.99);}
  .btn2{display:block; width:100%; padding:12px 14px; border-radius:16px; border:1px solid rgba(255,255,255,0.12); background:transparent; color:rgba(255,255,255,0.85); font-weight:700; font-size:14px; cursor:pointer; text-align:center; text-decoration:none;}
  input,textarea{width:100%; padding:14px 14px; border-radius:16px; border:1px solid rgba(255,255,255,0.14); background:#101018; color:#fff; font-size:16px; box-sizing:border-box;}
  textarea{min-height:88px; resize:vertical;}
  label{display:block; font-size:12px; color:rgba(255,255,255,0.7); margin:12px 0 6px;}
  .grid{display:grid; gap:10px;}
  .muted{color:rgba(255,255,255,0.65); font-size:12px;}
  .row{display:flex; gap:10px;}
  .pill{display:inline-block; padding:6px 10px; border-radius:999px; background:rgba(255,255,255,0.08); border:1px solid rgba(255,255,255,0.10); font-size:12px;}
  table{width:100%; border-collapse:collapse; margin-top:10px;}
  th,td{border-bottom:1px solid rgba(255,255,255,0.10); padding:10px 6px; font-size:13px; vertical-align:top;}
  th{color:rgba(255,255,255,0.70); font-weight:700; text-align:left;}
  .actions{display:flex; gap:8px; flex-wrap:wrap;}
  .sbtn{padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.12); background:#2a2a3a; color:#fff; font-weight:700; cursor:pointer; font-size:12px;}
  .sbtn2{padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.12); background:transparent; color:rgba(255,255,255,0.85); font-weight:700; cursor:pointer; font-size:12px;}

  .decide-form{display:flex; flex-direction:column; gap:10px;}
  .decide-form .comment{width:100%; box-sizing:border-box; padding:10px 12px; border-radius:14px; border:1px solid rgba(255,255,255,0.12); background:rgba(255,255,255,0.06); color:#fff; font-size:14px;}
  .decide-form .actions{display:flex; gap:10px; align-items:center; justify-content:flex-start; flex-wrap:wrap;}
</style>
"""

M_LOGIN_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>PORA Mobile</title>
""" + MOBILE_BASE_CSS + """
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="title">PORA Mobile</div>
      <div class="sub">Ingresa tu PIN para continuar</div>

      {% if err %}<div class="err"><b>{{err}}</b></div>{% endif %}
      {% if msg %}<div class="ok"><b>{{msg}}</b></div>{% endif %}

      <form method="post" class="grid" autocomplete="off">
        <label>PIN (4 digitos)</label>
        <input name="pin" inputmode="numeric" pattern="\\d{4}" maxlength="4" placeholder="****" required autofocus>
        <button class="btn" type="submit">Ingresar</button>
      </form>
      <div class="muted" style="margin-top:10px;">
        Si no tenes PIN, pediselo a un administrador.
      </div>
    </div>
  </div>
</body>
</html>
"""

M_MENU_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>PORA Mobile</title>
""" + MOBILE_BASE_CSS + """
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="title">PORA Mobile</div>
      <div class="sub">Hola, <b>{{u.username}}</b> <span class="pill">{{u.role}}</span></div>

      <div class="grid" style="margin-top:10px;">
        <a class="btn" href="{{ url_for('m_adv_new') }}">$ Solicitar adelanto</a>
        <a class="btn2" href="{{ url_for(\'m_adv_history\') }}">[Historial] Historial de adelantos</a>

        {% if u.role == 'admin' %}
          <a class="btn2" href="{{ url_for('m_adv_admin') }}">[Admin] Aprobar / Rechazar adelantos</a>
        {% endif %}

        <form method="post" action="{{ url_for('m_logout') }}">
          <button class="btn2" type="submit">Cerrar sesion</button>
        </form>
      </div>

      <div class="muted" style="margin-top:12px;">
        (Mas adelante: asistencia, stock, etc.)
      </div>
    </div>
  </div>
</body>
</html>
"""

M_ADV_NEW_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Adelanto</title>
""" + MOBILE_BASE_CSS + """
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="title">Adelanto</div>
      <div class="sub">Solicitud de adelanto</div>

      {% if err %}<div class="err"><b>{{err}}</b></div>{% endif %}
      {% if msg %}<div class="ok"><b>{{msg}}</b></div>{% endif %}

      <form method="post" class="grid">
        <label>Monto</label>
        <input name="amount" inputmode="numeric" placeholder="Ej: 20000" required>

        <label>Fecha de adelanto</label>
        <input name="req_date" type="date" required value="{{default_date}}">

        <label>Motivo (opcional)</label>
        <textarea name="reason" placeholder="Opcional..."></textarea>

        <button class="btn" type="submit">Enviar solicitud</button>
        <a class="btn2" href="{{ url_for('m_menu') }}">Volver</a>
      </form>
    </div>
  </div>
</body>
</html>
"""

M_ADV_ADMIN_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Adelantos</title>
""" + MOBILE_BASE_CSS + """
<style>
  /* Lista mobile mas estable que una tabla (evita desalineados) */
  .advlist{display:flex; flex-direction:column; gap:12px; margin-top:10px;}
  .advitem{border:1px solid rgba(255,255,255,0.10); background:rgba(255,255,255,0.04); border-radius:16px; padding:12px;}
  .advgrid{display:grid; grid-template-columns: 1fr 1fr; gap:10px 12px; align-items:start;}
  .advgrid .lbl{font-size:12px; opacity:.75; margin-bottom:2px;}
  .advgrid .val{font-size:14px;}
  .advgrid .span2{grid-column:1 / -1;}
  .decide-form{display:flex; flex-direction:column; gap:10px; margin-top:10px;}
  .decide-form .comment{width:100%; box-sizing:border-box; padding:10px 12px; border-radius:14px; border:1px solid rgba(255,255,255,0.12); background:rgba(255,255,255,0.06); color:#fff; font-size:14px;}
  .decide-form .actions{display:flex; gap:10px; flex-wrap:wrap;}
</style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="title">Adelantos</div>
      <div class="sub">Pendientes de decision</div>

      {% if err %}<div class="err"><b>{{err}}</b></div>{% endif %}
      {% if msg %}<div class="ok"><b>{{msg}}</b></div>{% endif %}

      {% if not rows %}
        <div class="muted" style="margin-top:12px;">No hay adelantos pendientes.</div>
      {% else %}
        <div class="advlist">
          {% for a,u in rows %}
            <div class="advitem">
              <div class="advgrid">
                <div>
                  <div class="lbl">Empleado</div>
                  <div class="val"><b>{{u.username}}</b></div>
                </div>
                <div>
                  <div class="lbl">Monto</div>
                  <div class="val">{{ a.amount_requested|money }}</div>
                </div>

                <div>
                  <div class="lbl">Dia</div>
                  <div class="val">{{ a.requested_for_date|weekday_es }}</div>
                </div>
                <div>
                  <div class="lbl">Fecha</div>
                  <div class="val">{{ a.requested_for_date }}</div>
                </div>

                <div class="span2">
                  <div class="lbl">Motivo</div>
                  <div class="val muted">{{ a.reason or "-" }}</div>
                </div>
              </div>

              <form method="post" action="{{ url_for('m_adv_decide', adv_id=a.id) }}" class="decide-form">
                <input type="text" name="admin_comment" placeholder="Comentario (opcional)" class="comment">
                <div class="actions">
                  <button class="sbtn" type="submit" name="decision" value="APPROVE">Aprobar</button>
                  <button class="sbtn2" type="submit" name="decision" value="REJECT">Rechazar</button>
                </div>
              </form>
            </div>
          {% endfor %}
        </div>
      {% endif %}

      <div style="margin-top:12px;">
        <a class="btn2" href="{{ url_for('m_menu') }}">Volver</a>
      </div>
    </div>
  </div>
</body>
</html>
"""


M_ADV_HISTORY_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Historial</title>
""" + MOBILE_BASE_CSS + """
<style>
  .hist{margin-top:10px;}
  .hist table{width:100%; border-collapse:collapse;}
  .hist th,.hist td{padding:10px 6px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align:top;}
  .pill2{display:inline-block; padding:4px 10px; border-radius:999px; border:1px solid rgba(255,255,255,0.14); font-size:12px;}
  .st-ok{background:rgba(90,255,140,0.16); border-color:rgba(90,255,140,0.35);}
  .st-no{background:rgba(255,90,90,0.16); border-color:rgba(255,90,90,0.35);}
  .st-pe{background:rgba(255,210,90,0.14); border-color:rgba(255,210,90,0.30);}
  .nowrap{white-space:nowrap;}
</style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="title">Historial</div>
      <div class="sub">Adelantos registrados</div>

      <div class="hist">
        <table>
          <tr>
            {% if is_admin %}<th class="nowrap">Empleado</th>{% endif %}
            <th class="nowrap">Dia</th>
            <th class="nowrap">Fecha</th>
            <th class="nowrap">Monto</th>
            <th>Estado</th>
            <th>Comentario</th>
          </tr>
          {% for r in rows %}
            <tr>
              {% if is_admin %}<td class="nowrap">{{ r.employee }}</td>{% endif %}
              <td class="nowrap">{{ r.weekday }}</td>
              <td class="nowrap">{{ r.req_date }}</td>
              <td class="nowrap">{{ r.amount|money }}</td>
              <td>
                {% if r.status == 'APPROVED' %}
                  <span class="pill2 st-ok">Aceptado</span>
                {% elif r.status == 'REJECTED' %}
                  <span class="pill2 st-no">Rechazado</span>
                {% else %}
                  <span class="pill2 st-pe">Pendiente</span>
                {% endif %}
              </td>
              <td>{{ r.comment or '-' }}</td>
            </tr>
          {% endfor %}
          {% if not rows %}
            <tr><td colspan="{{ 6 if is_admin else 5 }}" class="muted">Sin adelantos.</td></tr>
          {% endif %}
        </table>
      </div>

      <div style="margin-top:12px;">
        <a class="btn2" href="{{ url_for('m_menu') }}">Volver</a>
      </div>
    </div>
  </div>
</body>
</html>
"""



def _mobile_fail_state():
    # Estado de bloqueo simple para intentos sin usuario (sesion anonima)
    until_iso = session.get("m_lock_until")
    if until_iso:
        try:
            until = datetime.fromisoformat(until_iso)
            if until > datetime.utcnow():
                return True, until
        except Exception:
            pass
    return False, None

def _mobile_set_lock(minutes=10):
    until = datetime.utcnow() + timedelta(minutes=minutes)
    session["m_lock_until"] = until.isoformat()
    return until


@app.route("/admin/fix_responsables")
@login_required
@admin_required
def admin_fix_responsables():
    """Corrige turnos donde 'responsible' quedo con un numero (por error de carga/import).
    Heuristica: si responsible es numerico y opening_cash == 0, asumimos que ese numero era la caja inicial.
    Luego seteamos responsible con el primer admin configurado.
    """
    fixed = 0
    default_admin = ADMINS[0] if ADMINS else "Admin"
    for s in Shift.query.all():
        resp = (s.responsible or "").strip()
        if not resp:
            continue
        if resp in EMPLOYEES or resp in ADMINS:
            continue
        try:
            val = float(resp)
        except Exception:
            continue
        if int(s.opening_cash or 0) == 0 and val >= 0:
            s.opening_cash = int(round(val))
            s.responsible = default_admin
            fixed += 1
    if fixed:
        db.session.commit()
    backup_caja_local_y_drive()
    return render_template_string(
        BASE_MSG_HTML,
        base_css=BASE_CSS,
        title="Fix responsables",
        msg=f"Listo. Registros corregidos: {fixed}.",
        back=url_for("caja_index")
    )


@app.route("/m", methods=["GET","POST"])
def m_login():
    # Si ya esta logueado en mobile
    if mobile_current_user():
        return redirect(url_for("m_menu"))

    locked, until = _mobile_fail_state()
    if locked:
        mins = int((until - datetime.utcnow()).total_seconds() // 60) + 1
        return render_template_string(M_LOGIN_HTML, err=f"Demasiados intentos. Espera {mins} min.", msg=None)

    err = None
    msg = None
    if request.method == "POST":
        pin = (request.form.get("pin") or "").strip()

        # Validacion simple
        if not (pin.isdigit() and len(pin) == 4):
            err = "PIN invalido."
        else:
            fp = _pin_fingerprint(pin)
            u = User.query.filter_by(mobile_pin_fingerprint=fp, is_active=1).first()

            # Si no hay usuario con ese PIN (o no tiene PIN configurado)
            if not u or not u.mobile_pin_hash:
                fails = int(session.get("m_fail", 0)) + 1
                session["m_fail"] = fails
                if fails >= 5:
                    until = _mobile_set_lock(10)
                    mins = int((until - datetime.utcnow()).total_seconds() // 60) + 1
                    err = f"Demasiados intentos. Espera {mins} min."
                else:
                    err = "PIN incorrecto."
            else:
                if is_mobile_locked(u):
                    mins = int((u.mobile_pin_locked_until - datetime.utcnow()).total_seconds() // 60) + 1
                    err = f"PIN bloqueado. Espera {mins} min."
                elif not check_mobile_pin(u, pin):
                    u.mobile_pin_attempts = int(u.mobile_pin_attempts or 0) + 1
                    if u.mobile_pin_attempts >= 5:
                        u.mobile_pin_locked_until = datetime.utcnow() + timedelta(minutes=10)
                        u.mobile_pin_attempts = 0
                        db.session.commit()
                        backup_caja_local_y_drive()
                        err = "Demasiados intentos. Bloqueado 10 min."
                    else:
                        db.session.commit()
                        backup_caja_local_y_drive()
                        err = "PIN incorrecto."
                else:
                    # OK
                    u.mobile_pin_attempts = 0
                    u.mobile_pin_locked_until = None
                    db.session.commit()
                    backup_caja_local_y_drive()

                    session[MOBILE_SESSION_KEY] = u.id
                    session.pop("m_fail", None)
                    session.pop("m_lock_until", None)
                    return redirect(url_for("m_menu"))

    return render_template_string(M_LOGIN_HTML, err=err, msg=msg)

@app.route("/m/menu")
@mobile_login_required
def m_menu():
    return render_template_string(M_MENU_HTML, u=mobile_current_user())

@app.route("/m/logout", methods=["POST"])
def m_logout():
    session.pop(MOBILE_SESSION_KEY, None)
    return redirect(url_for("m_login"))

@app.route("/m/adelantos/nuevo", methods=["GET","POST"])
@mobile_login_required
def m_adv_new():
    u = mobile_current_user()
    err = None
    msg = None
    default_date = date.today().isoformat()

    if request.method == "POST":
        try:
            amount = safe_int(request.form.get("amount"))
            req_date_s = (request.form.get("req_date") or "").strip()
            reason = (request.form.get("reason") or "").strip() or None

            if amount is None or amount <= 0:
                raise ValueError("Monto invalido.")
            if not req_date_s:
                raise ValueError("Fecha invalida.")
            req_date = date.fromisoformat(req_date_s)

            ar = AdvanceRequest(
                user_id=u.id,
                amount_requested=int(amount),
                requested_for_date=req_date,
                reason=reason,
                status="PENDING",
                created_at=datetime.utcnow()
            )
            db.session.add(ar)
            db.session.commit()
            backup_caja_local_y_drive()
            msg = "Solicitud enviada "
        except Exception as ex:
            db.session.rollback()
            err = str(ex)

    return render_template_string(M_ADV_NEW_HTML, err=err, msg=msg, default_date=default_date)

@app.route("/m/admin/adelantos")
@mobile_login_required
def m_adv_admin():
    u = mobile_current_user()
    if not u or u.role != "admin":
        abort(403)

    rows = (
        db.session.query(AdvanceRequest, User)
        .join(User, User.id == AdvanceRequest.user_id)
        .filter(AdvanceRequest.status == "PENDING")
        .order_by(AdvanceRequest.created_at.asc())
        .all()
    )
    return render_template_string(M_ADV_ADMIN_HTML, rows=rows, err=None, msg=None)

@app.route("/m/admin/adelantos/<int:adv_id>/decide", methods=["POST"])
@mobile_login_required
def m_adv_decide(adv_id: int):
    u = mobile_current_user()
    if not u or u.role != "admin":
        abort(403)

    decision = (request.form.get("decision") or "").strip().upper()
    admin_comment = (request.form.get("admin_comment") or "").strip() or None
    if decision not in ("APPROVE", "REJECT"):
        return redirect(url_for("m_adv_admin"))

    ar = AdvanceRequest.query.get_or_404(adv_id)
    if ar.status != "PENDING":
        return redirect(url_for("m_adv_admin"))

    ar.status = "APPROVED" if decision == "APPROVE" else "REJECTED"
    ar.decided_at = datetime.utcnow()
    ar.decided_by_user_id = u.id
    ar.admin_comment = admin_comment
    if ar.status == "APPROVED":
        sync_advance_to_attendance(ar)
    db.session.commit()
    backup_caja_local_y_drive()
    return redirect(url_for("m_adv_admin"))



@app.route("/m/adelantos/historial")
@mobile_login_required
def m_adv_history():
    u = mobile_current_user()
    if not u:
        abort(403)

    q = (
        db.session.query(AdvanceRequest, User)
        .join(User, User.id == AdvanceRequest.user_id)
    )
    # Admin ve todos; user ve solo los suyos
    if u.role != "admin":
        q = q.filter(AdvanceRequest.user_id == u.id)

    q = q.order_by(AdvanceRequest.created_at.desc()).limit(200)
    rows_db = q.all()

    rows = []
    for ar, usr in rows_db:
        # Dia/fecha = requested_for_date (la fecha deseada)
        req_d = ar.requested_for_date or ar.created_at.date()
        rows.append({
            "employee": (getattr(usr, "username", None) or getattr(usr, "name", None) or ""),
            "weekday": weekday_es(req_d),
            "req_date": req_d.isoformat(),
            "amount": int(ar.amount_requested or 0),
            "status": ar.status or "PENDING",
            "comment": (ar.admin_comment or None),
        })

    return render_template_string(M_ADV_HISTORY_HTML, rows=rows, is_admin=(u.role=="admin"))


# ============================================================
# =====================  CAJA (UI)  ===========================
# ============================================================

def get_caja_summary(limit=200, start: Optional[date]=None, end: Optional[date]=None,
                     turn: str="ALL", responsible: str="ALL"):
    rows = []
    q = (
        db.session.query(Shift, ShiftClose)
        .join(ShiftClose, ShiftClose.shift_id == Shift.id)
        .filter(Shift.status == "CLOSED")
    )
    if start:
        q = q.filter(Shift.day >= start)
    if end:
        q = q.filter(Shift.day <= end)
    if turn and turn != "ALL":
        q = q.filter(Shift.turn == turn)
    if responsible and responsible != "ALL":
        q = q.filter(Shift.responsible == responsible)

    q = q.order_by(Shift.day.desc(), Shift.turn.asc()).limit(limit).all()

    for s, c in q:
        exp = expenses_total(s.id)
        cash_final = int(c.ending_cash or 0)
        bruto = calc_ingreso_bruto(s, exp, cash_final=cash_final)
        neto = calc_ingreso_neto(s, egresos=exp, cash_final=cash_final)

        rows.append({
            "day": s.day.isoformat(),
            "weekday": weekday_es(s.day),
            "turn_code": s.turn,
            "turn_name": TURN_NAMES.get(s.turn, s.turn),
            "responsible": responsible_name(s.responsible),
            "opening_cash": int(s.opening_cash or 0),
            "retirado": int(s.sales_cash or 0),
            "cash_final": cash_final,
            "sales_cash": int(cash_bruto(s, cash_final=cash_final)),
            "sales_mp": int(s.sales_mp or 0),
            "sales_pya": int(getattr(s, "sales_pya", 0) or 0),
            "sales_rappi": int(getattr(s, "sales_rappi", 0) or 0),
            "ventas_bruto": int(bruto),
            "ventas_neto": int(neto),
            "expenses": int(exp),
            "withdrawn": int(s.sales_cash or 0),
            "ending_real": int(c.ending_cash or 0),
            "difference": int(c.difference or 0),
            "close_ok": int(c.close_ok or 1),
        })
    return rows

CAJA_INDEX_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Control de Caja - PORA</title>
  {{ base_css|safe }}
  <style>
    body{max-width:1300px;}
    .filters{display:flex; gap:10px; flex-wrap:wrap; align-items:end; margin:8px 0 14px;}
    .filters label{display:block; font-size:12px; color:#666;}
    .smallnote{font-size:12px; color:#666; margin-top:6px;}
    .holiday-bg{background:#ffe3ec;}
    .holiday-bg table tr{background:#fff;}
    .summary-table{table-layout:auto;}
    .summary-table th,
    .summary-table td{padding:6px 7px; font-size:12px;}
    .summary-table th{
      white-space:normal;
      line-height:1.1;
      text-align:center;
      vertical-align:middle;
    }
    .summary-table td{
      white-space:nowrap;
      vertical-align:middle;
    }
    .summary-table td.wrapday{white-space:normal;}
  </style>
</head>
<body class="{{ 'holiday-bg' if hday else '' }}">

<div class="top">
  <div class="logo">
    <img src="{{ url_for('static', filename='img/pora_logo.png') }}" alt="PORA">
    <div>
      <h1 style="margin:0;">Control de Caja</h1>
      <div class="muted"><a href="{{ url_for('home') }}"><- Volver al menu</a></div>

      <form method="get" action="{{ url_for('caja_index') }}" style="margin-top:6px; display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
        <label class="muted">Fecha:</label>
        <input type="date" name="d" value="{{d}}">
        <button class="btn" type="submit">Ir</button>
      </form>
    </div>
  </div>
  <div class="row">
    <a class="btn" href="{{ url_for('export_excel') }}">Exportar Excel</a>
    <a class="btn" href="{{ url_for('export_caja_json') }}">Exportar JSON</a>
    <a class="btn" href="{{ url_for('import_caja') }}">Importar Excel</a>
    <a class="btn" href="{{ url_for('import_caja_json') }}">Importar JSON</a>
  </div>
</div>

<h2>Turnos del dia</h2>
<table class="tight">
  <tr>
    <th>Dia</th>
    <th>Turno</th>
    <th>Responsable</th>
    <th>Estado</th>
    <th>Cierre</th>
    <th>Ventas netas</th>
    <th>Efectivo disponible</th>
    <th>Accion</th>
  </tr>

  {% for code,name in turns %}
    {% set s = shifts.get(code) %}
    <tr>
      <td>{{ weekday_name }}</td>
      <td><b>{{name}}</b></td>
      <td>{{ responsible_name(s.responsible) if s else "-" }}</td>
      <td>
        {% if not s %}
          <span class="badge badge-open">No abierto</span>
        {% elif s.status == "OPEN" %}
          <span class="badge badge-open">Abierto</span>
        {% else %}
          <span class="badge badge-closed">Cerrado</span>
        {% endif %}
      </td>
      <td>
        {% if s and s.status == "CLOSED" %}
          {% set c = closes.get(s.id) %}
          {% if c and c.close_ok == 1 %}
            <span class="badge badge-ok">OK</span>
          {% else %}
            <span class="badge badge-warn">NO OK</span>
          {% endif %}
        {% else %}
          -
        {% endif %}
      </td>

      <td>
        {% if s and s.status == "CLOSED" %}
          {{ ventas_netas_map.get(s.id, 0)|money }}
        {% else %}-{% endif %}
      </td>

      <td>
        {% if s and s.status == "CLOSED" %}
          {{ efectivo_disp_map.get(s.id, 0)|money }}
        {% else %}-{% endif %}
      </td>

      <td>
        {% if not s %}
          <a href="{{ url_for('open_shift', turn=code, d=d) }}">Abrir</a>
        {% else %}
          {% if s.status == "OPEN" %}
            <a href="{{ url_for('shift', id=s.id, d=d) }}">Entrar</a>
          {% else %}
            <a href="{{ url_for('shift', id=s.id, d=d) }}">Ver</a>
            {% if can_edit_map.get(s.id) %}
              | <a href="{{ url_for('edit_all', id=s.id, d=d) }}">Editar</a>
            {% endif %}
          {% endif %}
        {% endif %}
      </td>
    </tr>
  {% endfor %}
</table>

{% if efectivo_total_dia is not none %}
  <div class="smallnote">
    <b>Total efectivo disponible del dia:</b> {{ efectivo_total_dia|money }}
    <br><b>Total ingreso del dia (neto):</b> {{ ingreso_neto_total_dia|money }}
  </div>
{% endif %}

<h2 style="margin-top:18px;">Resumen de cierres</h2>

<form method="get" action="{{ url_for('caja_index') }}" class="filters">
  <input type="hidden" name="d" value="{{d}}">

  <div>
    <label>Desde</label>
    <input type="date" name="start" value="{{start}}">
  </div>

  <div>
    <label>Hasta</label>
    <input type="date" name="end" value="{{end}}">
  </div>

  <div>
    <label>Turno</label>
    <select name="turn">
      <option value="ALL" {% if turn=='ALL' %}selected{% endif %}>Todos</option>
      <option value="MORNING" {% if turn=='MORNING' %}selected{% endif %}>Manana</option>
      <option value="AFTERNOON" {% if turn=='AFTERNOON' %}selected{% endif %}>Tarde</option>
    </select>
  </div>

  <div>
    <label>Responsable</label>
    <select name="resp">
      <option value="ALL" {% if resp=='ALL' %}selected{% endif %}>Todos</option>
      {% for r in responsibles %}
        <option value="{{r}}" {% if resp==r %}selected{% endif %}>{{r}}</option>
      {% endfor %}
    </select>
  </div>

  <div>
    <button class="btn" type="submit">Aplicar</button>
  </div>
</form>

<table class="tight summary-table">
  <tr>
    <th>Dia</th>
    <th class="nowrap">Fecha</th>
    <th>Turno</th>
    <th>Responsable</th>
    <th>Caja<br>Inicial</th>
    <th>Efectivo<br>bruto</th>
    <th>Ventas<br>Mercado Pago</th>
    <th>Ventas<br>Pedidos Ya</th>
    <th>Ventas<br>Rappi</th>
    <th>Egresos</th>
    <th>Ventas<br>totales</th>
    <th class="net-soft">Ventas<br>netas</th>
    <th class="net-soft">Ventas netas<br>del dia</th>
    <th>Retirado</th>
    <th>Caja Final<br>(Real)</th>
  </tr>

  {% for row in summary %}
    <tr>
      {% if row.show_day %}
        <td class="wrapday" rowspan="{{ row.day_rowspan }}"><b>{{ row.weekday }}</b></td>
      {% endif %}
      <td class="nowrap">{{ row.day }}</td>
      <td>{{ row.turn_name }}</td>
      <td>{{ row.responsible }}</td>
      <td>{{ row.opening_cash|money }}</td>
      <td>{{ row.sales_cash|money }}</td>
      <td>{{ row.sales_mp|money }}</td>
      <td>{{ row.sales_pya|money }}</td>
      <td>{{ row.sales_rappi|money }}</td>
      <td>{{ row.expenses|money }}</td>
      <td>{{ row.ventas_bruto|money }}</td>
      <td class="net-soft"><b>{{ row.ventas_neto|money }}</b></td>
      {% if row.show_day_total %}
        <td class="net-soft" rowspan="{{ row.day_rowspan }}"><b>{{ row.day_total_neto|money }}</b></td>
      {% endif %}
      <td>{{ row.withdrawn|money }}</td>
      <td>{{ row.ending_real|money }}</td>
    </tr>
  {% endfor %}

  {% if not summary %}
    <tr><td colspan="15" class="muted">Sin cierres para el filtro seleccionado.</td></tr>
  {% endif %}
</table>

<p class="muted">DB: <code>{{db_path}}</code></p>
</body>
</html>
"""

OPEN_HTML = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Abrir turno</title>
{{ base_css|safe }}
<style>
  body{max-width:540px;}
  input,button{padding:10px; font-size:14px;}
  .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body>
<a href="{{ url_for('caja_index', d=d) }}"><- Volver</a>
<h3>Abrir turno {{turn_name}} ({{day}})</h3>

<div class="box">
  <form method="post">
    <input type="hidden" name="d" value="{{d}}">

    {% if locked_opening is not none %}
      <p><b>Caja inicial</b> (del cierre anterior):</p>
      <p><input type="number" name="opening_cash" value="{{locked_opening}}" readonly></p>
      <p class="muted">No se puede editar.</p>
    {% else %}
      <p><b>Caja inicial</b> (no hay cierre previo):</p>
      <p><input type="number" name="opening_cash" min="0" required></p>
      <p class="muted">Solo se usa si es el primer turno sin historial.</p>
    {% endif %}

    <p><b>Responsable:</b></p>
    <input name="responsible" required style="width:100%;" value="{{default_responsible}}" readonly>

    <br><br>
    <button class="btn">Abrir</button>
  </form>
</div>
</body>
</html>
"""

SHIFT_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Turno</title>
  {{ base_css|safe }}
  <style>
    body{max-width:1100px;}
    input,select,button,textarea{padding:8px; font-size:14px;}
    .bad{color:#842029; font-weight:bold;}
    .good{color:#0f5132; font-weight:bold;}
    .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
  .delivery-actions{display:flex; gap:10px; align-items:center; flex-wrap:wrap;}
  .delivery-backdrop{display:none; position:fixed; inset:0; background:rgba(0,0,0,.35); z-index:999;}
  .delivery-modal{display:none; position:fixed; top:50%; left:50%; transform:translate(-50%,-50%); width:min(960px, calc(100vw - 24px)); max-height:85vh; overflow:auto; background:#fff; border:1px solid #ccc; border-radius:18px; padding:18px; z-index:1000; box-shadow:0 18px 45px rgba(0,0,0,.22);}
  .delivery-modal h3{margin:0 0 8px;}
  .delivery-modal .head{display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:10px;}
  .delivery-table{width:100%; border-collapse:collapse; table-layout:fixed;}
  .delivery-table th,.delivery-table td{padding:10px 8px; border:1px solid #ddd; vertical-align:middle;}
  .delivery-table th{text-align:left; background:#f3f3f3;}
  .delivery-table .money-input-wrap{display:inline-flex; align-items:center; gap:6px; white-space:nowrap;}
  .delivery-table .money-input-wrap input{width:84px; text-align:center; padding:8px 6px;}
  .delivery-table .qty-input{width:72px; text-align:center; padding:8px 6px;}
  .delivery-table .concept-col{width:24%;}
  .delivery-table .rate-col{width:24%;}
  .delivery-table .qty-col{width:22%;}
  .delivery-table .tot-col{width:30%;}
  .delivery-total-row td{font-weight:bold; background:#fafafa;}
  .delivery-consumos-box{margin-top:12px; padding:12px; border:1px solid #ddd; border-radius:12px; background:#fafafa;}
  .delivery-consumos-grid{display:grid; grid-template-columns:160px 1fr; gap:10px; align-items:center;}
  .delivery-consumos-grid input{padding:8px 10px;}
  .delivery-consumos-grid .money-input-wrap input{width:96px; text-align:center;}
  .delivery-impact-box{margin-top:12px; display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap;}
  .delivery-impact-total{font-weight:bold; font-size:18px;}
</style>
</head>
<body>
<a href="{{ url_for('caja_index', d=d) }}"><- Volver</a>
<h2>Turno {{turn_name}} - {{s.day}}</h2>

<p>
  Responsable: <b>{{s.responsible}}</b> |
  Estado: <b>{{s.status}}</b> |
  Caja inicial: <b>{{ s.opening_cash|money }}</b>
</p>

{% if close and can_edit %}
  <p class="muted">
    <a class="btn" href="{{ url_for('edit_all', id=s.id, d=d) }}">Editar (todo)</a>
    {% if close.edit_count %}
      <span class="muted">Ediciones: {{close.edit_count}}</span>
    {% endif %}
  </p>
{% elif close and not can_edit %}
  <p class="muted">Este turno ya no puede editarse con tu usuario.</p>
{% endif %}


<h3>Ventas</h3>
<div class="box">
  <div class="row" style="align-items:flex-start;">
    <div style="min-width:280px;">
      <label><b>Retirado (Efectivo total)</b></label><br>
      <input type="number" name="sales_cash" min="0" value="{{s.sales_cash}}" {{'disabled' if s.status=='CLOSED' else ''}} id="retirado_input">
      <div class="muted" style="margin-top:6px;">
        Efectivo bruto (auto): <b id="ef_bruto_lbl">{{ (s.sales_cash + (close.ending_cash if close else 0))|money }}</b><br>
        Efectivo neto (auto): <b id="ef_neto_lbl">{{ (s.sales_cash - s.opening_cash)|money }}</b>
      </div>
    </div>

    <div style="min-width:320px;">
      <label><b>Caja final (efectivo que queda en la caja)</b></label><br>
      <input type="number" name="cash_final" min="0" value="{{ (close.ending_cash if close else 0) }}" {{'disabled' if s.status=='CLOSED' else ''}} id="caja_final_input">
      <div class="muted" style="margin-top:6px;">
        (Se carga manualmente)
      </div>
    </div>
  </div>

  <div class="row" style="margin-top:10px;">
    <div>
      Ventas Mercado Pago:<br>
      <input type="number" name="sales_mp" min="0" value="{{s.sales_mp}}" {{'disabled' if s.status=='CLOSED' else ''}} id="mp_input">
    </div>
    <div>
      Ventas Pedidos Ya:<br>
      <input type="number" name="sales_pya" min="0" value="{{s.sales_pya}}" {{'disabled' if s.status=='CLOSED' else ''}} id="pya_input">
    </div>
    <div>
      Ventas Rappi:<br>
      <input type="number" name="sales_rappi" min="0" value="{{s.sales_rappi}}" {{'disabled' if s.status=='CLOSED' else ''}} id="rappi_input">
    </div>
  </div>
</div>


<h3>Egresos (caja chica, efectivo)</h3>
{% if s.status != 'CLOSED' %}
<div class="delivery-actions">
  <form method="post" action="{{ url_for('expense', id=s.id, d=d) }}" id="expenseForm">
    <select name="category" id="expenseCategory">
      {% for c in categories %}<option>{{c}}</option>{% endfor %}
    </select>
    <input type="number" name="amount" min="1" required id="expenseAmount">
    <input name="note" placeholder="nota (obligatoria si Otros)" id="expenseNote">
    <button class="btn">Agregar</button>
  </form>

  <button type="button" class="btn" id="openDeliveryCalcBtn">Calculo para delivery</button>
</div>

<div class="delivery-backdrop" id="deliveryBackdrop"></div>
<div class="delivery-modal" id="deliveryModal" aria-hidden="true">
  <div class="head">
    <div>
      <h3>Calculadora pago delivery</h3>
      <div class="muted">Los importes por pedido/hora vienen precargados, pero podes editarlos. La informacion se guarda automaticamente en este turno.</div>
    </div>
    <button type="button" class="btn" id="closeDeliveryCalcBtn">Cerrar</button>
  </div>

  <table class="delivery-table">
    <tr>
      <th class="concept-col">Concepto</th>
      <th class="rate-col">$ por pedido/hora</th>
      <th class="qty-col">Cantidad</th>
      <th class="tot-col">Total</th>
    </tr>
    <tr>
      <td rowspan="4" style="text-align:center;"><b>Pedidos</b></td>
      <td>
        <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="0" value="1500"></div>
      </td>
      <td><input type="number" class="del-qty qty-input" data-row="0" value="0" min="0" step="1"></td>
      <td class="del-total" data-row="0">$ 0</td>
    </tr>
    <tr>
      <td>
        <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="1" value="2000"></div>
      </td>
      <td><input type="number" class="del-qty qty-input" data-row="1" value="0" min="0" step="1"></td>
      <td class="del-total" data-row="1">$ 0</td>
    </tr>
    <tr>
      <td>
        <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="2" value="2500"></div>
      </td>
      <td><input type="number" class="del-qty qty-input" data-row="2" value="0" min="0" step="1"></td>
      <td class="del-total" data-row="2">$ 0</td>
    </tr>
    <tr>
      <td>
        <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="3" value="3000"></div>
      </td>
      <td><input type="number" class="del-qty qty-input" data-row="3" value="0" min="0" step="1"></td>
      <td class="del-total" data-row="3">$ 0</td>
    </tr>
    <tr>
      <td style="text-align:center;"><b>Horas</b></td>
      <td>
        <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="4" value="2500"></div>
      </td>
      <td>
        <div style="display:grid; gap:8px;">
          <div>
            <label class="muted" for="deliveryHourShift">Turno</label><br>
            <select id="deliveryHourShift">
              <option value="MORNING">Mañana</option>
              <option value="AFTERNOON">Tarde</option>
            </select>
          </div>
          <div class="row" style="gap:8px; align-items:end;">
            <div>
              <label class="muted" for="deliveryHourIn">Entrada</label><br>
              <input type="time" id="deliveryHourIn" class="qty-input" style="width:120px;">
            </div>
            <div>
              <label class="muted" for="deliveryHourOut">Salida</label><br>
              <input type="time" id="deliveryHourOut" class="qty-input" style="width:120px;">
            </div>
          </div>
          <div class="muted">Horas calculadas automáticamente: <b id="deliveryHoursLabel">0.00</b></div>
          <input type="number" class="del-qty qty-input" data-row="4" value="0" min="0" step="0.01" readonly>
        </div>
      </td>
      <td class="del-total" data-row="4">$ 0</td>
    </tr>
    <tr class="delivery-total-row">
      <td colspan="3" style="text-align:right;">Total viajes / horas</td>
      <td id="deliveryGrandTotal">$ 0</td>
    </tr>
  </table>

  <div class="delivery-consumos-box">
    <div style="font-weight:bold; margin-bottom:8px;">Consumos</div>
    <div class="delivery-consumos-grid">
      <div>
        <label class="muted" for="deliveryConsumeAmount">Monto</label><br>
        <div class="money-input-wrap">$ <input type="number" id="deliveryConsumeAmount" min="0" value="0"></div>
      </div>
      <div>
        <label class="muted" for="deliveryConsumeNote">Notas</label><br>
        <input type="text" id="deliveryConsumeNote" placeholder="Ej: gaseosa, cena, peaje, etc.">
      </div>
    </div>
    <div class="muted" style="margin-top:8px;">El monto de consumos se descuenta del total de viajes / horas. Las horas se calculan solas según turno, entrada y salida.</div>
  </div>

  <div class="delivery-impact-box">
    <div>
      <div class="muted">Total a cargar como egreso Delivery / Cadete</div>
      <div class="delivery-impact-total" id="deliveryNetTotal">$ 0</div>
    </div>
    <button type="button" class="btn" id="applyDeliveryExpenseBtn">Impactar en egresos</button>
  </div>
</div>
{% else %}
<p class="muted">Turno cerrado: no se pueden agregar egresos desde aca. Usa Editar (todo).</p>
{% endif %}

<table style="margin-top:10px;">
  <tr><th>Categoria</th><th>Monto</th><th>Nota</th></tr>
  {% for e in expenses %}
    <tr><td>{{e.category}}</td><td>{{e.amount|money}}</td><td>{{e.note or ''}}</td></tr>
  {% endfor %}
  {% if not expenses %}
    <tr><td colspan="3" class="muted">Sin egresos</td></tr>
  {% endif %}
</table>


<h3>Cierre</h3>

<div class="box">
  <div class="row">
    <div><b>Egresos total:</b> <span id="egresos_total_lbl">{{ egresos_total|money }}</span></div>
    <div><b>Ingreso total (bruto):</b> <span id="ingreso_bruto_lbl">{{ ingreso_bruto|money }}</span></div>
    <div><b>Ingreso neto:</b> <span id="ingreso_neto_lbl">{{ ingreso_neto|money }}</span></div>
    <div><b>Efectivo disponible:</b> <span id="ef_disp_lbl">{{ efectivo_disponible|money }}</span></div>
  </div>
  <div class="muted" style="margin-top:6px;">
    Efectivo bruto = Retirado + Caja final.<br>
    Efectivo neto = Retirado - Caja inicial.<br>
    Ingreso total (bruto) = (Efectivo bruto + MP + PedidosYa + Rappi) + Egresos.<br>
    Ingreso neto = Ingreso total (bruto) - Egresos total.<br>
    Efectivo disponible = Retirado.
  </div>
</div>

{% if close %}
  <div class="box" style="margin-top:10px;">
    <div class="row">
      <div><b>Retirado:</b> {{ s.sales_cash|money }}</div>
      <div><b>Caja final:</b> {{ close.ending_cash|money }}</div>
    </div>
  </div>
{% else %}
  <form method="post" action="{{ url_for('close_shift', id=s.id, d=d) }}" style="margin-top:14px;" id="saveForm">
    <input type="hidden" name="sales_cash" id="h_sales_cash">
    <input type="hidden" name="cash_final" id="h_cash_final">
    <input type="hidden" name="sales_mp" id="h_sales_mp">
    <input type="hidden" name="sales_pya" id="h_sales_pya">
    <input type="hidden" name="sales_rappi" id="h_sales_rappi">
    <button class="btn" style="margin-top:10px;">Guardar registro</button>
  </form>

  <script>
    const sid = "{{s.id}}";

    function toInt(v){
      v = (v || "").toString().trim();
      if(v === "") return 0;
      const n = parseInt(v, 10);
      return isNaN(n) ? 0 : n;
    }

    function toNum(v){
      v = (v || "").toString().trim().replace(",", ".");
      if(v === "") return 0;
      const n = parseFloat(v);
      return isNaN(n) ? 0 : n;
    }

    const retiradoEl = document.getElementById("retirado_input");
    const cajaFinalEl = document.getElementById("caja_final_input");
    const mpEl = document.getElementById("mp_input");
    const pyaEl = document.getElementById("pya_input");
    const rappiEl = document.getElementById("rappi_input");

    const efBrutoLbl = document.getElementById("ef_bruto_lbl");
    const efNetoLbl = document.getElementById("ef_neto_lbl");
    const ingresoBrutoLbl = document.getElementById("ingreso_bruto_lbl");
    const ingresoNetoLbl = document.getElementById("ingreso_neto_lbl");
    const efDispLbl = document.getElementById("ef_disp_lbl");

    const egresosTotal = {{ egresos_total|int }};

    function fmtMoney(n){
      try{ n = parseInt(n,10); }catch(e){ n = 0; }
      if(isNaN(n)) n = 0;
      const s = n.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
      return "$ " + s;
    }

    function saveDraft(){
      const payload = {
        retirado: toInt(retiradoEl.value),
        caja_final: toInt(cajaFinalEl.value),
        mp: toInt(mpEl.value),
        pya: toInt(pyaEl.value),
        rappi: toInt(rappiEl.value)
      };
      try{
        localStorage.setItem("caja_draft_"+sid, JSON.stringify(payload));
      }catch(e){}
    }

    const deliveryStorageKey = "delivery_calc_" + sid;
    const deliveryModal = document.getElementById("deliveryModal");
    const deliveryBackdrop = document.getElementById("deliveryBackdrop");
    const openDeliveryCalcBtn = document.getElementById("openDeliveryCalcBtn");
    const closeDeliveryCalcBtn = document.getElementById("closeDeliveryCalcBtn");
    const deliveryRateEls = Array.from(document.querySelectorAll(".del-rate"));
    const deliveryQtyEls = Array.from(document.querySelectorAll(".del-qty"));
    const deliveryTotalEls = Array.from(document.querySelectorAll(".del-total"));
    const deliveryGrandTotalEl = document.getElementById("deliveryGrandTotal");
    const deliveryConsumeAmountEl = document.getElementById("deliveryConsumeAmount");
    const deliveryConsumeNoteEl = document.getElementById("deliveryConsumeNote");
    const deliveryNetTotalEl = document.getElementById("deliveryNetTotal");
    const applyDeliveryExpenseBtn = document.getElementById("applyDeliveryExpenseBtn");
    const expenseFormEl = document.getElementById("expenseForm");
    const expenseCategoryEl = document.getElementById("expenseCategory");
    const expenseAmountEl = document.getElementById("expenseAmount");
    const expenseNoteEl = document.getElementById("expenseNote");

    const deliveryPresets = {{ delivery_shift_presets_json|safe }};
    const deliveryInitialData = {{ delivery_payload_json|safe }};
    const deliveryHourShiftEl = document.getElementById("deliveryHourShift");
    const deliveryHourInEl = document.getElementById("deliveryHourIn");
    const deliveryHourOutEl = document.getElementById("deliveryHourOut");
    const deliveryHoursLabelEl = document.getElementById("deliveryHoursLabel");

    function getDeliveryDefaults(){
      return JSON.parse(JSON.stringify(deliveryInitialData || {
        rates: [1500, 2000, 2500, 3000, 2500],
        qtys: [0, 0, 0, 0, 0],
        consume_amount: 0,
        consume_note: "",
        hour_shift: "MORNING",
        hour_in: "08:55",
        hour_out: "12:50"
      }));
    }

    function calcHourQty(){
      const hIn = (deliveryHourInEl && deliveryHourInEl.value) ? deliveryHourInEl.value : "";
      const hOut = (deliveryHourOutEl && deliveryHourOutEl.value) ? deliveryHourOutEl.value : "";
      if(!hIn || !hOut) return 0;
      const [ih, im] = hIn.split(":").map(v => parseInt(v, 10));
      const [oh, om] = hOut.split(":").map(v => parseInt(v, 10));
      if([ih, im, oh, om].some(v => isNaN(v))) return 0;
      const mins = ((oh * 60 + om) - (ih * 60 + im));
      if(mins < 0) return 0;
      return Math.round((mins / 60) * 100) / 100;
    }

    function updateHoursFromTimes(){
      const hours = calcHourQty();
      if(deliveryQtyEls[4]) deliveryQtyEls[4].value = hours.toFixed(2);
      if(deliveryHoursLabelEl) deliveryHoursLabelEl.textContent = hours.toFixed(2) + " hs";
      return hours;
    }

    function applyPresetForShift(force=false){
      if(!deliveryHourShiftEl) return;
      const shiftKey = deliveryHourShiftEl.value === "AFTERNOON" ? "AFTERNOON" : "MORNING";
      const preset = deliveryPresets[shiftKey] || deliveryPresets["MORNING"];
      if(force || !deliveryHourInEl.value) deliveryHourInEl.value = preset.hour_in || "";
      if(force || !deliveryHourOutEl.value) deliveryHourOutEl.value = preset.hour_out || "";
      updateHoursFromTimes();
    }

    let deliverySaveTimer = null;
    function persistDeliveryCalc(){
      saveDeliveryCalc();
      if(deliverySaveTimer) clearTimeout(deliverySaveTimer);
      deliverySaveTimer = setTimeout(async () => {
        try{
          await fetch("{{ url_for('save_delivery_draft', id=s.id) }}", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
              rates: deliveryRateEls.map(el => toInt(el.value)),
              qtys: deliveryQtyEls.map((el, i) => i === 4 ? toNum(el.value) : toInt(el.value)),
              consume_amount: toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0),
              consume_note: deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "") : "",
              hour_shift: deliveryHourShiftEl ? deliveryHourShiftEl.value : "MORNING",
              hour_in: deliveryHourInEl ? deliveryHourInEl.value : "",
              hour_out: deliveryHourOutEl ? deliveryHourOutEl.value : ""
            })
          });
        }catch(e){}
      }, 350);
    }

    function saveDeliveryCalc(){
      try{
        const payload = {
          rates: deliveryRateEls.map(el => toInt(el.value)),
          qtys: deliveryQtyEls.map((el, i) => i === 4 ? toNum(el.value) : toInt(el.value)),
          consume_amount: toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0),
          consume_note: deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "") : "",
          hour_shift: deliveryHourShiftEl ? deliveryHourShiftEl.value : "MORNING",
          hour_in: deliveryHourInEl ? deliveryHourInEl.value : "",
          hour_out: deliveryHourOutEl ? deliveryHourOutEl.value : ""
        };
        localStorage.setItem(deliveryStorageKey, JSON.stringify(payload));
      }catch(e){}
    }

    function loadDeliveryCalc(){
      const defaults = getDeliveryDefaults();
      try{
        const raw = localStorage.getItem(deliveryStorageKey);
        const payload = raw ? (JSON.parse(raw) || {}) : defaults;
        deliveryRateEls.forEach((el, i) => {
          el.value = Array.isArray(payload.rates) && payload.rates[i] != null ? payload.rates[i] : defaults.rates[i];
        });
        deliveryQtyEls.forEach((el, i) => {
          el.value = Array.isArray(payload.qtys) && payload.qtys[i] != null ? payload.qtys[i] : defaults.qtys[i];
        });
        if(deliveryConsumeAmountEl){
          deliveryConsumeAmountEl.value = payload.consume_amount != null ? payload.consume_amount : defaults.consume_amount;
        }
        if(deliveryConsumeNoteEl){
          deliveryConsumeNoteEl.value = payload.consume_note != null ? payload.consume_note : defaults.consume_note;
        }
        if(deliveryHourShiftEl){
          deliveryHourShiftEl.value = payload.hour_shift || defaults.hour_shift || "MORNING";
        }
        if(deliveryHourInEl){
          deliveryHourInEl.value = payload.hour_in || defaults.hour_in || "";
        }
        if(deliveryHourOutEl){
          deliveryHourOutEl.value = payload.hour_out || defaults.hour_out || "";
        }
        updateHoursFromTimes();
      }catch(e){
        deliveryRateEls.forEach((el, i) => el.value = defaults.rates[i]);
        deliveryQtyEls.forEach((el, i) => el.value = defaults.qtys[i]);
        if(deliveryConsumeAmountEl) deliveryConsumeAmountEl.value = defaults.consume_amount;
        if(deliveryConsumeNoteEl) deliveryConsumeNoteEl.value = defaults.consume_note;
        if(deliveryHourShiftEl) deliveryHourShiftEl.value = defaults.hour_shift || "MORNING";
        if(deliveryHourInEl) deliveryHourInEl.value = defaults.hour_in || "";
        if(deliveryHourOutEl) deliveryHourOutEl.value = defaults.hour_out || "";
        updateHoursFromTimes();
      }
    }

    function recalcDeliveryCalc(){
      updateHoursFromTimes();
      let grand = 0;
      deliveryRateEls.forEach((el, i) => {
        const rate = toNum(el.value);
        const qty = (i === 4) ? toNum(deliveryQtyEls[i].value) : toInt(deliveryQtyEls[i].value);
        const total = rate * qty;
        grand += total;
        if(deliveryTotalEls[i]){
          deliveryTotalEls[i].textContent = fmtMoney(Math.round(total));
        }
      });
      const consumeAmount = toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0);
      const net = Math.max(0, Math.round(grand - consumeAmount));
      if(deliveryGrandTotalEl){
        deliveryGrandTotalEl.textContent = fmtMoney(Math.round(grand));
      }
      if(deliveryNetTotalEl){
        deliveryNetTotalEl.textContent = fmtMoney(Math.round(net));
      }
    }

    function openDeliveryCalc(){
      if(deliveryModal) deliveryModal.style.display = "block";
      if(deliveryBackdrop) deliveryBackdrop.style.display = "block";
      document.body.style.overflow = "hidden";
    }

    function closeDeliveryCalc(){
      if(deliveryModal) deliveryModal.style.display = "none";
      if(deliveryBackdrop) deliveryBackdrop.style.display = "none";
      document.body.style.overflow = "";
    }

    if(openDeliveryCalcBtn){
      openDeliveryCalcBtn.addEventListener("click", openDeliveryCalc);
    }
    if(closeDeliveryCalcBtn){
      closeDeliveryCalcBtn.addEventListener("click", closeDeliveryCalc);
    }
    if(deliveryBackdrop){
      deliveryBackdrop.addEventListener("click", closeDeliveryCalc);
    }
    document.addEventListener("keydown", (ev) => {
      if(ev.key === "Escape" && deliveryModal && deliveryModal.style.display === "block"){
        closeDeliveryCalc();
      }
    });

    [...deliveryRateEls, ...deliveryQtyEls, deliveryConsumeAmountEl, deliveryConsumeNoteEl, deliveryHourInEl, deliveryHourOutEl].filter(Boolean).forEach(el => {
      el.addEventListener("input", () => {
        persistDeliveryCalc();
        recalcDeliveryCalc();
      });
    });
    if(deliveryHourShiftEl){
      deliveryHourShiftEl.addEventListener("change", () => {
        applyPresetForShift(true);
        persistDeliveryCalc();
        recalcDeliveryCalc();
      });
    }

    if(applyDeliveryExpenseBtn){
      applyDeliveryExpenseBtn.addEventListener("click", () => {
        const grand = deliveryRateEls.reduce((acc, el, i) => {
          const rate = toNum(el.value);
          const qty = (i === 4) ? toNum(deliveryQtyEls[i].value) : toInt(deliveryQtyEls[i].value);
          return acc + (rate * qty);
        }, 0);
        const consumeAmount = toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0);
        const net = Math.max(0, Math.round(grand - consumeAmount));
        const consumeNote = deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "").trim() : "";

        if(!expenseFormEl || !expenseCategoryEl || !expenseAmountEl || net <= 0){
          alert("El total a cargar debe ser mayor a 0.");
          return;
        }

        expenseCategoryEl.value = "Delivery / Cadete (efectivo)";
        expenseAmountEl.value = net;

        if(expenseNoteEl){
          let note = "Calculadora delivery";
          if(consumeAmount > 0){
            note += " - consumo descontado: " + fmtMoney(consumeAmount);
          }
          if(consumeNote){
            note += " - " + consumeNote;
          }
          expenseNoteEl.value = note;
        }

        persistDeliveryCalc();
        expenseFormEl.submit();
      });
    }

    function loadDraft(){
      try{
        const raw = localStorage.getItem("caja_draft_"+sid);
        if(!raw) return;
        const p = JSON.parse(raw);
        if(p && typeof p === "object"){
          if(retiradoEl.value === "" || retiradoEl.value === "0") retiradoEl.value = p.retirado ?? retiradoEl.value;
          if(cajaFinalEl.value === "" || cajaFinalEl.value === "0") cajaFinalEl.value = p.caja_final ?? cajaFinalEl.value;
          if(mpEl.value === "" || mpEl.value === "0") mpEl.value = p.mp ?? mpEl.value;
          if(pyaEl.value === "" || pyaEl.value === "0") pyaEl.value = p.pya ?? pyaEl.value;
          if(rappiEl.value === "" || rappiEl.value === "0") rappiEl.value = p.rappi ?? rappiEl.value;
        }
      }catch(e){}
    }

    function recalc(){
      const retirado = toInt(retiradoEl.value);
      const cajaFinal = toInt(cajaFinalEl.value);
      const mp = toInt(mpEl.value);
      const pya = toInt(pyaEl.value);
      const rappi = toInt(rappiEl.value);

      const efectivoBruto = retirado + cajaFinal;
      const efectivoNeto = retirado - {{ s.opening_cash|int }};

      const ingresoBruto = (efectivoBruto + mp + pya + rappi) + egresosTotal;
      const ingresoNeto = ingresoBruto - egresosTotal; // pedido: neto = bruto - egresos

      efBrutoLbl.textContent = fmtMoney(efectivoBruto);
      efNetoLbl.textContent = fmtMoney(efectivoNeto);
      ingresoBrutoLbl.textContent = fmtMoney(ingresoBruto);
      ingresoNetoLbl.textContent = fmtMoney(ingresoNeto);
      efDispLbl.textContent = fmtMoney(retirado);
    }

    [retiradoEl, cajaFinalEl, mpEl, pyaEl, rappiEl].forEach(el => {
      el.addEventListener("input", () => { saveDraft(); recalc(); });
    });

    // Restore draft after page reload (ej: al agregar egresos)
    loadDeliveryCalc();
    applyPresetForShift(false);
    recalcDeliveryCalc();
    loadDraft();
    recalc();

    // Submit: enviar valores actuales
    document.getElementById("saveForm").addEventListener("submit", (e) => {
      document.getElementById("h_sales_cash").value = toInt(retiradoEl.value);
      document.getElementById("h_cash_final").value = toInt(cajaFinalEl.value);
      document.getElementById("h_sales_mp").value = toInt(mpEl.value);
      document.getElementById("h_sales_pya").value = toInt(pyaEl.value);
      document.getElementById("h_sales_rappi").value = toInt(rappiEl.value);
      // no borramos draft: por si vuelve atras
    });
  </script>

{% endif %}

</body>

</html>
"""

EDIT_ALL_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Editar turno</title>
{{ base_css|safe }}
<style>
  body{max-width:1200px;}
  input,select,button,textarea{padding:8px; font-size:14px;}
  .row{display:grid; grid-template-columns:repeat(4, 1fr); gap:10px;}
  .row2{display:grid; grid-template-columns:1fr 2fr; gap:10px;}
  .right{text-align:right;}
  .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body>

<a href="{{ url_for('shift', id=s.id, d=d) }}"><- Volver</a>
<h2 style="margin-bottom:6px;">Editar (todo) - {{turn_name}} {{s.day}}</h2>
<p class="muted">Usuario: <b>{{user.username}}</b> ({{user.role}}) - Ediciones previas: {{close.edit_count}}</p>

<form method="post">
  <div class="box">
    <h3>Datos del turno</h3>
    <div class="row">
      <div>
        <label>Responsable</label><br>
        <input type="text" name="responsible" value="{{s.responsible}}" required style="width:100%;">
      </div>
      <div>
        <label>Caja inicial</label><br>
        <input type="number" name="opening_cash" min="0" value="{{s.opening_cash}}" required style="width:100%;">
      </div>
      <div>
        <label>Turno</label><br>
        <input type="text" value="{{turn_name}}" readonly style="width:100%;">
      </div>
      <div>
        <label>Fecha</label><br>
        <input type="text" value="{{s.day}}" readonly style="width:100%;">
      </div>
    </div>
    <p class="muted">Cambiar CI impacta calculos del turno siguiente.</p>
  </div>

  <div class="box">
    <h3>Ventas</h3>
    <div class="row">
      <div>Ventas efectivo (neto)<br><input type="number" name="sales_cash" min="0" value="{{s.sales_cash}}" style="width:100%;"></div>
      <div>Ventas Mercado Pago<br><input type="number" name="sales_mp" min="0" value="{{s.sales_mp}}" style="width:100%;"></div>
      <div>Ventas Pedidos Ya<br><input type="number" name="sales_pya" min="0" value="{{s.sales_pya}}" style="width:100%;"></div>
      <div>Ventas Rappi<br><input type="number" name="sales_rappi" min="0" value="{{s.sales_rappi}}" style="width:100%;"></div>
    </div>
  </div>

  <div class="box">
    <h3>Egresos (editar / borrar / agregar)</h3>
    <table>
      <tr><th>Eliminar</th><th>Categoria</th><th class="right">Monto</th><th>Nota</th></tr>
      {% for e in expenses %}
        <tr>
          <td><input type="checkbox" name="del_{{e.id}}"></td>
          <td>
            <select name="cat_{{e.id}}">
              {% for c in categories %}
                <option value="{{c}}" {% if e.category==c %}selected{% endif %}>{{c}}</option>
              {% endfor %}
            </select>
          </td>
          <td class="right"><input type="number" name="amt_{{e.id}}" min="0" value="{{e.amount}}" style="width:120px;"></td>
          <td><input type="text" name="note_{{e.id}}" value="{{e.note or ''}}" style="width:100%;"></td>
        </tr>
      {% endfor %}
      {% if not expenses %}
        <tr><td colspan="4" class="muted">Sin egresos.</td></tr>
      {% endif %}
    </table>

    <h4 style="margin-top:12px;">Agregar egresos (hasta 5)</h4>
    <p class="muted" style="margin-top:-6px;">Completa categoria y monto para cada egreso que quieras agregar.</p>
    <div class="grid">
      {% for i in range(1,6) %}
      <div class="row2" style="align-items:center;">
        <div>
          <select name="new_category_{{i}}">
            <option value="">(no agregar)</option>
            {% for c in categories %}<option value="{{c}}">{{c}}</option>{% endfor %}
          </select>
        </div>
        <div style="display:flex; gap:10px; align-items:center; width:100%;">
          <input type="number" name="new_amount_{{i}}" min="0" placeholder="$" style="width:140px;">
          <input type="text" name="new_note_{{i}}" placeholder="nota (obligatoria si Otros)" style="flex:1;">
        </div>
      </div>
      {% endfor %}
    </div>
  </div>

  <div class="box">
    
    <h3>Cierre</h3>
    <div class="row">
      <div>
        <label>Retirado (Efectivo total)</label><br>
        <input type="number" name="withdrawn" min="0" value="{{close.withdrawn_cash}}" required style="width:100%;">
      </div>
      <div>
        <label>Caja final</label><br>
        <input type="number" name="ending_real" min="0" value="{{close.ending_cash}}" required style="width:100%;">
      </div>
      <div>
        <label class="muted">Estado</label><br>
        <input type="text" value="OK" readonly style="width:100%;">
      </div>
      <div>
        <label class="muted">Observacion</label><br>
        <input type="text" value="{{close.note or ''}}" readonly style="width:100%;">
      </div>
    </div>
    <p class="muted">La diferencia se mantiene en 0 (sin validacion OK/NO OK por ahora).</p>

  </div>

  <div class="box">
    <h3>Auditoria</h3>
    <div class="row2">
      <div><b>Motivo de edicion (obligatorio)</b></div>
      <div><input type="text" name="reason" required placeholder="Ej: se cargo venta luego del cierre" style="width:100%;"></div>
    </div>
  </div>

  <button class="btn">Guardar edicion</button>
</form>

</body>
</html>
"""


def is_placeholder_shift(s: "Shift") -> bool:
    """Devuelve True si el turno existe en BD pero en la practica esta 'vacio'
    (tipico de importaciones/placeholder): todo en 0, sin egresos, y (si existe) cierre vacio.
    En ese caso, en el panel principal lo tratamos como NO ABIERTO para mostrar el boton 'Abrir'.
    """
    if not s:
        return True

    # Solo consideramos placeholder a turnos CERRADOS en cero (evita ocultar un turno abierto recien iniciado).
    if (s.status or "").upper() != "CLOSED":
        return False

    nums = [
        int(s.opening_cash or 0),
        int(s.sales_cash or 0),
        int(s.sales_mp or 0),
        int(s.sales_pya or 0),
        int(s.sales_rappi or 0),
        int(getattr(s, "sales_apps", 0) or 0),
    ]
    if any(v != 0 for v in nums):
        return False

    # Si tiene egresos, NO es placeholder
    if CashExpense.query.filter_by(shift_id=s.id).first():
        return False

    c = ShiftClose.query.filter_by(shift_id=s.id).first()
    if c:
        close_nums = [
            int(c.withdrawn_cash or 0),
            int(c.ending_calc or 0),
            int(c.ending_cash or 0),
            int(c.difference or 0),
        ]
        if any(v != 0 for v in close_nums):
            return False
        if (c.note or "").strip():
            return False

    return True


@app.route("/caja")
@login_required
def caja_index():
    day_obj, d = parse_day_param(request.args.get("d"))

    start_date, end_date = parse_range_params(request.args.get("start"), request.args.get("end"))
    start = start_date.isoformat()
    end = end_date.isoformat()
    turn = request.args.get("turn") or "ALL"
    resp = request.args.get("resp") or "ALL"

    shifts = {s.turn: s for s in Shift.query.filter_by(day=day_obj).all()}

    # Si por importacion o error existe un turno 'cerrado' totalmente en cero, lo tratamos como NO ABIERTO.
    # (Asi el usuario ve el boton 'Abrir' en vez de 'Ver/Editar'.)
    for _t, _s in list(shifts.items()):
        if _s and is_placeholder_shift(_s):
            shifts[_t] = None

    closes = {}
    can_edit_map = {}
    u = current_user()

    for s in shifts.values():
        if not s:
            continue
        if s.status == "CLOSED":
            c = ShiftClose.query.filter_by(shift_id=s.id).first()
            closes[s.id] = c
            can_edit_map[s.id] = bool(c) and can_edit_close(u, c)

    ventas_netas_map = {}
    efectivo_disp_map = {}

    for s in shifts.values():
        if not s:
            continue
        if s.status == "CLOSED":
            c = closes.get(s.id)
            cash_final = int(getattr(c, 'ending_cash', 0) or 0) if c else 0
            eg = expenses_total(s.id, CashExpense)
            ventas_netas_map[s.id] = int(calc_ingreso_neto(s, egresos=eg, cash_final=cash_final))
            c = closes.get(s.id)
            efectivo_disp_map[s.id] = int(s.sales_cash or 0)

    efectivo_total_dia = None
    ingreso_neto_total_dia = None
    if shifts.get("MORNING") and shifts.get("AFTERNOON"):
        if shifts["MORNING"].status == "CLOSED" and shifts["AFTERNOON"].status == "CLOSED":
            efectivo_total_dia = (
                efectivo_disp_map.get(shifts["MORNING"].id, 0) +
                efectivo_disp_map.get(shifts["AFTERNOON"].id, 0)
            )
            ingreso_neto_total_dia = (ventas_netas_map.get(shifts["MORNING"].id, 0) + ventas_netas_map.get(shifts["AFTERNOON"].id, 0))

    def is_valid_responsible(name: str) -> bool:
        name = (name or "").strip()
        if not name:
            return False
        if name in EMPLOYEES or name in ADMINS:
            return True
        return False

    responsibles = sorted({
        s.responsible for s in Shift.query.filter(Shift.status == "CLOSED").all()
        if is_valid_responsible(s.responsible)
    })

    summary = get_caja_summary(
        limit=400,
        start=start_date,
        end=end_date,
        turn=turn,
        responsible=resp
    )

    # Agrupar visualmente por fecha en "Resumen de cierres"
    grouped_summary = []
    i = 0
    while i < len(summary):
        day_key = summary[i]["day"]
        j = i
        day_total_neto = 0
        while j < len(summary) and summary[j]["day"] == day_key:
            day_total_neto += int(summary[j].get("ventas_neto", 0) or 0)
            j += 1

        rowspan = j - i
        for k in range(i, j):
            row = dict(summary[k])
            row["day_rowspan"] = rowspan
            row["show_day"] = (k == i)
            row["show_day_total"] = (k == i)
            row["day_total_neto"] = day_total_neto
            grouped_summary.append(row)
        i = j

    summary = grouped_summary

    weekday_name = weekday_es(day_obj)

    return render_template_string(
        CAJA_INDEX_HTML,
        base_css=BASE_CSS,
        d=d,
        turns=TURNS,
        shifts=shifts,
        closes=closes,
        can_edit_map=can_edit_map,
        ventas_netas_map=ventas_netas_map,
        efectivo_disp_map=efectivo_disp_map,
        efectivo_total_dia=efectivo_total_dia,
        ingreso_neto_total_dia=ingreso_neto_total_dia,
        summary=summary,
        db_path=DB_PATH,
        start=start,
        end=end,
        turn=turn,
        resp=resp,
        responsibles=responsibles,
        weekday_name=weekday_name
    )

@app.route("/caja/open/<turn>", methods=["GET","POST"])
@login_required
def open_shift(turn):
    day_obj, d = parse_day_param(request.args.get("d") or request.form.get("d"))

    existing = Shift.query.filter_by(day=day_obj, turn=turn).first()
    if existing:
        # Si existe como placeholder (cerrado y todo en 0), lo limpiamos para permitir "Abrir" correctamente.
        if is_placeholder_shift(existing):
            try:
                ShiftClose.query.filter_by(shift_id=existing.id).delete()
                CashExpense.query.filter_by(shift_id=existing.id).delete()
                db.session.delete(existing)
                db.session.commit()
                backup_caja_local_y_drive()
            except Exception:
                db.session.rollback()
            existing = None
        else:
            return redirect(url_for("shift", id=existing.id, d=d))

    locked_opening = get_locked_opening_cash(day_obj, turn)
    u = current_user()
    default_responsible = u.username if u else ""

    if request.method == "POST":
        responsible = (request.form.get("responsible") or "").strip()
        opening_cash = safe_int(request.form.get("opening_cash"))

        if opening_cash is None:
            return redirect(url_for("open_shift", turn=turn, d=d))

        # Responsible can be omitted (or was imported wrong). Default to current user / Bernardo.
        if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
            responsible = default_responsible or "Bernardo"

        if locked_opening is not None and int(opening_cash) != int(locked_opening):
            return redirect(url_for("open_shift", turn=turn, d=d))

        s = Shift(day=day_obj, turn=turn, responsible=responsible, opening_cash=int(opening_cash))
        db.session.add(s)
        db.session.commit()
        backup_caja_local_y_drive()
        return redirect(url_for("shift", id=s.id, d=d))

    return render_template_string(
        OPEN_HTML,
        base_css=BASE_CSS,
        d=d,
        day=day_obj.isoformat(),
        turn_name=TURN_NAMES.get(turn, turn),
        locked_opening=locked_opening,
        default_responsible=default_responsible
    )

@app.route("/caja/shift/<int:id>")
@login_required
def shift(id):
    d = request.args.get("d") or date.today().isoformat()
    s = Shift.query.get_or_404(id)

    expenses = CashExpense.query.filter_by(shift_id=id).order_by(CashExpense.created_at.asc()).all()
    close_row = ShiftClose.query.filter_by(shift_id=id).first()

    egresos_total = expenses_total(id, CashExpense)
    cash_final = int(getattr(close_row, 'ending_cash', 0) or 0) if close_row else 0
    ingreso_bruto = calc_ingreso_bruto(s, egresos_total, cash_final=cash_final)
    ingreso_neto = calc_ingreso_neto(s, egresos=egresos_total, cash_final=cash_final)

    efectivo_disponible = int(s.sales_cash or 0)

    u = current_user()
    can_edit = bool(close_row) and can_edit_close(u, close_row)

    delivery_payload = build_delivery_payload(s)

    return render_template_string(
        SHIFT_HTML,
        base_css=BASE_CSS,
        s=s,
        expenses=expenses,
        close=close_row,
        categories=CATEGORIES,
        turn_name=TURN_NAMES.get(s.turn, s.turn),
        d=d,
        egresos_total=egresos_total,
        ingreso_neto=ingreso_neto,
        ingreso_bruto=ingreso_bruto,
        efectivo_disponible=efectivo_disponible,
        can_edit=can_edit,
        delivery_payload_json=json.dumps(delivery_payload),
        delivery_shift_presets_json=json.dumps(DELIVERY_SHIFT_PRESETS)
    )

@app.route("/caja/shift/<int:id>/delivery_draft", methods=["POST"])
@login_required
def save_delivery_draft(id):
    s = Shift.query.get_or_404(id)
    if s.status != "OPEN":
        return {"ok": False, "error": "Turno cerrado"}, 400

    payload = request.get_json(silent=True) or {}
    hour_shift = (payload.get("hour_shift") or "MORNING").strip().upper()
    if hour_shift not in DELIVERY_SHIFT_PRESETS:
        hour_shift = "MORNING"

    hour_in = (payload.get("hour_in") or "").strip()
    hour_out = (payload.get("hour_out") or "").strip()
    if hour_in and not valid_time_str(hour_in):
        return {"ok": False, "error": "Hora de entrada invalida"}, 400
    if hour_out and not valid_time_str(hour_out):
        return {"ok": False, "error": "Hora de salida invalida"}, 400

    rates = payload.get("rates") if isinstance(payload.get("rates"), list) else [1500, 2000, 2500, 3000, 2500]
    qtys = payload.get("qtys") if isinstance(payload.get("qtys"), list) else [0, 0, 0, 0, 0]
    rates = (rates + [0, 0, 0, 0, 0])[:5]
    qtys = (qtys + [0, 0, 0, 0, 0])[:5]

    safe_rates = []
    for v in rates:
        try:
            safe_rates.append(int(float(v or 0)))
        except Exception:
            safe_rates.append(0)

    hours_qty = delivery_hours_decimal(hour_in, hour_out)
    safe_qtys = []
    for i, v in enumerate(qtys):
        if i == 4:
            safe_qtys.append(hours_qty)
        else:
            try:
                safe_qtys.append(int(float(v or 0)))
            except Exception:
                safe_qtys.append(0)

    try:
        consume_amount = int(float(payload.get("consume_amount") or 0))
    except Exception:
        consume_amount = 0
    consume_note = str(payload.get("consume_note") or "")[:200]

    clean_payload = {
        "rates": safe_rates,
        "qtys": safe_qtys,
        "consume_amount": consume_amount,
        "consume_note": consume_note,
        "hour_shift": hour_shift,
        "hour_in": hour_in,
        "hour_out": hour_out,
    }

    s.hour_shift = hour_shift
    s.hour_in = hour_in or None
    s.hour_out = hour_out or None
    s.delivery_data_json = json.dumps(clean_payload, ensure_ascii=False)
    db.session.commit()
    backup_caja_local_y_drive()
    return {"ok": True, "hours": safe_qtys[4]}

@app.route("/caja/shift/<int:id>/sales", methods=["POST"])
@login_required
def sales(id):
    d = request.args.get("d") or date.today().isoformat()
    s = Shift.query.get_or_404(id)
    if s.status != "OPEN":
        return redirect(url_for("shift", id=id, d=d))

    s.sales_cash = to_int(request.form.get("sales_cash"))
    s.sales_mp = to_int(request.form.get("sales_mp"))
    s.sales_pya = to_int(request.form.get("sales_pya"))
    s.sales_rappi = to_int(request.form.get("sales_rappi"))

    db.session.commit()
    backup_caja_local_y_drive()
    return redirect(url_for("shift", id=id, d=d))

@app.route("/caja/shift/<int:id>/expense", methods=["POST"])
@login_required
def expense(id):
    d = request.args.get("d") or date.today().isoformat()
    s = Shift.query.get_or_404(id)
    if s.status != "OPEN":
        return redirect(url_for("shift", id=id, d=d))

    category = (request.form.get("category") or "").strip()
    amount = safe_int(request.form.get("amount"))
    note = (request.form.get("note") or "").strip()

    if category not in CATEGORIES or amount is None or amount <= 0:
        return redirect(url_for("shift", id=id, d=d))

    if category.startswith("Otros") and not note:
        return redirect(url_for("shift", id=id, d=d))

    e = CashExpense(shift_id=id, category=category, amount=int(amount), note=note or None)
    db.session.add(e)
    db.session.commit()
    backup_caja_local_y_drive()
    return redirect(url_for("shift", id=id, d=d))


@app.route("/caja/shift/<int:id>/close", methods=["POST"])
@login_required
def close_shift(id):
    d = request.args.get("d") or date.today().isoformat()
    s = Shift.query.get_or_404(id)
    if s.status != "OPEN":
        return redirect(url_for("shift", id=id, d=d))

    # Nuevo modelo:
    # - Retirado (efectivo total): se carga manualmente -> lo guardamos en Shift.sales_cash
    # - Caja final: se carga manualmente -> lo guardamos en ShiftClose.ending_cash
    retirado = safe_int(request.form.get("sales_cash"))
    caja_final = safe_int(request.form.get("cash_final"))

    if retirado is None or retirado < 0:
        return redirect(url_for("shift", id=id, d=d))
    if caja_final is None or caja_final < 0:
        return redirect(url_for("shift", id=id, d=d))

    # Guardar ventas (mismos nombres para no romper import/export)
    s.sales_cash = int(retirado)
    s.sales_mp = to_int(request.form.get("sales_mp"))
    s.sales_pya = to_int(request.form.get("sales_pya"))
    s.sales_rappi = to_int(request.form.get("sales_rappi"))

    # Cierre (sin estado OK/NO OK por ahora)
    ending_calc = calc_ending_calc(int(caja_final), int(retirado))
    ending_real = int(caja_final)
    difference = 0

    c = ShiftClose(
        shift_id=id,
        withdrawn_cash=int(retirado),
        ending_calc=int(ending_calc),
        ending_cash=int(ending_real),
        difference=int(difference),
        note=None,
        close_ok=1,
        edit_count=0
    )
    s.status = "CLOSED"
    s.closed_at = datetime.utcnow()
    db.session.add(c)
    db.session.commit()
    backup_caja_local_y_drive()
    return redirect(url_for("caja_index", d=d))




@app.route("/caja/shift/<int:id>/edit_all", methods=["GET","POST"])
@login_required
def edit_all(id):
    d = request.args.get("d") or date.today().isoformat()
    s = Shift.query.get_or_404(id)
    close_row = ShiftClose.query.filter_by(shift_id=id).first()
    if not close_row or s.status != "CLOSED":
        abort(404)

    u = current_user()
    if not can_edit_close(u, close_row):
        abort(403)

    expenses = CashExpense.query.filter_by(shift_id=id).order_by(CashExpense.created_at.asc()).all()

    if request.method == "POST":
        reason = (request.form.get("reason") or "").strip()
        if not reason:
            return redirect(url_for("edit_all", id=id, d=d))

        responsible = (request.form.get("responsible") or "").strip()
        opening_cash = safe_int(request.form.get("opening_cash"))
        if opening_cash is None:
            return redirect(url_for("edit_all", id=id, d=d))

        if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
            responsible = (u.username if u else None) or "Bernardo"

        s.responsible = responsible
        s.opening_cash = int(opening_cash)

        s.sales_cash = to_int(request.form.get("sales_cash"))
        s.sales_mp = to_int(request.form.get("sales_mp"))
        s.sales_pya = to_int(request.form.get("sales_pya"))
        s.sales_rappi = to_int(request.form.get("sales_rappi"))

        for e in expenses:
            if request.form.get(f"del_{e.id}") == "on":
                db.session.delete(e)
                continue

            cat = (request.form.get(f"cat_{e.id}") or "").strip()
            amt = safe_int(request.form.get(f"amt_{e.id}"))
            note = (request.form.get(f"note_{e.id}") or "").strip()

            if cat not in CATEGORIES:
                return redirect(url_for("edit_all", id=id, d=d))
            if amt is None or amt < 0:
                return redirect(url_for("edit_all", id=id, d=d))
            if cat.startswith("Otros") and not note:
                return redirect(url_for("edit_all", id=id, d=d))

            e.category = cat
            e.amount = int(amt)
            e.note = note or None

        # Agregar multiples egresos nuevos (hasta 5)
        for i in range(1, 6):
            new_cat = (request.form.get(f"new_category_{i}") or "").strip()
            new_amt = safe_int(request.form.get(f"new_amount_{i}"))
            new_note = (request.form.get(f"new_note_{i}") or "").strip()
            if not new_cat:
                continue
            if new_cat not in CATEGORIES:
                return redirect(url_for("edit_all", id=id, d=d))
            if new_amt is None or new_amt <= 0:
                return redirect(url_for("edit_all", id=id, d=d))
            if new_cat.startswith("Otros") and not new_note:
                return redirect(url_for("edit_all", id=id, d=d))
            db.session.add(CashExpense(
                shift_id=id,
                category=new_cat,
                amount=int(new_amt),
                note=new_note or None
            ))

        db.session.flush()

        withdrawn = safe_int(request.form.get("withdrawn"))
        ending_real = safe_int(request.form.get("ending_real"))

        if withdrawn is None or withdrawn < 0 or ending_real is None or ending_real < 0:
            return redirect(url_for("edit_all", id=id, d=d))

        # Nuevo modelo: caja final se carga manualmente
        ending_calc = calc_ending_calc(int(ending_real), int(withdrawn))
        close_ok = 1
        note_to_save = None
        difference = 0

        close_row.withdrawn_cash = int(withdrawn)
        close_row.ending_calc = int(ending_calc)
        close_row.ending_cash = int(ending_real)
        close_row.difference = int(difference)
        close_row.close_ok = close_ok
        close_row.note = note_to_save

        close_row.edit_count = int(close_row.edit_count or 0) + 1
        close_row.edited_by = u.username
        close_row.edited_at = datetime.utcnow()
        close_row.edit_reason = reason

        db.session.commit()
        backup_caja_local_y_drive()
        return redirect(url_for("shift", id=id, d=d))

    return render_template_string(
        EDIT_ALL_HTML,
        base_css=BASE_CSS,
        s=s,
        close=close_row,
        expenses=expenses,
        categories=CATEGORIES,
        user=u,
        d=d,
        turn_name=TURN_NAMES.get(s.turn, s.turn)
    )

@app.route("/caja/export/excel")
@login_required
def export_excel():
    data = get_caja_summary(
        db,
        Shift,
        ShiftClose,
        CashExpense,
        limit=10000,
        weekday_es=weekday_es,
        responsible_name=responsible_name,
        turn_names=TURN_NAMES,
        expenses_total=expenses_total,
        calc_ingreso_bruto=calc_ingreso_bruto,
        calc_ingreso_neto=calc_ingreso_neto,
        cash_bruto=cash_bruto,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cierres"

    ws.append([
        "Fecha","Turno","Responsable",
        "Caja Inicial","Efectivo bruto","Ventas Mercado Pago","Ventas Pedidos Ya","Ventas Rappi",
        "Egresos","Ventas Totales","Ventas Netas","Retirado","Caja Final (Real)","Diferencia"
    ])

    for r in data:
        ws.append([
            r["day"], r["turn_name"], r["responsible"],
            r["opening_cash"], r["sales_cash"], r["sales_mp"], r["sales_pya"], r["sales_rappi"],
            r["expenses"], r["ventas_bruto"], r["ventas_neto"],
            r["withdrawn"], r["ending_real"], r["difference"]
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"cierres_caja_{date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/caja/export/json")
@login_required
def export_caja_json():
    # Backup JSON de Caja: shifts + expenses + close
    start_raw = request.args.get("start")
    end_raw = request.args.get("end")
    start_date, end_date = parse_range_params(start_raw, end_raw)

    shifts = Shift.query.order_by(Shift.day.asc(), Shift.turn.asc()).all()
    expenses = CashExpense.query.order_by(CashExpense.created_at.asc()).all()
    closes = ShiftClose.query.order_by(ShiftClose.created_at.asc()).all()

    # si mandan start/end por query, filtramos (por day del shift)
    if start_raw or end_raw:
        shift_ids_in_range = [
            s.id for s in shifts
            if (s.day >= start_date and s.day <= end_date)
        ]
        shifts = [s for s in shifts if s.id in shift_ids_in_range]
        expenses = [e for e in expenses if e.shift_id in shift_ids_in_range]
        closes = [c for c in closes if c.shift_id in shift_ids_in_range]

    shift_lookup = {s.id: s for s in shifts}
    close_lookup = {c.shift_id: c for c in closes}

    shifts_payload = []
    for s in shifts:
        c = close_lookup.get(s.id)
        cash_final = int(c.ending_cash or 0) if c else 0
        shifts_payload.append({
            "day": s.day.isoformat(),
            "turn": s.turn,
            "responsible": s.responsible,
            "opening_cash": int(s.opening_cash or 0),
            "retirado": int(s.sales_cash or 0),
            "cash_final": cash_final,
            "sales_cash": int(cash_bruto(s, cash_final=cash_final)),
            "sales_mp": int(s.sales_mp or 0),
            "sales_pya": int(getattr(s, "sales_pya", 0) or 0),
            "sales_rappi": int(getattr(s, "sales_rappi", 0) or 0),
            "delivery_data_json": s.delivery_data_json,
            "hour_shift": s.hour_shift,
            "hour_in": s.hour_in,
            "hour_out": s.hour_out,
            "status": s.status,
            "closed_at": s.closed_at.isoformat() if s.closed_at else None,
        })

    expenses_payload = []
    for e in expenses:
        sh = shift_lookup.get(e.shift_id)
        expenses_payload.append({
            "shift_day": sh.day.isoformat() if sh else None,
            "shift_turn": sh.turn if sh else None,
            "category": e.category,
            "amount": int(e.amount or 0),
            "note": e.note,
            "created_at": e.created_at.isoformat() if e.created_at else None,
        })

    closes_payload = []
    for c in closes:
        sh = shift_lookup.get(c.shift_id)
        closes_payload.append({
            "shift_day": sh.day.isoformat() if sh else None,
            "shift_turn": sh.turn if sh else None,
            "withdrawn_cash": int(c.withdrawn_cash or 0),
            "ending_calc": int(c.ending_calc or 0),
            "ending_cash": int(c.ending_cash or 0),
            "difference": int(c.difference or 0),
            "note": c.note,
            "close_ok": int(c.close_ok or 1),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "edited_by": c.edited_by,
            "edited_at": c.edited_at.isoformat() if c.edited_at else None,
            "edit_reason": c.edit_reason,
            "edit_count": int(c.edit_count or 0),
        })

    payload = {
        "type": "caja_backup",
        "version": 1,
        "exported_at": datetime.utcnow().isoformat(),
        "range": {
            "start": start_date.isoformat() if (start_raw or end_raw) else None,
            "end": end_date.isoformat() if (start_raw or end_raw) else None
        },
        "shifts": shifts_payload,
        "expenses": expenses_payload,
        "closes": closes_payload,
    }

    bio = BytesIO()
    bio.write(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    bio.seek(0)
    filename = f"caja_backup_{date.today().isoformat()}.json"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/json")

# ==============================
# IMPORT CAJA (Opcion A)
# ==============================
IMPORT_CAJA_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Import Caja</title>{{ base_css|safe }}</head>
<body style="max-width:950px;">
  <a href="{{ url_for('caja_index') }}"><- Volver</a>
  <h2>Importar datos (Excel)</h2>

  <div class="box" style="margin-bottom:12px; border-color:#ffbf66;background:#fff0d6;">
    <b>Formato:</b> hojas <code>shifts</code> y <code>expenses</code>.<br>
    Se importan turnos y egresos, y el turno queda <b>CERRADO</b>.<br>
    <b>Regla aplicada:</b> Retirado = Efectivo neto (Efectivo bruto - Caja inicial).
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
    <div class="box">
      <label><b>Archivo Excel</b></label><br>
      <input type="file" name="file" accept=".xlsx" required>
      <br><br>
      <label><b>Modo</b></label><br>
      <select name="mode">
        <option value="skip" selected>Si existe el turno, NO tocarlo (skip)</option>
        <option value="replace">Si existe el turno, REEMPLAZAR (replace)</option>
      </select>
      <div class="muted" style="margin-top:6px;">Tip: para "pisar todo", usa modo <b>replace</b>.</div>
    </div>
    <br>
    <button class="btn">Importar</button>
  </form>
</body>
</html>
"""

def _excel_to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except:
        pass
    try:
        dd, mm, yy = s.split("/")
        return date(int(yy), int(mm), int(dd))
    except:
        return None

@app.route("/import/caja", methods=["GET","POST"])
@login_required
def import_caja():
    err = None
    msg = None

    if request.method == "POST":
        mode = (request.form.get("mode") or "skip").strip()
        f = request.files.get("file")
        if not f:
            err = "No se recibio archivo."
        else:
            try:
                wb = load_workbook(f, data_only=True)
                if "shifts" not in wb.sheetnames:
                    raise ValueError("Error: Falta la hoja 'shifts'.")
                if "expenses" not in wb.sheetnames:
                    raise ValueError("Error: Falta la hoja 'expenses'.")

                ws_s = wb["shifts"]
                ws_e = wb["expenses"]

                rows_s = list(ws_s.iter_rows(values_only=True))
                if len(rows_s) < 2:
                    raise ValueError("Error: La hoja 'shifts' no tiene datos.")

                headers_s = [str(x).strip() if x is not None else "" for x in rows_s[0]]
                need_s = ["day","turn","responsible","opening_cash","sales_cash","sales_mp","sales_pya","sales_rappi"]
                for n in need_s:
                    if n not in headers_s:
                        raise ValueError(f"Error: Falta columna '{n}' en shifts.")
                idxs = {k:i for i,k in enumerate(headers_s)}

                rows_e = list(ws_e.iter_rows(values_only=True))
                headers_e = [str(x).strip() if x is not None else "" for x in rows_e[0]] if rows_e else []
                need_e = ["day","turn","category","amount","note"]
                for n in need_e:
                    if n not in headers_e:
                        raise ValueError(f"Error: Falta columna '{n}' en expenses.")
                idxe = {k:i for i,k in enumerate(headers_e)}

                imported = 0
                replaced = 0

                exp_map = {}
                for r in rows_e[1:]:
                    if not r or all(v is None or str(v).strip()=="" for v in r):
                        continue
                    dd = _excel_to_date(r[idxe["day"]])
                    tt = (str(r[idxe["turn"]] or "").strip().upper())
                    if tt not in ("MORNING","AFTERNOON"):
                        continue
                    cat = (str(r[idxe["category"]] or "").strip())
                    amt = int(r[idxe["amount"]] or 0)
                    note = (str(r[idxe["note"]] or "").strip() or None)
                    exp_map.setdefault((dd, tt), []).append((cat, amt, note))

                for r in rows_s[1:]:
                    if not r or all(v is None or str(v).strip()=="" for v in r):
                        continue

                    dd = _excel_to_date(r[idxs["day"]])
                    tt = (str(r[idxs["turn"]] or "").strip().upper())
                    if dd is None:
                        continue
                    if tt not in ("MORNING","AFTERNOON"):
                        continue

                    responsible = str(r[idxs["responsible"]] or "").strip() or ""
                    opening_cash = int(r[idxs["opening_cash"]] or 0)
                    # Some historical imports had the "responsable" column shifted with opening cash.
                    if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
                        if opening_cash == 0:
                            try:
                                opening_cash = int(float(str(r[idxs["responsible"]] or 0).replace(",", ".")))
                            except Exception:
                                pass
                        responsible = "Bernardo"
                    sales_cash = int(r[idxs["sales_cash"]] or 0)
                    sales_mp = int(r[idxs["sales_mp"]] or 0)
                    sales_pya = int(r[idxs["sales_pya"]] or 0)
                    sales_rappi = int(r[idxs["sales_rappi"]] or 0)

                    existing = Shift.query.filter_by(day=dd, turn=tt).first()
                    if existing:
                        if mode == "skip":
                            continue
                        c = ShiftClose.query.filter_by(shift_id=existing.id).first()
                        if c:
                            db.session.delete(c)
                        CashExpense.query.filter_by(shift_id=existing.id).delete()
                        db.session.delete(existing)
                        db.session.flush()
                        replaced += 1

                    s = Shift(
                        day=dd, turn=tt, responsible=responsible,
                        opening_cash=opening_cash,
                        sales_cash=sales_cash, sales_mp=sales_mp,
                        sales_pya=sales_pya, sales_rappi=sales_rappi,
                        delivery_data_json=None,
                        hour_shift=None,
                        hour_in=None,
                        hour_out=None,
                        status="CLOSED", closed_at=datetime.utcnow()
                    )
                    db.session.add(s)
                    db.session.flush()

                    for (cat, amt, note) in exp_map.get((dd, tt), []):
                        if cat not in CATEGORIES:
                            cat = "Delivery / Cadete (efectivo)"
                        if cat.startswith("Otros") and not note:
                            note = "import"
                        if amt and amt > 0:
                            db.session.add(CashExpense(shift_id=s.id, category=cat, amount=int(amt), note=note))

                    withdrawn = max(0, int(sales_cash) - int(opening_cash))
                    ending_calc = calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn)
                    ending_real = ending_calc
                    diff = 0

                    db.session.add(ShiftClose(
                        shift_id=s.id,
                        withdrawn_cash=withdrawn,
                        ending_calc=ending_calc,
                        ending_cash=ending_real,
                        difference=diff,
                        note=None,
                        close_ok=1,
                        edit_count=0
                    ))

                    imported += 1

                db.session.commit()
                backup_caja_local_y_drive()
                msg = f"Import OK. Turnos importados: {imported}. Reemplazados: {replaced}."

            except Exception as ex:
                db.session.rollback()
                err = str(ex)

    return render_template_string(IMPORT_CAJA_HTML, base_css=BASE_CSS, err=err, msg=msg)

IMPORT_CAJA_JSON_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Import Caja (JSON)</title>{{ base_css|safe }}</head>
<body style="max-width:950px;">
  <a href="{{ url_for('caja_index') }}"><- Volver</a>
  <h2>Importar Backup JSON - Caja</h2>

  <div class="box" style="margin-bottom:12px;">
    <b>Formato:</b> archivo JSON exportado por "Exportar JSON".<br>
    <b>Modo skip:</b> si el turno (dia+turno) ya existe, no lo toca.<br>
    <b>Modo replace:</b> borra ese turno (incluye egresos + cierre) y lo recrea.
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
    <div class="box">
      <label><b>Archivo JSON</b></label><br>
      <input type="file" name="file" accept=".json" required>
      <br><br>
      <label><b>Modo</b></label><br>
      <select name="mode">
        <option value="skip" selected>Si existe el turno, NO tocarlo (skip)</option>
        <option value="replace">Si existe el turno, REEMPLAZAR (replace)</option>
      </select>
    </div>
    <br>
    <button class="btn">Importar JSON</button>
  </form>
</body>
</html>
"""

@app.route("/import/caja/json", methods=["GET","POST"])
@login_required
def import_caja_json():
    err = None
    msg = None

    if request.method == "POST":
        mode = (request.form.get("mode") or "skip").strip()
        f = request.files.get("file")
        if not f:
            err = "No se recibio archivo."
        else:
            try:
                payload = json.load(f)

                if not isinstance(payload, dict) or payload.get("type") != "caja_backup":
                    raise ValueError("JSON invalido: no parece un backup de Caja.")

                shifts = payload.get("shifts") or []
                expenses = payload.get("expenses") or []
                closes = payload.get("closes") or []

                # index de expenses/closes por (day, turn)
                exp_map = {}
                for e in expenses:
                    dd = (e.get("shift_day") or "").strip()
                    tt = (e.get("shift_turn") or "").strip().upper()
                    if not dd or tt not in ("MORNING","AFTERNOON"):
                        continue
                    exp_map.setdefault((dd, tt), []).append(e)

                close_map = {}
                for c in closes:
                    dd = (c.get("shift_day") or "").strip()
                    tt = (c.get("shift_turn") or "").strip().upper()
                    if not dd or tt not in ("MORNING","AFTERNOON"):
                        continue
                    close_map[(dd, tt)] = c

                imported = 0
                replaced = 0
                skipped = 0

                for s in shifts:
                    dd = (s.get("day") or "").strip()
                    tt = (s.get("turn") or "").strip().upper()
                    if not dd or tt not in ("MORNING","AFTERNOON"):
                        continue

                    day_obj = date.fromisoformat(dd)
                    responsible = (s.get("responsible") or "").strip()
                    opening_cash = int(s.get("opening_cash") or 0)
                    if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", str(responsible)):
                        if opening_cash == 0:
                            try:
                                opening_cash = int(float(str(s.get("responsible") or 0).replace(",", ".")))
                            except Exception:
                                pass
                        responsible = "Bernardo"
                    sales_cash = int(s.get("sales_cash") or 0)
                    sales_mp = int(s.get("sales_mp") or 0)
                    sales_pya = int(s.get("sales_pya") or 0)
                    sales_rappi = int(s.get("sales_rappi") or 0)

                    existing = Shift.query.filter_by(day=day_obj, turn=tt).first()
                    if existing:
                        if mode == "skip":
                            skipped += 1
                            continue

                        # replace: borrar cierre + egresos + shift
                        c_old = ShiftClose.query.filter_by(shift_id=existing.id).first()
                        if c_old:
                            db.session.delete(c_old)
                        CashExpense.query.filter_by(shift_id=existing.id).delete()
                        db.session.delete(existing)
                        db.session.flush()
                        replaced += 1

                    # crear shift
                    shift_status = (s.get("status") or "CLOSED").strip().upper()
                    if shift_status not in ("OPEN","CLOSED"):
                        shift_status = "CLOSED"

                    new_shift = Shift(
                        day=day_obj,
                        turn=tt,
                        responsible=responsible,
                        opening_cash=opening_cash,
                        sales_cash=sales_cash,
                        sales_mp=sales_mp,
                        sales_pya=sales_pya,
                        sales_rappi=sales_rappi,
                        delivery_data_json=s.get("delivery_data_json"),
                        hour_shift=s.get("hour_shift"),
                        hour_in=s.get("hour_in"),
                        hour_out=s.get("hour_out"),
                        status=shift_status,
                        closed_at=datetime.utcnow() if shift_status == "CLOSED" else None
                    )
                    db.session.add(new_shift)
                    db.session.flush()

                    # egresos
                    for e in exp_map.get((dd, tt), []):
                        cat = (e.get("category") or "").strip() or "Mantenimiento / Varios"
                        amt = int(e.get("amount") or 0)
                        note = (e.get("note") or None)

                        if cat not in CATEGORIES:
                            cat = "Mantenimiento / Varios"
                        if cat.startswith("Otros") and not note:
                            note = "import json"

                        if amt > 0:
                            db.session.add(CashExpense(
                                shift_id=new_shift.id,
                                category=cat,
                                amount=amt,
                                note=note
                            ))

                    # cierre (si existe en JSON). Si no, lo calculamos.
                    cdata = close_map.get((dd, tt))
                    if shift_status == "CLOSED":
                        if cdata:
                            withdrawn = int(cdata.get("withdrawn_cash") or 0)
                            ending_calc = int(cdata.get("ending_calc") or calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn))
                            ending_real = int(cdata.get("ending_cash") or ending_calc)
                            diff = int(cdata.get("difference") or (ending_real - ending_calc))
                            close_ok = int(cdata.get("close_ok") or 1)
                            note = cdata.get("note") or None

                            db.session.add(ShiftClose(
                                shift_id=new_shift.id,
                                withdrawn_cash=withdrawn,
                                ending_calc=ending_calc,
                                ending_cash=ending_real,
                                difference=diff,
                                note=note,
                                close_ok=close_ok,
                                edit_count=int(cdata.get("edit_count") or 0),
                                edited_by=cdata.get("edited_by"),
                                edited_at=datetime.fromisoformat(cdata["edited_at"]) if cdata.get("edited_at") else None,
                                edit_reason=cdata.get("edit_reason"),
                            ))
                        else:
                            withdrawn = max(0, int(sales_cash) - int(opening_cash))
                            ending_calc = calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn)
                            ending_real = ending_calc
                            db.session.add(ShiftClose(
                                shift_id=new_shift.id,
                                withdrawn_cash=withdrawn,
                                ending_calc=ending_calc,
                                ending_cash=ending_real,
                                difference=0,
                                note=None,
                                close_ok=1,
                                edit_count=0
                            ))

                    imported += 1

                db.session.commit()
                backup_caja_local_y_drive()
                msg = f"Import JSON OK. Importados: {imported}. Reemplazados: {replaced}. Omitidos (skip): {skipped}."
            except Exception as ex:
                db.session.rollback()
                err = str(ex)

    return render_template_string(IMPORT_CAJA_JSON_HTML, base_css=BASE_CSS, err=err, msg=msg)

# ============================================================
# =====================  ASISTENCIA (UI)  =====================
# ============================================================

ATT_CONFIG_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Config Asistencia</title>{{ base_css|safe }}</head>
<body style="max-width:1000px;">
  <a href="{{ url_for('asistencia') }}"><- Volver</a>
  <h2>Configurar grupos / vacaciones</h2>

  <div class="box">
    <b>Horarios por grupo</b><br>
    <span class="muted">
      <b>Grupo A:</b> {{ga.morning_in}}-{{ga.morning_out}} / {{ga.afternoon_in}}-{{ga.afternoon_out}}<br>
      <b>Grupo B:</b> {{gb.morning_in}}-{{gb.morning_out}} / {{gb.afternoon_in}}-{{gb.afternoon_out}}
    </span>
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <h3>Rotacion semanal (A/B)</h3>
  <div class="box">
    <p class="muted">
      Rotacion automatica semanal (AB).<br>
      {% if cfg_exists %}
        <b>Edicion:</b> solo Admin puede modificar configuracion existente.
      {% else %}
        <b>Inicial:</b> cualquier usuario puede configurar por primera vez.
      {% endif %}
    </p>
    <form method="post">
      <input type="hidden" name="action" value="set_rotation">
      <table>
        <tr><th>Empleado</th><th>Grupo (Semana base)</th></tr>
        {% for e in employees %}
        <tr>
          <td><b>{{e}}</b></td>
          <td>
            <select name="g_{{e}}" {% if lock_rotation %}disabled{% endif %}>
              <option value="A" {% if week0.get(e,'A')=='A' %}selected{% endif %}>A</option>
              <option value="B" {% if week0.get(e,'A')=='B' %}selected{% endif %}>B</option>
            </select>
          </td>
        </tr>
        {% endfor %}
      </table>
      <br>
      <button class="btn" {% if lock_rotation %}disabled{% endif %}>Guardar rotacion</button>
      {% if lock_rotation %}
        <div class="muted">Solo Admin puede editar esta configuracion.</div>
      {% endif %}
    </form>
  </div>

  <h3>Vacaciones (pre-carga)</h3>
  <div class="box">
    <form method="post" style="display:flex; gap:10px; flex-wrap:wrap; align-items:end;">
      <input type="hidden" name="action" value="add_vac">
      <div>
        <label class="muted">Empleado</label><br>
        <select name="emp">
          {% for e in employees %}<option value="{{e}}">{{e}}</option>{% endfor %}
        </select>
      </div>
      <div>
        <label class="muted">Desde</label><br>
        <input type="date" name="start" required>
      </div>
      <div>
        <label class="muted">Hasta</label><br>
        <input type="date" name="end" required>
      </div>
      <div>
        <button class="btn">Agregar</button>
      </div>
    </form>

    <table style="margin-top:12px;">
      <tr><th>Empleado</th><th>Desde</th><th>Hasta</th><th>Accion</th></tr>
      {% for v in vacations %}
        <tr>
          <td>{{v.employee}}</td>
          <td>{{v.start_day}}</td>
          <td>{{v.end_day}}</td>
          <td>
            <form method="post" style="display:inline;">
              <input type="hidden" name="action" value="del_vac">
              <input type="hidden" name="id" value="{{v.id}}">
              <button class="btn">Borrar</button>
            </form>
          </td>
        </tr>
      {% endfor %}
      {% if not vacations %}
        <tr><td colspan="4" class="muted">Sin vacaciones configuradas.</td></tr>
      {% endif %}
    </table>
  </div>

</body>
</html>
"""

@app.route("/asistencia/config", methods=["GET","POST"])
@login_required
def att_config():
    u = current_user()
    err = None
    msg = None

    cfg = rotation_config_get()
    cfg_exists = bool(cfg and cfg.week0_map)
    week0 = parse_week0_map(cfg.week0_map) if cfg_exists else {}

    lock_rotation = cfg_exists and (u.role != "admin")

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        try:
            if action == "set_rotation":
                if lock_rotation:
                    raise ValueError("Solo Admin puede editar la rotacion una vez creada.")
                m = {}
                for e in EMPLOYEES:
                    g = (request.form.get(f"g_{e}") or "A").strip().upper()
                    if g not in ("A","B"):
                        g = "A"
                    m[e] = g
                if not cfg:
                    cfg = RotationConfig(week0_map=make_week0_map_str(m), created_by=u.username)
                    db.session.add(cfg)
                else:
                    cfg.week0_map = make_week0_map_str(m)
                    cfg.created_by = cfg.created_by or u.username
                db.session.commit()
                backup_caja_local_y_drive()
                msg = "Rotacion guardada."
            elif action == "add_vac":
                emp = (request.form.get("emp") or "").strip()
                start = date.fromisoformat(request.form.get("start"))
                end = date.fromisoformat(request.form.get("end"))
                if emp not in EMPLOYEES:
                    raise ValueError("Empleado invalido.")
                if start > end:
                    start, end = end, start
                db.session.add(Vacation(employee=emp, start_day=start, end_day=end))
                db.session.commit()
                backup_caja_local_y_drive()
                msg = "Vacaciones agregadas."
            elif action == "del_vac":
                vid = int(request.form.get("id") or 0)
                v = Vacation.query.get(vid)
                if v:
                    db.session.delete(v)
                    db.session.commit()
                    backup_caja_local_y_drive()
                msg = "Vacaciones borradas."
        except Exception as ex:
            db.session.rollback()
            err = str(ex)

    vacations = Vacation.query.order_by(Vacation.employee.asc(), Vacation.start_day.asc()).all()

    return render_template_string(
        ATT_CONFIG_HTML,
        base_css=BASE_CSS,
        ga=GROUP_A, gb=GROUP_B,
        employees=EMPLOYEES,
        cfg_exists=cfg_exists,
        lock_rotation=lock_rotation,
        week0=week0,
        vacations=vacations,
        err=err,
        msg=msg
    )

ASISTENCIA_HTML = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Asistencia</title>
{{ base_css|safe }}
<style>
  body{max-width:1500px;}
  th,td{padding:8px;}
  .grid{display:flex; gap:10px; flex-wrap:wrap; align-items:end; margin:12px 0;}
  details{background:#fafafa;border:1px solid #ddd;border-radius:10px;padding:8px;}
  summary{cursor:pointer;font-weight:bold;}
  .w-time{width:110px;}
  .w-emp{width:120px;}
  .w-group{width:120px;}
  .w-nov{width:220px;}
  .w-notes{width:260px;}
  .filters{display:flex; gap:10px; flex-wrap:wrap; align-items:end; margin:12px 0;}
  .holiday-bg{background:#ffe3ec;}
  .holiday-bg table tr{background:#fff;}
</style>
</head>
<body class="{{ 'holiday-bg' if hday else '' }}">

<div class="top">
  <div>
    <h2 style="margin:0;">Asistencia</h2>
    <div class="muted"><a href="{{ url_for('home') }}"><- Volver al menu</a></div>
  </div>
    <div class="row">
      <a class="btn" href="{{ url_for('att_export_excel', emp=emp, start=start, end=end, novf=novf) }}">Exportar Excel</a>
      <a class="btn" href="{{ url_for('att_export_json', emp=emp, start=start, end=end, novf=novf) }}">Exportar JSON</a>

      <a class="btn" href="{{ url_for('import_asistencia') }}">Importar Excel</a>
      <a class="btn" href="{{ url_for('import_asistencia_json') }}">Importar JSON</a>

      <a class="btn" href="{{ url_for('att_config') }}">Configurar grupos / vacaciones</a>
    </div>
</div>

<div class="box" style="margin-top:10px;">
  <b>Horarios por grupo</b><br>
  <span class="muted">
    <b>Grupo A:</b> {{ga.morning_in}}-{{ga.morning_out}} / {{ga.afternoon_in}}-{{ga.afternoon_out}} &nbsp;&nbsp;|&nbsp;&nbsp;
    <b>Grupo B:</b> {{gb.morning_in}}-{{gb.morning_out}} / {{gb.afternoon_in}}-{{gb.afternoon_out}}
  </span>
</div>

<form method="get" action="{{ url_for('asistencia') }}" class="grid">
  <div>
    <label class="muted">Fecha:</label><br>
    <input type="date" name="d" value="{{d}}">
  </div>
  <div>
    <button class="btn" type="submit">Ir</button>
  </div>
</form>

<form method="post" action="{{ url_for('asistencia') }}" class="grid" style="margin-top:6px;">
  <input type="hidden" name="action" value="set_holiday">
  <input type="hidden" name="d" value="{{d}}">
  <div>
    <label class="muted">Feriado:</label><br>
    <select name="hday" style="min-width:220px;">
      <option value="" {% if hday=='' %}selected{% endif %}>-</option>
      <option value="LABORABLE" {% if hday=='LABORABLE' %}selected{% endif %}>Feriado Laborable</option>
      <option value="NO_LABORABLE" {% if hday=='NO_LABORABLE' %}selected{% endif %}>Feriado No laborable (se paga el dia)</option>
    </select>
  </div>
  <div>
    <button class="btn" type="submit">Aplicar</button>
  </div>
</form>

{% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}
{% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}

<form method="post">
<input type="hidden" name="d" value="{{d}}">

<table>
  <tr>
    <th class="w-emp">Empleado</th>
    <th class="w-group">Grupo</th>
    <th colspan="3">Manana</th>
    <th colspan="3">Tarde</th>
    <th>Horas (paga)</th>
    <th class="w-nov">Novedad</th>
    <th class="w-notes">Notas</th>
    <th>Consumos</th>
  </tr>
  <tr>
    <th></th>
    <th></th>
    <th class="w-time">Ing</th><th class="w-time">Egr</th><th>Horas</th>
    <th class="w-time">Ing</th><th class="w-time">Egr</th><th>Horas</th>
    <th></th>
    <th></th><th></th><th></th>
  </tr>

  {% for r in rows %}
  <tr class="{% if r.is_vac %}vac-row{% elif r.has_warning %}warn-row{% endif %}">
    <td>
      <b>{{r.employee}}</b><br>
      {% if r.chip %}
        <span class="chip {{r.chip_class}}">{{r.chip}}</span>
      {% endif %}
    </td>

    <td>
      <select name="{{r.key}}_mode">
        <option value="AUTO" {% if r.mode=='AUTO' %}selected{% endif %}>Auto</option>
        <option value="MANUAL" {% if r.mode=='MANUAL' %}selected{% endif %}>Manual</option>
      </select>
      <div class="muted">Sug: {{r.group}}</div>
      <input type="hidden" name="{{r.key}}_group" value="{{r.group}}">
    </td>

    <td><input class="w-time {{'preset' if r.mi_preset else ''}}" type="time" name="{{r.key}}_mi" value="{{r.morning_in}}" {% if r.lock_times %}readonly{% endif %}></td>
    <td><input class="w-time {{'preset' if r.mo_preset else ''}}" type="time" name="{{r.key}}_mo" value="{{r.morning_out}}" {% if r.lock_times %}readonly{% endif %}></td>
    <td>{{r.morning_time}}</td>

    <td><input class="w-time {{'preset' if r.ai_preset else ''}}" type="time" name="{{r.key}}_ai" value="{{r.afternoon_in}}" {% if r.lock_times %}readonly{% endif %}></td>
    <td><input class="w-time {{'preset' if r.ao_preset else ''}}" type="time" name="{{r.key}}_ao" value="{{r.afternoon_out}}" {% if r.lock_times %}readonly{% endif %}></td>
    <td>{{r.afternoon_time}}</td>

    <td><b>{{r.payable_time}}</b><div class="muted">{{r.warn_text}}</div></td>

    <td>
      <select name="{{r.key}}_nov" onchange="toggleNovelty('{{r.key}}');">
        {% for it in novelty_items %}
          <option value="{{it}}" {% if r.novelty==it %}selected{% endif %}>{{it if it else "-"}}</option>
        {% endfor %}
      </select>

      <div id="{{r.key}}_nov_extra" style="margin-top:6px; display:none;">
        <div class="muted" id="{{r.key}}_nov_hint"></div>
        <input type="text" name="{{r.key}}_novx" id="{{r.key}}_novx" style="width:140px;" placeholder="">
      </div>

      <span data-key="{{r.key}}" data-nov="{{r.novelty}}" data-novm="{{r.novelty_minutes}}" style="display:none;"></span>
    </td>

    <td><input style="width:100%;" name="{{r.key}}_notes" value="{{r.notes or ''}}" placeholder="detalle opcional"></td>

    <td>
      <details>
        <summary>Ver / Editar</summary>
        <div class="muted" style="margin-top:8px;">Cantidad:</div>
        <select name="{{r.key}}_cq" id="{{r.key}}_cq" onchange="setRows('{{r.key}}')">
          {% for i in range(0, maxc+1) %}
            <option value="{{i}}" {% if r.cons_count==i %}selected{% endif %}>{{i}}</option>
          {% endfor %}
        </select>

        <table style="margin-top:8px;">
          <tr><th>#</th><th>Consumo</th><th>Monto</th></tr>
          {% for i in range(1, maxc+1) %}
          <tr class="{{r.key}}_row" data-i="{{i}}">
            <td>{{i}}</td>
            <td><input style="width:260px;" type="text" name="{{r.key}}_ci{{i}}" value="{{ r.cons[i-1].item if r.cons|length>=i else '' }}"></td>
            <td><input style="width:120px;" type="number" min="0" name="{{r.key}}_cm{{i}}" value="{{ r.cons[i-1].amount if r.cons|length>=i and r.cons[i-1].amount is not none else '' }}"></td>
          </tr>
          {% endfor %}
        </table>
      </details>
    </td>
  </tr>
  {% endfor %}
</table>

<br>
<button class="btn">Guardar</button>
</form>

<hr>

<h2>Resumen / Filtros</h2>

<div class="row" style="margin:10px 0; gap:10px;">
  <div class="box" style="min-width:220px;"><b>Horas totales:</b> {{ sum_hours_txt }}</div>
  <div class="box" style="min-width:220px;"><b>Consumos totales:</b> {{ sum_cons_money|money }}</div>
</div>


<form method="get" action="{{ url_for('asistencia') }}" class="filters">
  <input type="hidden" name="d" value="{{d}}">

  <div>
    <label class="muted">Empleado:</label><br>
    <select name="emp">
      <option value="ALL" {% if emp=='ALL' %}selected{% endif %}>Todos</option>
      {% for e in employees %}
        <option value="{{e}}" {% if emp==e %}selected{% endif %}>{{e}}</option>
      {% endfor %}
    </select>
  </div>

  <div>
    <label class="muted">Desde:</label><br>
    <input type="date" name="start" value="{{start}}">
  </div>

  <div>
    <label class="muted">Hasta:</label><br>
    <input type="date" name="end" value="{{end}}">
  </div>

  <div>
    <label class="muted">Novedad:</label><br>
    <select name="novf">
      <option value="ALL" {% if novf=='ALL' %}selected{% endif %}>Todas</option>
      {% for it in novelty_items %}
        {% if it %}
          <option value="{{it}}" {% if novf==it %}selected{% endif %}>{{it}}</option>
        {% endif %}
      {% endfor %}
    </select>
  </div>

  <div>
    <button class="btn" type="submit">Aplicar</button>
  </div>
</form>

<table>
  <tr>
    <th>Fecha</th>
    <th>Empleado</th>
    <th>Grupo</th>
    <th>Manana</th>
    <th>Tarde</th>
    <th>Horas pagas</th>
    <th>Novedad</th>
    <th>Notas</th>
    <th>Consumos</th>
  </tr>
  {% for r in summary %}
  <tr>
    <td>{{r.day}}</td>
    <td>{{r.employee}}</td>
    <td>{{r.group}}</td>
    <td>{{r.morning}}</td>
    <td>{{r.afternoon}}</td>
    <td><b>{{r.payable}}</b></td>
    <td>{{r.novelty}}</td>
    <td>{{r.notes}}</td>
    <td>
      <b>{{r.cons_total|money}}</b>
      {% if r.cons_items %}
        <div class="muted">{{r.cons_items}}</div>
      {% endif %}
    </td>
  </tr>
  {% endfor %}
  {% if not summary %}
    <tr><td colspan="9" class="muted">Sin registros en el rango.</td></tr>
  {% endif %}
</table>

<script>
  function setRows(key){
    const n = parseInt(document.getElementById(key + "_cq").value || "0", 10);
    const rows = document.querySelectorAll("tr." + key + "_row");
    rows.forEach(r => {
      const i = parseInt(r.getAttribute("data-i"), 10);
      r.style.display = (i <= n) ? "table-row" : "none";
    });
  }

  function toggleNovelty(key){
    const sel = document.querySelector("select[name='"+key+"_nov']");
    const box = document.getElementById(key+"_nov_extra");
    const hint = document.getElementById(key+"_nov_hint");
    const inp = document.getElementById(key+"_novx");

    const v = sel.value || "";
    if (v === "Razones particulares") {
      box.style.display = "block";
      hint.textContent = "Cargar minutos tomados hoy (ej: 30).";
      inp.placeholder = "minutos";
      inp.disabled = false;
    } else if (v === "Enfermedad") {
      box.style.display = "block";
      hint.textContent = "Cargar horas de enfermedad hoy (ej: 2.5).";
      inp.placeholder = "horas";
      inp.disabled = false;
    } else {
      box.style.display = "none";
      inp.disabled = true;
      inp.value = "";
    }
  }

  window.addEventListener("load", () => {
    document.querySelectorAll("select[id$='_cq']").forEach(sel => {
      const key = sel.id.replace("_cq","");
      setRows(key);
    });

    document.querySelectorAll("select[name$='_nov']").forEach(sel => {
      const key = sel.name.replace("_nov","");
      toggleNovelty(key);
    });

    document.querySelectorAll("span[data-novm]").forEach(sp => {
      const key = sp.getAttribute("data-key");
      const nov = sp.getAttribute("data-nov") || "";
      const novm = parseInt(sp.getAttribute("data-novm") || "0", 10);
      const inp = document.getElementById(key+"_novx");
      if (!inp) return;

      if (nov === "Razones particulares") {
        inp.value = String(novm || 0);
      } else if (nov === "Enfermedad") {
        inp.value = String((novm || 0) / 60.0);
      }
    });
  });
</script>

</body>
</html>
"""

def attendance_summary_rows(start: date, end: date, employee: str, novf: str):
    q = Attendance.query.filter(Attendance.day >= start, Attendance.day <= end)
    if employee and employee != "ALL":
        q = q.filter(Attendance.employee == employee)
    # NO filtramos por novf aca: Vacaciones puede venir de tabla Vacation
    q = q.order_by(Attendance.day.desc(), Attendance.employee.asc())

    rows = []
    for a in q.all():
        vac = is_vacation(a.employee, a.day) or (a.novelty == "Vacaciones")
        novelty_display = "Vacaciones" if vac else (a.novelty or "")

        if novf and novf != "ALL":
            if novelty_display != novf:
                continue

        g = a.group_code or group_for_employee_on_day(a.employee, a.day)
        exp = expected_times_for_group(g, a.day)

        mi = a.morning_in or exp["morning_in"]
        mo = a.morning_out or exp["morning_out"]
        ai = a.afternoon_in or exp["afternoon_in"]
        ao = a.afternoon_out or exp["afternoon_out"]

        calc = compute_work_minutes_and_flags(a)
        cons_total, cons_items = consumptions_summary_for_attendance(a.id)

        rows.append({
            "day": a.day.isoformat(),
            "employee": a.employee,
            "group": g,
            "morning": f"{mi}-{mo}".strip("-"),
            "afternoon": f"{ai}-{ao}".strip("-"),
            "payable": fmt_minutes(calc["payable_min"]),
            "novelty": novelty_display,
            "notes": a.notes or "",
            "cons_total": cons_total,
            "cons_items": cons_items,
        })
    return rows

@app.route("/asistencia", methods=["GET","POST"])
@login_required
def asistencia():
    day_obj, d = parse_day_param(request.args.get("d") or request.form.get("d"))

    err = None
    msg = None

    # filtros de resumen
    empf = request.args.get("emp") or "ALL"
    novf = request.args.get("novf") or "ALL"
    start_date, end_date = parse_range_params(request.args.get("start"), request.args.get("end"))
    start = start_date.isoformat()
    end = end_date.isoformat()

    # feriado guardado para el dia
    hday = get_holiday_type(day_obj)

    # ======= Aplicar feriado (POST) =======
    if request.method == "POST" and (request.form.get("action") or "").strip() == "set_holiday":
        try:
            h = (request.form.get("hday") or "").strip().upper()
            if h not in ("", "LABORABLE", "NO_LABORABLE"):
                h = ""
            set_holiday_type(day_obj, h)
            db.session.commit()
            backup_caja_local_y_drive()
            msg = "Feriado aplicado."
            return redirect(url_for("asistencia", d=d, emp=empf, start=start, end=end, novf=novf))
        except Exception as ex:
            db.session.rollback()
            err = str(ex)

    # ======= Guardar asistencia (POST) =======
    if request.method == "POST" and err is None:
        try:
            for emp_name in EMPLOYEES:
                key = emp_key(emp_name)
                row = Attendance.query.filter_by(day=day_obj, employee=emp_name).first()
                if not row:
                    row = Attendance(day=day_obj, employee=emp_name, mode="AUTO")
                    db.session.add(row)
                    db.session.flush()

                mode = (request.form.get(f"{key}_mode") or "AUTO").strip().upper()
                if mode not in ("AUTO","MANUAL"):
                    mode = "AUTO"

                g = (request.form.get(f"{key}_group") or "").strip().upper()
                if g not in ("A","B"):
                    g = group_for_employee_on_day(emp_name, day_obj)

                row.mode = mode
                row.group_code = g

                exp = expected_times_for_group(g, day_obj)
                vac = is_vacation(emp_name, day_obj)

                nov = (request.form.get(f"{key}_nov") or "").strip()
                if nov not in NOVELTY_ITEMS:
                    nov = ""
                row.novelty = nov or None

                def get_post_time(field_name: str) -> Optional[str]:
                    v = (request.form.get(field_name) or "").strip()
                    if v and not valid_time_str(v):
                        raise ValueError(f"Hora invalida en {emp_name}: {v}")
                    return v or None

                mi_post = get_post_time(f"{key}_mi")
                mo_post = get_post_time(f"{key}_mo")
                ai_post = get_post_time(f"{key}_ai")
                ao_post = get_post_time(f"{key}_ao")
                if vac or nov == "Vacaciones":
                    # Vacaciones: se fija el horario esperado completo (segun grupo/dia)
                    row.morning_in = exp.get("morning_in")
                    row.morning_out = exp.get("morning_out")
                    row.afternoon_in = exp.get("afternoon_in")
                    row.afternoon_out = exp.get("afternoon_out")

                elif nov == "Delivery":
                    # Delivery: pago aparte, no computa horas. No permitir carga de horarios.
                    row.morning_in = None
                    row.morning_out = None
                    row.afternoon_in = None
                    row.afternoon_out = None

                else:
                    if mode == "AUTO":
                        # Auto: precarga valores esperados como fallback (gris), y se sobreescriben si se editan.
                        row.morning_in = mi_post or row.morning_in or exp.get("morning_in")
                        row.morning_out = mo_post or row.morning_out or exp.get("morning_out")
                        row.afternoon_in = ai_post or row.afternoon_in or exp.get("afternoon_in")
                        row.afternoon_out = ao_post or row.afternoon_out or exp.get("afternoon_out")
                    else:
                        # Manual: se guarda exactamente lo cargado (puede quedar vacio)
                        row.morning_in = mi_post
                        row.morning_out = mo_post
                        row.afternoon_in = ai_post
                        row.afternoon_out = ao_post


                novx = (request.form.get(f"{key}_novx") or "").strip()
                if nov == "Razones particulares":
                    m = safe_int(novx) or 0
                    row.novelty_minutes = max(0, int(m))
                elif nov == "Enfermedad":
                    h = safe_float(novx) or 0.0
                    mins = int(round(max(0.0, h) * 60.0))
                    row.novelty_minutes = mins
                else:
                    row.novelty_minutes = 0

                row.notes = (request.form.get(f"{key}_notes") or "").strip() or None

                # Validaciones de jornada (solo si no es vacaciones/inasistencia)
                calc = compute_work_minutes_and_flags(row)
                jornada_min = jornada_min_for_day(day_obj)

                if not (vac or nov == "Vacaciones" or nov == "Inasistencia"):
                    if calc["total_worked_min"] < jornada_min and not (row.novelty or ""):
                        raise ValueError(
                            f"{emp_name}: si trabajo menos de {fmt_minutes(jornada_min)} debe cargar Novedad."
                        )
                    if calc["total_worked_min"] > jornada_min and not (row.notes and row.notes.strip()):
                        raise ValueError(
                            f"{emp_name}: si trabajo mas de {fmt_minutes(jornada_min)} debe justificar en Notas."
                        )

                db.session.flush()

                # Consumptions (recrear)
                AttendanceConsumption.query.filter_by(attendance_id=row.id).delete()
                qty = safe_int(request.form.get(f"{key}_cq")) or 0
                qty = max(0, min(qty, MAX_CONSUMOS))
                for i in range(1, qty + 1):
                    item = (request.form.get(f"{key}_ci{i}") or "").strip() or None
                    amount = safe_int(request.form.get(f"{key}_cm{i}"))
                    db.session.add(AttendanceConsumption(attendance_id=row.id, idx=i, item=item, amount=amount))

            db.session.commit()
            backup_caja_local_y_drive()
            msg = "Guardado."
            return redirect(url_for("asistencia", d=d, emp=empf, start=start, end=end, novf=novf))

        except Exception as ex:
            db.session.rollback()
            err = str(ex)

    # refrescar feriado guardado (por si se modifico)
    hday = get_holiday_type(day_obj)

    rows = []
    auto_fix_commit = False
    for emp_name in EMPLOYEES:
        key = emp_key(emp_name)
        row = Attendance.query.filter_by(day=day_obj, employee=emp_name).first()
        if not row:
            row = Attendance(day=day_obj, employee=emp_name, mode="AUTO")
            db.session.add(row)
            db.session.flush()
            db.session.commit()

        g = row.group_code or group_for_employee_on_day(emp_name, day_obj)
        exp = expected_times_for_group(g, day_obj)
        # Si viene en MANUAL pero no tiene diferencias respecto al horario esperado,
        # lo volvemos a AUTO para que se vea como 'precargado' (gris) por defecto.
        if (row.mode or '').upper() == 'MANUAL':
            # no tocar si hay novedad / notas / consumos
            cons_cnt = AttendanceConsumption.query.filter_by(attendance_id=row.id).count()
            if (not (row.novelty or '').strip()) and (not (row.notes or '').strip()) and cons_cnt == 0:
                def _eq(a_, b_):
                    return (a_ or '') == (b_ or '')
                if (_eq(row.morning_in, exp.get('morning_in')) and _eq(row.morning_out, exp.get('morning_out')) and
                    _eq(row.afternoon_in, exp.get('afternoon_in')) and _eq(row.afternoon_out, exp.get('afternoon_out'))):
                    row.mode = 'AUTO'
                    auto_fix_commit = True

        vac = is_vacation(emp_name, day_obj) or (row.novelty == "Vacaciones")

        mi_preset = (row.mode or "AUTO") == "AUTO" and (row.morning_in is None) and not vac and bool(exp.get("morning_in"))
        mo_preset = (row.mode or "AUTO") == "AUTO" and (row.morning_out is None) and not vac and bool(exp.get("morning_out"))
        ai_preset = (row.mode or "AUTO") == "AUTO" and (row.afternoon_in is None) and not vac and bool(exp.get("afternoon_in"))
        ao_preset = (row.mode or "AUTO") == "AUTO" and (row.afternoon_out is None) and not vac and bool(exp.get("afternoon_out"))

        def show_time(v: Optional[str], default: str) -> str:
            if vac:
                return default or ""
            if (row.mode or "AUTO") == "AUTO":
                return v or (default or "")
            return v or ""

        mi = show_time(row.morning_in, exp["morning_in"])
        mo = show_time(row.morning_out, exp["morning_out"])
        ai = show_time(row.afternoon_in, exp["afternoon_in"])
        ao = show_time(row.afternoon_out, exp["afternoon_out"])

        mm = diff_minutes(mi, mo)
        ma = diff_minutes(ai, ao)

        cons_rows = AttendanceConsumption.query.filter_by(attendance_id=row.id).order_by(AttendanceConsumption.idx.asc()).all()

        calc = compute_work_minutes_and_flags(row)
        payable_time = fmt_minutes(calc["payable_min"])
        warn_text = " | ".join(calc["warnings"]) if calc["warnings"] else ""

        chip = ""
        chip_class = "chip-rp"
        if vac:
            chip = "VAC"
            chip_class = "chip-vac"
        elif row.novelty == "Razones particulares":
            chip = "RP"
            chip_class = "chip-rp"
        elif row.novelty == "Curso":
            chip = "Curso"
            chip_class = "chip-curso"
        elif row.novelty == "Delivery":
            chip = "DEL"
            chip_class = "chip-curso"

        novelty_ui = (row.novelty or "")
        if vac and novelty_ui != "Vacaciones":
            novelty_ui = "Vacaciones"

        rows.append({
            "employee": emp_name,
            "key": key,
            "group": g,
            "mode": (row.mode or "AUTO"),
            "morning_in": mi,
            "morning_out": mo,
            "afternoon_in": ai,
            "afternoon_out": ao,
            "morning_time": fmt_minutes(mm),
            "afternoon_time": fmt_minutes(ma),
            "payable_time": payable_time,
            "novelty": novelty_ui,
            "novelty_minutes": int(row.novelty_minutes or 0),
            "notes": row.notes or "",
            "cons_count": len(cons_rows),
            "cons": cons_rows,
            "is_vac": bool(vac),
            "lock_times": bool(vac or (row.novelty == "Delivery")),
            "has_warning": bool(calc["warnings"]),
            "warn_text": warn_text,
            "chip": chip,
            "chip_class": chip_class,
            "mi_preset": mi_preset,
            "mo_preset": mo_preset,
            "ai_preset": ai_preset,
            "ao_preset": ao_preset,
        })

    if auto_fix_commit:
        db.session.commit()
    backup_caja_local_y_drive()
    summary = attendance_summary_rows(start_date, end_date, empf, novf)

    # Totales del periodo filtrado
    sum_payable_min = 0
    sum_cons = 0
    for rr in summary:
        # rr['payable'] es HH:MM
        try:
            hh, mm = str(rr.get('payable','0:0')).split(':')
            sum_payable_min += int(hh)*60 + int(mm)
        except:
            pass
        sum_cons += int(rr.get('cons_total') or 0)
    sum_hours_txt = fmt_minutes(sum_payable_min)

    return render_template_string(
        ASISTENCIA_HTML,
        base_css=BASE_CSS,
        d=d,
        hday=hday,
        ga=GROUP_A, gb=GROUP_B,
        rows=rows,
        novelty_items=NOVELTY_ITEMS,
        maxc=MAX_CONSUMOS,
        err=err,
        msg=msg,
        employees=EMPLOYEES,
        emp=empf,
        start=start,
        end=end,
        novf=novf,
        summary=summary,
        sum_hours_txt=sum_hours_txt,
        sum_cons_money=sum_cons,
    )
@app.route("/asistencia/export/excel")
@login_required
def att_export_excel():
    emp = request.args.get("emp") or "ALL"
    novf = request.args.get("novf") or "ALL"
    start_date, end_date = parse_range_params(request.args.get("start"), request.args.get("end"))
    data = attendance_summary_rows(start_date, end_date, emp, novf)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.append(["Fecha","Empleado","Grupo","Manana","Tarde","Horas pagas","Novedad","Notas","Consumos total","Consumos items"])

    for r in data:
        ws.append([
            r["day"], r["employee"], r["group"], r["morning"], r["afternoon"],
            r["payable"], r["novelty"], r["notes"],
            r["cons_total"], r["cons_items"]
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"asistencia_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/asistencia/export/json")
@login_required
def att_export_json():
    emp = request.args.get("emp") or "ALL"
    novf = request.args.get("novf") or "ALL"
    start_date, end_date = parse_range_params(request.args.get("start"), request.args.get("end"))

    # Export del resumen filtrado (attendance + consumptions)
    data = attendance_summary_rows(start_date, end_date, emp, novf)

    # Ademas exportamos los registros completos (attendance + consumptions)
    q = Attendance.query.filter(Attendance.day >= start_date, Attendance.day <= end_date)
    if emp != "ALL":
        q = q.filter(Attendance.employee == emp)

    attendances = []
    for a in q.order_by(Attendance.day.asc(), Attendance.employee.asc()).all():
        cons = (
            AttendanceConsumption.query
            .filter_by(attendance_id=a.id)
            .order_by(AttendanceConsumption.idx.asc())
            .all()
        )
        attendances.append({
            "day": a.day.isoformat(),
            "employee": a.employee,
            "group_code": a.group_code,
            "mode": a.mode,
            "morning_in": a.morning_in,
            "morning_out": a.morning_out,
            "afternoon_in": a.afternoon_in,
            "afternoon_out": a.afternoon_out,
            "novelty": a.novelty,
            "novelty_minutes": int(a.novelty_minutes or 0),
            "notes": a.notes,
            "consumptions": [
                {"idx": int(c.idx), "item": c.item, "amount": int(c.amount or 0)}
                for c in cons
            ]
        })

    payload = {
        "type": "asistencia_backup",
        "version": 1,
        "exported_at": datetime.utcnow().isoformat(),
        "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "filters": {"emp": emp, "novf": novf},
        "attendances": attendances,
        "summary": data,
        # opcional (backup extra): config y calendarios
        "rotation_config": (
            {"week0_map": rotation_config_get().week0_map, "created_by": rotation_config_get().created_by,
             "created_at": rotation_config_get().created_at.isoformat()}
            if rotation_config_get() else None
        ),
        "vacations": [
            {"employee": v.employee, "start_day": v.start_day.isoformat(), "end_day": v.end_day.isoformat()}
            for v in Vacation.query.order_by(Vacation.employee.asc(), Vacation.start_day.asc()).all()
        ],
        "calendar_days": [
            {"day": cd.day.isoformat(), "holiday_type": cd.holiday_type or ""}
            for cd in CalendarDay.query.order_by(CalendarDay.day.asc()).all()
        ]
    }

    bio = BytesIO()
    bio.write(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    bio.seek(0)
    filename = f"asistencia_backup_{start_date.isoformat()}_{end_date.isoformat()}.json"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/json")

# ==============================
# IMPORT ASISTENCIA
# ==============================
IMPORT_ATT_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Import Asistencia</title>{{ base_css|safe }}</head>
<body style="max-width:950px;">
  <a href="{{ url_for('asistencia') }}"><- Volver</a>
  <h2>Importar datos (Excel) - Asistencia</h2>

  <div class="box" style="margin-bottom:12px;">
    <b>Formato:</b> hoja <code>attendance</code> (y opcional <code>consumptions</code>).<br>
    Columnas minimas: <code>day</code>, <code>employee</code>, <code>group</code>, <code>morning_in</code>, <code>morning_out</code>,
    <code>afternoon_in</code>, <code>afternoon_out</code>, <code>novelty</code>, <code>novelty_minutes</code>, <code>notes</code>.
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
    <div class="box">
      <label><b>Archivo Excel</b></label><br>
      <input type="file" name="file" accept=".xlsx" required>
      <br><br>
      <label><b>Modo</b></label><br>
      <select name="mode">
        <option value="skip" selected>Si existe el dia/empleado, NO tocar (skip)</option>
        <option value="replace">Si existe el dia/empleado, REEMPLAZAR (replace)</option>
      </select>
    </div>
    <br>
    <button class="btn">Importar</button>
  </form>
</body>
</html>
"""

@app.route("/import/asistencia", methods=["GET","POST"])
@login_required
def import_asistencia():
    err = None
    msg = None
    if request.method == "POST":
        mode = (request.form.get("mode") or "skip").strip()
        f = request.files.get("file")
        if not f:
            err = "No se recibio archivo."
        else:
            try:
                wb = load_workbook(f, data_only=True)
                if "attendance" not in wb.sheetnames:
                    raise ValueError("No existe la hoja 'attendance'.")

                ws = wb["attendance"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    raise ValueError("La hoja 'attendance' no tiene datos.")

                h = [str(x).strip() if x is not None else "" for x in rows[0]]
                need = ["day","employee","group","morning_in","morning_out","afternoon_in","afternoon_out","novelty","novelty_minutes","notes"]
                for n in need:
                    if n not in h:
                        raise ValueError(f"Falta columna '{n}' en attendance.")
                idx = {k:i for i,k in enumerate(h)}

                imported = 0
                replaced = 0

                for r in rows[1:]:
                    if not r or all(v is None or str(v).strip()=="" for v in r):
                        continue

                    day_v = r[idx["day"]]
                    if isinstance(day_v, datetime):
                        day_v = day_v.date()
                    elif isinstance(day_v, date):
                        pass
                    else:
                        day_v = date.fromisoformat(str(day_v).strip()[:10])

                    emp = str(r[idx["employee"]]).strip()
                    if emp not in EMPLOYEES:
                        continue

                    existing = Attendance.query.filter_by(day=day_v, employee=emp).first()
                    if existing:
                        if mode == "skip":
                            continue
                        AttendanceConsumption.query.filter_by(attendance_id=existing.id).delete()
                        db.session.delete(existing)
                        db.session.flush()
                        replaced += 1

                    grp = (str(r[idx["group"]]).strip().upper() or "A")
                    if grp not in ("A","B"):
                        grp = "A"

                    mi = (str(r[idx["morning_in"]]).strip() if r[idx["morning_in"]] else None)
                    mo = (str(r[idx["morning_out"]]).strip() if r[idx["morning_out"]] else None)
                    ai = (str(r[idx["afternoon_in"]]).strip() if r[idx["afternoon_in"]] else None)
                    ao = (str(r[idx["afternoon_out"]]).strip() if r[idx["afternoon_out"]] else None)

                    nov = (str(r[idx["novelty"]]).strip() if r[idx["novelty"]] else "") or ""
                    if nov not in NOVELTY_ITEMS:
                        nov = ""

                    novm = int(r[idx["novelty_minutes"]] or 0)

                    raw_notes_cell = r[idx["notes"]]  # no uses "if cell"
                    raw_notes = None if raw_notes_cell is None else str(raw_notes_cell).strip()

                    cons, notes_clean = extract_consumptions_and_clean_notes(
                        raw_notes, max_items=MAX_CONSUMOS
                    )

                    a = Attendance(
                        day=day_v,
                        employee=emp,
                        group_code=grp,
                        mode="MANUAL",
                        morning_in=mi,
                        morning_out=mo,
                        afternoon_in=ai,
                        afternoon_out=ao,
                        novelty=nov or None,
                        novelty_minutes=novm,
                        notes=notes_clean
                    )
                    db.session.add(a)
                    db.session.flush()

                    # Crear consumos detectados
                    for i, (item, amount) in enumerate(cons, start=1):
                        db.session.add(
                            AttendanceConsumption(
                                attendance_id=a.id,
                                idx=i,
                                item=item,
                                amount=amount
                            )
                        )

                    imported += 1

                db.session.commit()
                backup_caja_local_y_drive()
                msg = f"Import OK. Registros: {imported}. Reemplazados: {replaced}."
            except Exception as ex:
                db.session.rollback()
                err = str(ex)

    return render_template_string(IMPORT_ATT_HTML, base_css=BASE_CSS, err=err, msg=msg)

IMPORT_ATT_JSON_HTML = """
<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Import Asistencia (JSON)</title>{{ base_css|safe }}</head>
<body style="max-width:950px;">
  <a href="{{ url_for('asistencia') }}"><- Volver</a>
  <h2>Importar Backup JSON - Asistencia</h2>

  <div class="box" style="margin-bottom:12px;">
    <b>Formato:</b> archivo JSON exportado por "Exportar JSON".<br>
    <b>Modo skip:</b> si existe (dia+empleado), no lo toca.<br>
    <b>Modo replace:</b> reemplaza (dia+empleado) y recrea consumos.<br>
    <span class="muted">Ademas: si el JSON trae vacaciones / feriados / rotacion, se "mergea" (upsert simple).</span>
  </div>

  {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
  {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

  <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
    <div class="box">
      <label><b>Archivo JSON</b></label><br>
      <input type="file" name="file" accept=".json" required>
      <br><br>
      <label><b>Modo</b></label><br>
      <select name="mode">
        <option value="skip" selected>Si existe el dia/empleado, NO tocar (skip)</option>
        <option value="replace">Si existe el dia/empleado, REEMPLAZAR (replace)</option>
      </select>
    </div>
    <br>
    <button class="btn">Importar JSON</button>
  </form>
</body>
</html>
"""

@app.route("/import/asistencia/json", methods=["GET","POST"])
@login_required
def import_asistencia_json():
    err = None
    msg = None

    if request.method == "POST":
        mode = (request.form.get("mode") or "skip").strip()
        f = request.files.get("file")
        if not f:
            err = "No se recibio archivo."
        else:
            try:
                payload = json.load(f)
                if not isinstance(payload, dict) or payload.get("type") != "asistencia_backup":
                    raise ValueError("JSON invalido: no parece un backup de Asistencia.")

                attendances = payload.get("attendances") or []

                imported = 0
                replaced = 0
                skipped = 0

                for a in attendances:
                    dd = (a.get("day") or "").strip()
                    emp = (a.get("employee") or "").strip()
                    if not dd or emp not in EMPLOYEES:
                        continue

                    day_obj = date.fromisoformat(dd)

                    existing = Attendance.query.filter_by(day=day_obj, employee=emp).first()
                    if existing:
                        if mode == "skip":
                            skipped += 1
                            continue
                        AttendanceConsumption.query.filter_by(attendance_id=existing.id).delete()
                        db.session.delete(existing)
                        db.session.flush()
                        replaced += 1

                    grp = (a.get("group_code") or "A").strip().upper()
                    if grp not in ("A","B"):
                        grp = "A"

                    nov = (a.get("novelty") or "").strip()
                    if nov and nov not in NOVELTY_ITEMS:
                        nov = ""

                    row = Attendance(
                        day=day_obj,
                        employee=emp,
                        group_code=grp,
                        mode=(a.get("mode") or "MANUAL"),
                        morning_in=a.get("morning_in"),
                        morning_out=a.get("morning_out"),
                        afternoon_in=a.get("afternoon_in"),
                        afternoon_out=a.get("afternoon_out"),
                        novelty=nov or None,
                        novelty_minutes=int(a.get("novelty_minutes") or 0),
                        notes=(a.get("notes") or None),
                    )
                    db.session.add(row)
                    db.session.flush()

                    cons = a.get("consumptions") or []
                    for c in cons:
                        idx = int(c.get("idx") or 0)
                        if idx <= 0:
                            continue
                        db.session.add(AttendanceConsumption(
                            attendance_id=row.id,
                            idx=idx,
                            item=(c.get("item") or None),
                            amount=int(c.get("amount") or 0)
                        ))

                    imported += 1

                # merge extras (no pisa todo, solo upsert simple)
                rc = payload.get("rotation_config")
                if rc and isinstance(rc, dict) and rc.get("week0_map"):
                    # guardamos como nueva config (la mas nueva manda)
                    db.session.add(RotationConfig(
                        week0_map=str(rc.get("week0_map")),
                        created_by=(rc.get("created_by") or "import-json")
                    ))

                vacs = payload.get("vacations") or []
                for v in vacs:
                    emp = (v.get("employee") or "").strip()
                    sd = (v.get("start_day") or "").strip()
                    ed = (v.get("end_day") or "").strip()
                    if emp not in EMPLOYEES or not sd or not ed:
                        continue
                    sd2 = date.fromisoformat(sd)
                    ed2 = date.fromisoformat(ed)
                    exists = Vacation.query.filter_by(employee=emp, start_day=sd2, end_day=ed2).first()
                    if not exists:
                        db.session.add(Vacation(employee=emp, start_day=sd2, end_day=ed2))

                cds = payload.get("calendar_days") or []
                for cd in cds:
                    dd = (cd.get("day") or "").strip()
                    if not dd:
                        continue
                    d2 = date.fromisoformat(dd)
                    ht = (cd.get("holiday_type") or "").strip().upper()
                    if ht not in ("", "LABORABLE", "NO_LABORABLE"):
                        ht = ""
                    set_holiday_type(d2, ht)

                db.session.commit()
                backup_caja_local_y_drive()
                msg = f"Import JSON OK. Registros: {imported}. Reemplazados: {replaced}. Omitidos (skip): {skipped}."
            except Exception as ex:
                db.session.rollback()
                err = str(ex)

    return render_template_string(IMPORT_ATT_JSON_HTML, base_css=BASE_CSS, err=err, msg=msg)

# ==============================
# MAIN
# ==============================
if __name__ == "__main__":
    # En local (python app.py) si usamos app.run.
    # En Render esto NO corre (porque Render usa gunicorn).
    port = int(os.getenv("PORT", "5050"))
    app.run(host="127.0.0.1", port=port, debug=False)
