from __future__ import annotations

import json
import re
from datetime import date, datetime
from io import BytesIO

from flask import abort, redirect, render_template_string, request, send_file, url_for
from openpyxl import Workbook, load_workbook


def register_caja_routes(
    app,
    *,
    login_required,
    current_user,
    backup_caja_local_y_drive,
    parse_day_param,
    parse_range_params,
    safe_int,
    to_int,
    weekday_es,
    responsible_name,
    Shift,
    ShiftClose,
    CashExpense,
    TURNS,
    TURN_NAMES,
    CATEGORIES,
    EMPLOYEES,
    ADMINS,
    DB_PATH,
    BASE_CSS,
    DELIVERY_SHIFT_PRESETS,
    build_delivery_payload,
    sanitize_delivery_payload,
    expenses_total,
    cash_bruto,
    calc_ingreso_bruto,
    calc_ingreso_neto,
    calc_ending_calc,
    get_locked_opening_cash,
    is_placeholder_shift,
    get_caja_summary,
    can_edit_close,
):
    CAJA_INDEX_HTML = """
    <!doctype html>
    <html lang="es">
    <head>
      <meta charset="utf-8">
      <title>Control de Caja - PORA</title>
      {{ base_css|safe }}
      <style>
        body{max-width:1300px;}
        .filters{display:flex; gap:10px; flex-wrap:wrap; align-items:end; margin:8px 0 14px;}
        .filters label{display:block; font-size:12px; color:#666;}
        .smallnote{font-size:12px; color:#666; margin-top:6px;}
        .holiday-bg{background:#ffe3ec;}
        .holiday-bg table tr{background:#fff;}
        .summary-table{table-layout:auto;}
        .summary-table th,
        .summary-table td{padding:6px 7px; font-size:12px;}
        .summary-table th{
          white-space:normal;
          line-height:1.1;
          text-align:center;
          vertical-align:middle;
        }
        .summary-table td{
          white-space:nowrap;
          vertical-align:middle;
        }
        .summary-table td.wrapday{white-space:normal;}
      </style>
    </head>
    <body class="{{ 'holiday-bg' if hday else '' }}">

    <div class="top">
      <div class="logo">
        <img src="{{ url_for('static', filename='img/pora_logo.png') }}" alt="PORA">
        <div>
          <h1 style="margin:0;">Control de Caja</h1>
          <div class="muted"><a href="{{ url_for('home') }}"><- Volver al menu</a></div>

          <form method="get" action="{{ url_for('caja_index') }}" style="margin-top:6px; display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
            <label class="muted">Fecha:</label>
            <input type="date" name="d" value="{{d}}">
            <button class="btn" type="submit">Ir</button>
          </form>
        </div>
      </div>
      <div class="row">
        <a class="btn" href="{{ url_for('export_excel') }}">Exportar Excel</a>
        <a class="btn" href="{{ url_for('export_caja_json') }}">Exportar JSON</a>
        <a class="btn" href="{{ url_for('import_caja') }}">Importar Excel</a>
        <a class="btn" href="{{ url_for('import_caja_json') }}">Importar JSON</a>
      </div>
    </div>

    <h2>Turnos del dia</h2>
    <table class="tight">
      <tr>
        <th>Dia</th>
        <th>Turno</th>
        <th>Responsable</th>
        <th>Estado</th>
        <th>Cierre</th>
        <th>Ventas netas</th>
        <th>Efectivo disponible</th>
        <th>Accion</th>
      </tr>

      {% for code,name in turns %}
        {% set s = shifts.get(code) %}
        <tr>
          <td>{{ weekday_name }}</td>
          <td><b>{{name}}</b></td>
          <td>{{ responsible_name(s.responsible) if s else "-" }}</td>
          <td>
            {% if not s %}
              <span class="badge badge-open">No abierto</span>
            {% elif s.status == "OPEN" %}
              <span class="badge badge-open">Abierto</span>
            {% else %}
              <span class="badge badge-closed">Cerrado</span>
            {% endif %}
          </td>
          <td>
            {% if s and s.status == "CLOSED" %}
              {% set c = closes.get(s.id) %}
              {% if c and c.close_ok == 1 %}
                <span class="badge badge-ok">OK</span>
              {% else %}
                <span class="badge badge-warn">NO OK</span>
              {% endif %}
            {% else %}
              -
            {% endif %}
          </td>

          <td>
            {% if s and s.status == "CLOSED" %}
              {{ ventas_netas_map.get(s.id, 0)|money }}
            {% else %}-{% endif %}
          </td>

          <td>
            {% if s and s.status == "CLOSED" %}
              {{ efectivo_disp_map.get(s.id, 0)|money }}
            {% else %}-{% endif %}
          </td>

          <td>
            {% if not s %}
              <a href="{{ url_for('open_shift', turn=code, d=d) }}">Abrir</a>
            {% else %}
              {% if s.status == "OPEN" %}
                <a href="{{ url_for('shift', id=s.id, d=d) }}">Entrar</a>
              {% else %}
                <a href="{{ url_for('shift', id=s.id, d=d) }}">Ver</a>
                {% if can_edit_map.get(s.id) %}
                  | <a href="{{ url_for('edit_all', id=s.id, d=d) }}">Editar</a>
                {% endif %}
              {% endif %}
            {% endif %}
          </td>
        </tr>
      {% endfor %}
    </table>

    {% if efectivo_total_dia is not none %}
      <div class="smallnote">
        <b>Total efectivo disponible del dia:</b> {{ efectivo_total_dia|money }}
        <br><b>Total ingreso del dia (neto):</b> {{ ingreso_neto_total_dia|money }}
      </div>
    {% endif %}

    <h2 style="margin-top:18px;">Resumen de cierres</h2>

    <form method="get" action="{{ url_for('caja_index') }}" class="filters">
      <input type="hidden" name="d" value="{{d}}">

      <div>
        <label>Desde</label>
        <input type="date" name="start" value="{{start}}">
      </div>

      <div>
        <label>Hasta</label>
        <input type="date" name="end" value="{{end}}">
      </div>

      <div>
        <label>Turno</label>
        <select name="turn">
          <option value="ALL" {% if turn=='ALL' %}selected{% endif %}>Todos</option>
          <option value="MORNING" {% if turn=='MORNING' %}selected{% endif %}>Manana</option>
          <option value="AFTERNOON" {% if turn=='AFTERNOON' %}selected{% endif %}>Tarde</option>
        </select>
      </div>

      <div>
        <label>Responsable</label>
        <select name="resp">
          <option value="ALL" {% if resp=='ALL' %}selected{% endif %}>Todos</option>
          {% for r in responsibles %}
            <option value="{{r}}" {% if resp==r %}selected{% endif %}>{{r}}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <button class="btn" type="submit">Aplicar</button>
      </div>
    </form>

    <table class="tight summary-table">
      <tr>
        <th>Dia</th>
        <th class="nowrap">Fecha</th>
        <th>Turno</th>
        <th>Responsable</th>
        <th>Caja<br>Inicial</th>
        <th>Efectivo<br>bruto</th>
        <th>Ventas<br>Mercado Pago</th>
        <th>Ventas<br>Pedidos Ya</th>
        <th>Ventas<br>Rappi</th>
        <th>Egresos</th>
        <th>Ventas<br>totales</th>
        <th class="net-soft">Ventas<br>netas</th>
        <th class="net-soft">Ventas netas<br>del dia</th>
        <th>Retirado</th>
        <th>Caja Final<br>(Real)</th>
      </tr>

      {% for row in summary %}
        <tr>
          {% if row.show_day %}
            <td class="wrapday" rowspan="{{ row.day_rowspan }}"><b>{{ row.weekday }}</b></td>
          {% endif %}
          <td class="nowrap">{{ row.day }}</td>
          <td>{{ row.turn_name }}</td>
          <td>{{ row.responsible }}</td>
          <td>{{ row.opening_cash|money }}</td>
          <td>{{ row.sales_cash|money }}</td>
          <td>{{ row.sales_mp|money }}</td>
          <td>{{ row.sales_pya|money }}</td>
          <td>{{ row.sales_rappi|money }}</td>
          <td>{{ row.expenses|money }}</td>
          <td>{{ row.ventas_bruto|money }}</td>
          <td class="net-soft"><b>{{ row.ventas_neto|money }}</b></td>
          {% if row.show_day_total %}
            <td class="net-soft" rowspan="{{ row.day_rowspan }}"><b>{{ row.day_total_neto|money }}</b></td>
          {% endif %}
          <td>{{ row.withdrawn|money }}</td>
          <td>{{ row.ending_real|money }}</td>
        </tr>
      {% endfor %}

      {% if not summary %}
        <tr><td colspan="15" class="muted">Sin cierres para el filtro seleccionado.</td></tr>
      {% endif %}
    </table>

    <p class="muted">DB: <code>{{db_path}}</code></p>
    </body>
    </html>
    """

    OPEN_HTML = """
    <!doctype html>
    <html lang="es">
    <head>
    <meta charset="utf-8">
    <title>Abrir turno</title>
    {{ base_css|safe }}
    <style>
      body{max-width:540px;}
      input,button{padding:10px; font-size:14px;}
      .holiday-bg{background:#ffe3ec;}
      .holiday-bg table tr{background:#fff;}
    </style>
    </head>
    <body>
    <a href="{{ url_for('caja_index', d=d) }}"><- Volver</a>
    <h3>Abrir turno {{turn_name}} ({{day}})</h3>

    <div class="box">
      <form method="post">
        <input type="hidden" name="d" value="{{d}}">

        {% if locked_opening is not none %}
          <p><b>Caja inicial</b> (del cierre anterior):</p>
          <p><input type="number" name="opening_cash" value="{{locked_opening}}" readonly></p>
          <p class="muted">No se puede editar.</p>
        {% else %}
          <p><b>Caja inicial</b> (no hay cierre previo):</p>
          <p><input type="number" name="opening_cash" min="0" required></p>
          <p class="muted">Solo se usa si es el primer turno sin historial.</p>
        {% endif %}

        <p><b>Responsable:</b></p>
        <input name="responsible" required style="width:100%;" value="{{default_responsible}}" readonly>

        <br><br>
        <button class="btn">Abrir</button>
      </form>
    </div>
    </body>
    </html>
    """

    SHIFT_HTML = """
    <!doctype html>
    <html lang="es">
    <head>
      <meta charset="utf-8">
      <title>Turno</title>
      {{ base_css|safe }}
      <style>
        body{max-width:1100px;}
        input,select,button,textarea{padding:8px; font-size:14px;}
        .bad{color:#842029; font-weight:bold;}
        .good{color:#0f5132; font-weight:bold;}
        .holiday-bg{background:#ffe3ec;}
      .holiday-bg table tr{background:#fff;}
      .delivery-actions{display:flex; gap:10px; align-items:center; flex-wrap:wrap;}
      .delivery-backdrop{display:none; position:fixed; inset:0; background:rgba(0,0,0,.35); z-index:999;}
      .delivery-modal{display:none; position:fixed; top:50%; left:50%; transform:translate(-50%,-50%); width:min(960px, calc(100vw - 24px)); max-height:85vh; overflow:auto; background:#fff; border:1px solid #ccc; border-radius:18px; padding:18px; z-index:1000; box-shadow:0 18px 45px rgba(0,0,0,.22);}
      .delivery-modal h3{margin:0 0 8px;}
      .delivery-modal .head{display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:10px;}
      .delivery-table{width:100%; border-collapse:collapse; table-layout:fixed;}
      .delivery-table th,.delivery-table td{padding:10px 8px; border:1px solid #ddd; vertical-align:middle;}
      .delivery-table th{text-align:left; background:#f3f3f3;}
      .delivery-table .money-input-wrap{display:inline-flex; align-items:center; gap:6px; white-space:nowrap;}
      .delivery-table .money-input-wrap input{width:84px; text-align:center; padding:8px 6px;}
      .delivery-table .qty-input{width:72px; text-align:center; padding:8px 6px;}
      .delivery-table .concept-col{width:24%;}
      .delivery-table .rate-col{width:24%;}
      .delivery-table .qty-col{width:22%;}
      .delivery-table .tot-col{width:30%;}
      .delivery-total-row td{font-weight:bold; background:#fafafa;}
      .delivery-consumos-box{margin-top:12px; padding:12px; border:1px solid #ddd; border-radius:12px; background:#fafafa;}
      .delivery-consumos-grid{display:grid; grid-template-columns:160px 1fr; gap:10px; align-items:center;}
      .delivery-consumos-grid input{padding:8px 10px;}
      .delivery-consumos-grid .money-input-wrap input{width:96px; text-align:center;}
      .delivery-impact-box{margin-top:12px; display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap;}
      .delivery-impact-total{font-weight:bold; font-size:18px;}
    </style>
    </head>
    <body>
    <a href="{{ url_for('caja_index', d=d) }}"><- Volver</a>
    <h2>Turno {{turn_name}} - {{s.day}}</h2>

    <p>
      Responsable: <b>{{s.responsible}}</b> |
      Estado: <b>{{s.status}}</b> |
      Caja inicial: <b>{{ s.opening_cash|money }}</b>
    </p>

    {% if close and can_edit %}
      <p class="muted">
        <a class="btn" href="{{ url_for('edit_all', id=s.id, d=d) }}">Editar (todo)</a>
        {% if close.edit_count %}
          <span class="muted">Ediciones: {{close.edit_count}}</span>
        {% endif %}
      </p>
    {% elif close and not can_edit %}
      <p class="muted">Este turno ya no puede editarse con tu usuario.</p>
    {% endif %}


    <h3>Ventas</h3>
    <div class="box">
      <div class="row" style="align-items:flex-start;">
        <div style="min-width:280px;">
          <label><b>Retirado (Efectivo total)</b></label><br>
          <input type="number" name="sales_cash" min="0" value="{{s.sales_cash}}" {{'disabled' if s.status=='CLOSED' else ''}} id="retirado_input">
          <div class="muted" style="margin-top:6px;">
            Efectivo bruto (auto): <b id="ef_bruto_lbl">{{ (s.sales_cash + (close.ending_cash if close else 0))|money }}</b><br>
            Efectivo neto (auto): <b id="ef_neto_lbl">{{ (s.sales_cash - s.opening_cash)|money }}</b>
          </div>
        </div>

        <div style="min-width:320px;">
          <label><b>Caja final (efectivo que queda en la caja)</b></label><br>
          <input type="number" name="cash_final" min="0" value="{{ (close.ending_cash if close else 0) }}" {{'disabled' if s.status=='CLOSED' else ''}} id="caja_final_input">
          <div class="muted" style="margin-top:6px;">
            (Se carga manualmente)
          </div>
        </div>
      </div>

      <div class="row" style="margin-top:10px;">
        <div>
          Ventas Mercado Pago:<br>
          <input type="number" name="sales_mp" min="0" value="{{s.sales_mp}}" {{'disabled' if s.status=='CLOSED' else ''}} id="mp_input">
        </div>
        <div>
          Ventas Pedidos Ya:<br>
          <input type="number" name="sales_pya" min="0" value="{{s.sales_pya}}" {{'disabled' if s.status=='CLOSED' else ''}} id="pya_input">
        </div>
        <div>
          Ventas Rappi:<br>
          <input type="number" name="sales_rappi" min="0" value="{{s.sales_rappi}}" {{'disabled' if s.status=='CLOSED' else ''}} id="rappi_input">
        </div>
      </div>
    </div>


    <h3>Egresos (caja chica, efectivo)</h3>
    {% if s.status != 'CLOSED' %}
    <div class="delivery-actions">
      <form method="post" action="{{ url_for('expense', id=s.id, d=d) }}" id="expenseForm">
        <select name="category" id="expenseCategory">
          {% for c in categories %}<option>{{c}}</option>{% endfor %}
        </select>
        <input type="number" name="amount" min="1" required id="expenseAmount">
        <input name="note" placeholder="nota (obligatoria si Otros)" id="expenseNote">
        <button class="btn">Agregar</button>
      </form>

      <button type="button" class="btn" id="openDeliveryCalcBtn">Calculo para delivery</button>
    </div>

    <div class="delivery-backdrop" id="deliveryBackdrop"></div>
    <div class="delivery-modal" id="deliveryModal" aria-hidden="true">
      <div class="head">
        <div>
          <h3>Calculadora pago delivery</h3>
          <div class="muted">Los importes por pedido/hora vienen precargados, pero podes editarlos. La informacion se guarda automaticamente en este turno.</div>
        </div>
        <button type="button" class="btn" id="closeDeliveryCalcBtn">Cerrar</button>
      </div>

      <table class="delivery-table">
        <tr>
          <th class="concept-col">Concepto</th>
          <th class="rate-col">$ por pedido/hora</th>
          <th class="qty-col">Cantidad</th>
          <th class="tot-col">Total</th>
        </tr>
        <tr>
          <td rowspan="4" style="text-align:center;"><b>Pedidos</b></td>
          <td>
            <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="0" value="1500"></div>
          </td>
          <td><input type="number" class="del-qty qty-input" data-row="0" value="0" min="0" step="1"></td>
          <td class="del-total" data-row="0">$ 0</td>
        </tr>
        <tr>
          <td>
            <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="1" value="2000"></div>
          </td>
          <td><input type="number" class="del-qty qty-input" data-row="1" value="0" min="0" step="1"></td>
          <td class="del-total" data-row="1">$ 0</td>
        </tr>
        <tr>
          <td>
            <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="2" value="2500"></div>
          </td>
          <td><input type="number" class="del-qty qty-input" data-row="2" value="0" min="0" step="1"></td>
          <td class="del-total" data-row="2">$ 0</td>
        </tr>
        <tr>
          <td>
            <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="3" value="3000"></div>
          </td>
          <td><input type="number" class="del-qty qty-input" data-row="3" value="0" min="0" step="1"></td>
          <td class="del-total" data-row="3">$ 0</td>
        </tr>
        <tr>
          <td style="text-align:center;"><b>Horas</b></td>
          <td>
            <div class="money-input-wrap">$ <input type="number" class="del-rate" data-row="4" value="2500"></div>
          </td>
          <td>
            <div style="display:grid; gap:8px;">
              <div>
                <label class="muted" for="deliveryHourShift">Turno</label><br>
                <select id="deliveryHourShift">
                  <option value="MORNING">Mañana</option>
                  <option value="AFTERNOON">Tarde</option>
                </select>
              </div>
              <div class="row" style="gap:8px; align-items:end;">
                <div>
                  <label class="muted" for="deliveryHourIn">Entrada</label><br>
                  <input type="time" id="deliveryHourIn" class="qty-input" style="width:120px;">
                </div>
                <div>
                  <label class="muted" for="deliveryHourOut">Salida</label><br>
                  <input type="time" id="deliveryHourOut" class="qty-input" style="width:120px;">
                </div>
              </div>
              <div class="muted">Horas calculadas automáticamente: <b id="deliveryHoursLabel">0.00</b></div>
              <input type="number" class="del-qty qty-input" data-row="4" value="0" min="0" step="0.01" readonly>
            </div>
          </td>
          <td class="del-total" data-row="4">$ 0</td>
        </tr>
        <tr class="delivery-total-row">
          <td colspan="3" style="text-align:right;">Total viajes / horas</td>
          <td id="deliveryGrandTotal">$ 0</td>
        </tr>
      </table>

      <div class="delivery-consumos-box">
        <div style="font-weight:bold; margin-bottom:8px;">Consumos</div>
        <div class="delivery-consumos-grid">
          <div>
            <label class="muted" for="deliveryConsumeAmount">Monto</label><br>
            <div class="money-input-wrap">$ <input type="number" id="deliveryConsumeAmount" min="0" value="0"></div>
          </div>
          <div>
            <label class="muted" for="deliveryConsumeNote">Notas</label><br>
            <input type="text" id="deliveryConsumeNote" placeholder="Ej: gaseosa, cena, peaje, etc.">
          </div>
        </div>
        <div class="muted" style="margin-top:8px;">El monto de consumos se descuenta del total de viajes / horas. Las horas se calculan solas según turno, entrada y salida.</div>
      </div>

      <div class="delivery-impact-box">
        <div>
          <div class="muted">Total a cargar como egreso Delivery / Cadete</div>
          <div class="delivery-impact-total" id="deliveryNetTotal">$ 0</div>
        </div>
        <button type="button" class="btn" id="applyDeliveryExpenseBtn">Impactar en egresos</button>
      </div>
    </div>
    {% else %}
    <p class="muted">Turno cerrado: no se pueden agregar egresos desde aca. Usa Editar (todo).</p>
    {% endif %}

    <table style="margin-top:10px;">
      <tr><th>Categoria</th><th>Monto</th><th>Nota</th></tr>
      {% for e in expenses %}
        <tr><td>{{e.category}}</td><td>{{e.amount|money}}</td><td>{{e.note or ''}}</td></tr>
      {% endfor %}
      {% if not expenses %}
        <tr><td colspan="3" class="muted">Sin egresos</td></tr>
      {% endif %}
    </table>


    <h3>Cierre</h3>

    <div class="box">
      <div class="row">
        <div><b>Egresos total:</b> <span id="egresos_total_lbl">{{ egresos_total|money }}</span></div>
        <div><b>Ingreso total (bruto):</b> <span id="ingreso_bruto_lbl">{{ ingreso_bruto|money }}</span></div>
        <div><b>Ingreso neto:</b> <span id="ingreso_neto_lbl">{{ ingreso_neto|money }}</span></div>
        <div><b>Efectivo disponible:</b> <span id="ef_disp_lbl">{{ efectivo_disponible|money }}</span></div>
      </div>
      <div class="muted" style="margin-top:6px;">
        Efectivo bruto = Retirado + Caja final.<br>
        Efectivo neto = Retirado - Caja inicial.<br>
        Ingreso total (bruto) = (Efectivo bruto + MP + PedidosYa + Rappi) + Egresos.<br>
        Ingreso neto = Ingreso total (bruto) - Egresos total.<br>
        Efectivo disponible = Retirado.
      </div>
    </div>

    {% if close %}
      <div class="box" style="margin-top:10px;">
        <div class="row">
          <div><b>Retirado:</b> {{ s.sales_cash|money }}</div>
          <div><b>Caja final:</b> {{ close.ending_cash|money }}</div>
        </div>
      </div>
    {% else %}
      <form method="post" action="{{ url_for('close_shift', id=s.id, d=d) }}" style="margin-top:14px;" id="saveForm">
        <input type="hidden" name="sales_cash" id="h_sales_cash">
        <input type="hidden" name="cash_final" id="h_cash_final">
        <input type="hidden" name="sales_mp" id="h_sales_mp">
        <input type="hidden" name="sales_pya" id="h_sales_pya">
        <input type="hidden" name="sales_rappi" id="h_sales_rappi">
        <button class="btn" style="margin-top:10px;">Guardar registro</button>
      </form>

      <script>
        const sid = "{{s.id}}";

        function toInt(v){
          v = (v || "").toString().trim();
          if(v === "") return 0;
          const n = parseInt(v, 10);
          return isNaN(n) ? 0 : n;
        }

        function toNum(v){
          v = (v || "").toString().trim().replace(",", ".");
          if(v === "") return 0;
          const n = parseFloat(v);
          return isNaN(n) ? 0 : n;
        }

        const retiradoEl = document.getElementById("retirado_input");
        const cajaFinalEl = document.getElementById("caja_final_input");
        const mpEl = document.getElementById("mp_input");
        const pyaEl = document.getElementById("pya_input");
        const rappiEl = document.getElementById("rappi_input");

        const efBrutoLbl = document.getElementById("ef_bruto_lbl");
        const efNetoLbl = document.getElementById("ef_neto_lbl");
        const ingresoBrutoLbl = document.getElementById("ingreso_bruto_lbl");
        const ingresoNetoLbl = document.getElementById("ingreso_neto_lbl");
        const efDispLbl = document.getElementById("ef_disp_lbl");

        const egresosTotal = {{ egresos_total|int }};

        function fmtMoney(n){
          try{ n = parseInt(n,10); }catch(e){ n = 0; }
          if(isNaN(n)) n = 0;
          const s = n.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ".");
          return "$ " + s;
        }

        function saveDraft(){
          const payload = {
            retirado: toInt(retiradoEl.value),
            caja_final: toInt(cajaFinalEl.value),
            mp: toInt(mpEl.value),
            pya: toInt(pyaEl.value),
            rappi: toInt(rappiEl.value)
          };
          try{
            localStorage.setItem("caja_draft_"+sid, JSON.stringify(payload));
          }catch(e){}
        }

        const deliveryStorageKey = "delivery_calc_" + sid;
        const deliveryModal = document.getElementById("deliveryModal");
        const deliveryBackdrop = document.getElementById("deliveryBackdrop");
        const openDeliveryCalcBtn = document.getElementById("openDeliveryCalcBtn");
        const closeDeliveryCalcBtn = document.getElementById("closeDeliveryCalcBtn");
        const deliveryRateEls = Array.from(document.querySelectorAll(".del-rate"));
        const deliveryQtyEls = Array.from(document.querySelectorAll(".del-qty"));
        const deliveryTotalEls = Array.from(document.querySelectorAll(".del-total"));
        const deliveryGrandTotalEl = document.getElementById("deliveryGrandTotal");
        const deliveryConsumeAmountEl = document.getElementById("deliveryConsumeAmount");
        const deliveryConsumeNoteEl = document.getElementById("deliveryConsumeNote");
        const deliveryNetTotalEl = document.getElementById("deliveryNetTotal");
        const applyDeliveryExpenseBtn = document.getElementById("applyDeliveryExpenseBtn");
        const expenseFormEl = document.getElementById("expenseForm");
        const expenseCategoryEl = document.getElementById("expenseCategory");
        const expenseAmountEl = document.getElementById("expenseAmount");
        const expenseNoteEl = document.getElementById("expenseNote");

        const deliveryPresets = {{ delivery_shift_presets_json|safe }};
        const deliveryInitialData = {{ delivery_payload_json|safe }};
        const deliveryHourShiftEl = document.getElementById("deliveryHourShift");
        const deliveryHourInEl = document.getElementById("deliveryHourIn");
        const deliveryHourOutEl = document.getElementById("deliveryHourOut");
        const deliveryHoursLabelEl = document.getElementById("deliveryHoursLabel");

        function getDeliveryDefaults(){
          return JSON.parse(JSON.stringify(deliveryInitialData || {
            rates: [1500, 2000, 2500, 3000, 2500],
            qtys: [0, 0, 0, 0, 0],
            consume_amount: 0,
            consume_note: "",
            hour_shift: "MORNING",
            hour_in: "08:55",
            hour_out: "12:50"
          }));
        }

        function calcHourQty(){
          const hIn = (deliveryHourInEl && deliveryHourInEl.value) ? deliveryHourInEl.value : "";
          const hOut = (deliveryHourOutEl && deliveryHourOutEl.value) ? deliveryHourOutEl.value : "";
          if(!hIn || !hOut) return 0;
          const [ih, im] = hIn.split(":").map(v => parseInt(v, 10));
          const [oh, om] = hOut.split(":").map(v => parseInt(v, 10));
          if([ih, im, oh, om].some(v => isNaN(v))) return 0;
          const mins = ((oh * 60 + om) - (ih * 60 + im));
          if(mins < 0) return 0;
          return Math.round((mins / 60) * 100) / 100;
        }

        function updateHoursFromTimes(){
          const hours = calcHourQty();
          if(deliveryQtyEls[4]) deliveryQtyEls[4].value = hours.toFixed(2);
          if(deliveryHoursLabelEl) deliveryHoursLabelEl.textContent = hours.toFixed(2) + " hs";
          return hours;
        }

        function applyPresetForShift(force=false){
          if(!deliveryHourShiftEl) return;
          const shiftKey = deliveryHourShiftEl.value === "AFTERNOON" ? "AFTERNOON" : "MORNING";
          const preset = deliveryPresets[shiftKey] || deliveryPresets["MORNING"];
          if(force || !deliveryHourInEl.value) deliveryHourInEl.value = preset.hour_in || "";
          if(force || !deliveryHourOutEl.value) deliveryHourOutEl.value = preset.hour_out || "";
          updateHoursFromTimes();
        }

        let deliverySaveTimer = null;
        function persistDeliveryCalc(){
          saveDeliveryCalc();
          if(deliverySaveTimer) clearTimeout(deliverySaveTimer);
          deliverySaveTimer = setTimeout(async () => {
            try{
              await fetch("{{ url_for('save_delivery_draft', id=s.id) }}", {
                method: "POST",
                headers: {"Content-Type": "application/json"},
                body: JSON.stringify({
                  rates: deliveryRateEls.map(el => toInt(el.value)),
                  qtys: deliveryQtyEls.map((el, i) => i === 4 ? toNum(el.value) : toInt(el.value)),
                  consume_amount: toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0),
                  consume_note: deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "") : "",
                  hour_shift: deliveryHourShiftEl ? deliveryHourShiftEl.value : "MORNING",
                  hour_in: deliveryHourInEl ? deliveryHourInEl.value : "",
                  hour_out: deliveryHourOutEl ? deliveryHourOutEl.value : ""
                })
              });
            }catch(e){}
          }, 350);
        }

        function saveDeliveryCalc(){
          try{
            const payload = {
              rates: deliveryRateEls.map(el => toInt(el.value)),
              qtys: deliveryQtyEls.map((el, i) => i === 4 ? toNum(el.value) : toInt(el.value)),
              consume_amount: toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0),
              consume_note: deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "") : "",
              hour_shift: deliveryHourShiftEl ? deliveryHourShiftEl.value : "MORNING",
              hour_in: deliveryHourInEl ? deliveryHourInEl.value : "",
              hour_out: deliveryHourOutEl ? deliveryHourOutEl.value : ""
            };
            localStorage.setItem(deliveryStorageKey, JSON.stringify(payload));
          }catch(e){}
        }

        function loadDeliveryCalc(){
          const defaults = getDeliveryDefaults();
          try{
            const raw = localStorage.getItem(deliveryStorageKey);
            const payload = raw ? (JSON.parse(raw) || {}) : defaults;
            deliveryRateEls.forEach((el, i) => {
              el.value = Array.isArray(payload.rates) && payload.rates[i] != null ? payload.rates[i] : defaults.rates[i];
            });
            deliveryQtyEls.forEach((el, i) => {
              el.value = Array.isArray(payload.qtys) && payload.qtys[i] != null ? payload.qtys[i] : defaults.qtys[i];
            });
            if(deliveryConsumeAmountEl){
              deliveryConsumeAmountEl.value = payload.consume_amount != null ? payload.consume_amount : defaults.consume_amount;
            }
            if(deliveryConsumeNoteEl){
              deliveryConsumeNoteEl.value = payload.consume_note != null ? payload.consume_note : defaults.consume_note;
            }
            if(deliveryHourShiftEl){
              deliveryHourShiftEl.value = payload.hour_shift || defaults.hour_shift || "MORNING";
            }
            if(deliveryHourInEl){
              deliveryHourInEl.value = payload.hour_in || defaults.hour_in || "";
            }
            if(deliveryHourOutEl){
              deliveryHourOutEl.value = payload.hour_out || defaults.hour_out || "";
            }
            updateHoursFromTimes();
          }catch(e){
            deliveryRateEls.forEach((el, i) => el.value = defaults.rates[i]);
            deliveryQtyEls.forEach((el, i) => el.value = defaults.qtys[i]);
            if(deliveryConsumeAmountEl) deliveryConsumeAmountEl.value = defaults.consume_amount;
            if(deliveryConsumeNoteEl) deliveryConsumeNoteEl.value = defaults.consume_note;
            if(deliveryHourShiftEl) deliveryHourShiftEl.value = defaults.hour_shift || "MORNING";
            if(deliveryHourInEl) deliveryHourInEl.value = defaults.hour_in || "";
            if(deliveryHourOutEl) deliveryHourOutEl.value = defaults.hour_out || "";
            updateHoursFromTimes();
          }
        }

        function recalcDeliveryCalc(){
          updateHoursFromTimes();
          let grand = 0;
          deliveryRateEls.forEach((el, i) => {
            const rate = toNum(el.value);
            const qty = (i === 4) ? toNum(deliveryQtyEls[i].value) : toInt(deliveryQtyEls[i].value);
            const total = rate * qty;
            grand += total;
            if(deliveryTotalEls[i]){
              deliveryTotalEls[i].textContent = fmtMoney(Math.round(total));
            }
          });
          const consumeAmount = toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0);
          const net = Math.max(0, Math.round(grand - consumeAmount));
          if(deliveryGrandTotalEl){
            deliveryGrandTotalEl.textContent = fmtMoney(Math.round(grand));
          }
          if(deliveryNetTotalEl){
            deliveryNetTotalEl.textContent = fmtMoney(Math.round(net));
          }
        }

        function openDeliveryCalc(){
          if(deliveryModal) deliveryModal.style.display = "block";
          if(deliveryBackdrop) deliveryBackdrop.style.display = "block";
          document.body.style.overflow = "hidden";
        }

        function closeDeliveryCalc(){
          if(deliveryModal) deliveryModal.style.display = "none";
          if(deliveryBackdrop) deliveryBackdrop.style.display = "none";
          document.body.style.overflow = "";
        }

        if(openDeliveryCalcBtn){
          openDeliveryCalcBtn.addEventListener("click", openDeliveryCalc);
        }
        if(closeDeliveryCalcBtn){
          closeDeliveryCalcBtn.addEventListener("click", closeDeliveryCalc);
        }
        if(deliveryBackdrop){
          deliveryBackdrop.addEventListener("click", closeDeliveryCalc);
        }
        document.addEventListener("keydown", (ev) => {
          if(ev.key === "Escape" && deliveryModal && deliveryModal.style.display === "block"){
            closeDeliveryCalc();
          }
        });

        [...deliveryRateEls, ...deliveryQtyEls, deliveryConsumeAmountEl, deliveryConsumeNoteEl, deliveryHourInEl, deliveryHourOutEl].filter(Boolean).forEach(el => {
          el.addEventListener("input", () => {
            persistDeliveryCalc();
            recalcDeliveryCalc();
          });
        });
        if(deliveryHourShiftEl){
          deliveryHourShiftEl.addEventListener("change", () => {
            applyPresetForShift(true);
            persistDeliveryCalc();
            recalcDeliveryCalc();
          });
        }

        if(applyDeliveryExpenseBtn){
          applyDeliveryExpenseBtn.addEventListener("click", () => {
            const grand = deliveryRateEls.reduce((acc, el, i) => {
              const rate = toNum(el.value);
              const qty = (i === 4) ? toNum(deliveryQtyEls[i].value) : toInt(deliveryQtyEls[i].value);
              return acc + (rate * qty);
            }, 0);
            const consumeAmount = toInt(deliveryConsumeAmountEl ? deliveryConsumeAmountEl.value : 0);
            const net = Math.max(0, Math.round(grand - consumeAmount));
            const consumeNote = deliveryConsumeNoteEl ? (deliveryConsumeNoteEl.value || "").trim() : "";

            if(!expenseFormEl || !expenseCategoryEl || !expenseAmountEl || net <= 0){
              alert("El total a cargar debe ser mayor a 0.");
              return;
            }

            expenseCategoryEl.value = "Delivery / Cadete (efectivo)";
            expenseAmountEl.value = net;

            if(expenseNoteEl){
              let note = "Calculadora delivery";
              if(consumeAmount > 0){
                note += " - consumo descontado: " + fmtMoney(consumeAmount);
              }
              if(consumeNote){
                note += " - " + consumeNote;
              }
              expenseNoteEl.value = note;
            }

            persistDeliveryCalc();
            expenseFormEl.submit();
          });
        }

        function loadDraft(){
          try{
            const raw = localStorage.getItem("caja_draft_"+sid);
            if(!raw) return;
            const p = JSON.parse(raw);
            if(p && typeof p === "object"){
              if(retiradoEl.value === "" || retiradoEl.value === "0") retiradoEl.value = p.retirado ?? retiradoEl.value;
              if(cajaFinalEl.value === "" || cajaFinalEl.value === "0") cajaFinalEl.value = p.caja_final ?? cajaFinalEl.value;
              if(mpEl.value === "" || mpEl.value === "0") mpEl.value = p.mp ?? mpEl.value;
              if(pyaEl.value === "" || pyaEl.value === "0") pyaEl.value = p.pya ?? pyaEl.value;
              if(rappiEl.value === "" || rappiEl.value === "0") rappiEl.value = p.rappi ?? rappiEl.value;
            }
          }catch(e){}
        }

        function recalc(){
          const retirado = toInt(retiradoEl.value);
          const cajaFinal = toInt(cajaFinalEl.value);
          const mp = toInt(mpEl.value);
          const pya = toInt(pyaEl.value);
          const rappi = toInt(rappiEl.value);

          const efectivoBruto = retirado + cajaFinal;
          const efectivoNeto = retirado - {{ s.opening_cash|int }};

          const ingresoBruto = (efectivoBruto + mp + pya + rappi) + egresosTotal;
          const ingresoNeto = ingresoBruto - egresosTotal; // pedido: neto = bruto - egresos

          efBrutoLbl.textContent = fmtMoney(efectivoBruto);
          efNetoLbl.textContent = fmtMoney(efectivoNeto);
          ingresoBrutoLbl.textContent = fmtMoney(ingresoBruto);
          ingresoNetoLbl.textContent = fmtMoney(ingresoNeto);
          efDispLbl.textContent = fmtMoney(retirado);
        }

        [retiradoEl, cajaFinalEl, mpEl, pyaEl, rappiEl].forEach(el => {
          el.addEventListener("input", () => { saveDraft(); recalc(); });
        });

        // Restore draft after page reload (ej: al agregar egresos)
        loadDeliveryCalc();
        applyPresetForShift(false);
        recalcDeliveryCalc();
        loadDraft();
        recalc();

        // Submit: enviar valores actuales
        document.getElementById("saveForm").addEventListener("submit", (e) => {
          document.getElementById("h_sales_cash").value = toInt(retiradoEl.value);
          document.getElementById("h_cash_final").value = toInt(cajaFinalEl.value);
          document.getElementById("h_sales_mp").value = toInt(mpEl.value);
          document.getElementById("h_sales_pya").value = toInt(pyaEl.value);
          document.getElementById("h_sales_rappi").value = toInt(rappiEl.value);
          // no borramos draft: por si vuelve atras
        });
      </script>

    {% endif %}

    </body>

    </html>
    """

    EDIT_ALL_HTML = """
    <!doctype html>
    <html lang="es">
    <head><meta charset="utf-8"><title>Editar turno</title>
    {{ base_css|safe }}
    <style>
      body{max-width:1200px;}
      input,select,button,textarea{padding:8px; font-size:14px;}
      .row{display:grid; grid-template-columns:repeat(4, 1fr); gap:10px;}
      .row2{display:grid; grid-template-columns:1fr 2fr; gap:10px;}
      .right{text-align:right;}
      .holiday-bg{background:#ffe3ec;}
      .holiday-bg table tr{background:#fff;}
    </style>
    </head>
    <body>

    <a href="{{ url_for('shift', id=s.id, d=d) }}"><- Volver</a>
    <h2 style="margin-bottom:6px;">Editar (todo) - {{turn_name}} {{s.day}}</h2>
    <p class="muted">Usuario: <b>{{user.username}}</b> ({{user.role}}) - Ediciones previas: {{close.edit_count}}</p>

    <form method="post">
      <div class="box">
        <h3>Datos del turno</h3>
        <div class="row">
          <div>
            <label>Responsable</label><br>
            <input type="text" name="responsible" value="{{s.responsible}}" required style="width:100%;">
          </div>
          <div>
            <label>Caja inicial</label><br>
            <input type="number" name="opening_cash" min="0" value="{{s.opening_cash}}" required style="width:100%;">
          </div>
          <div>
            <label>Turno</label><br>
            <input type="text" value="{{turn_name}}" readonly style="width:100%;">
          </div>
          <div>
            <label>Fecha</label><br>
            <input type="text" value="{{s.day}}" readonly style="width:100%;">
          </div>
        </div>
        <p class="muted">Cambiar CI impacta calculos del turno siguiente.</p>
      </div>

      <div class="box">
        <h3>Ventas</h3>
        <div class="row">
          <div>Ventas efectivo (neto)<br><input type="number" name="sales_cash" min="0" value="{{s.sales_cash}}" style="width:100%;"></div>
          <div>Ventas Mercado Pago<br><input type="number" name="sales_mp" min="0" value="{{s.sales_mp}}" style="width:100%;"></div>
          <div>Ventas Pedidos Ya<br><input type="number" name="sales_pya" min="0" value="{{s.sales_pya}}" style="width:100%;"></div>
          <div>Ventas Rappi<br><input type="number" name="sales_rappi" min="0" value="{{s.sales_rappi}}" style="width:100%;"></div>
        </div>
      </div>

      <div class="box">
        <h3>Egresos (editar / borrar / agregar)</h3>
        <table>
          <tr><th>Eliminar</th><th>Categoria</th><th class="right">Monto</th><th>Nota</th></tr>
          {% for e in expenses %}
            <tr>
              <td><input type="checkbox" name="del_{{e.id}}"></td>
              <td>
                <select name="cat_{{e.id}}">
                  {% for c in categories %}
                    <option value="{{c}}" {% if e.category==c %}selected{% endif %}>{{c}}</option>
                  {% endfor %}
                </select>
              </td>
              <td class="right"><input type="number" name="amt_{{e.id}}" min="0" value="{{e.amount}}" style="width:120px;"></td>
              <td><input type="text" name="note_{{e.id}}" value="{{e.note or ''}}" style="width:100%;"></td>
            </tr>
          {% endfor %}
          {% if not expenses %}
            <tr><td colspan="4" class="muted">Sin egresos.</td></tr>
          {% endif %}
        </table>

        <h4 style="margin-top:12px;">Agregar egresos (hasta 5)</h4>
        <p class="muted" style="margin-top:-6px;">Completa categoria y monto para cada egreso que quieras agregar.</p>
        <div class="grid">
          {% for i in range(1,6) %}
          <div class="row2" style="align-items:center;">
            <div>
              <select name="new_category_{{i}}">
                <option value="">(no agregar)</option>
                {% for c in categories %}<option value="{{c}}">{{c}}</option>{% endfor %}
              </select>
            </div>
            <div style="display:flex; gap:10px; align-items:center; width:100%;">
              <input type="number" name="new_amount_{{i}}" min="0" placeholder="$" style="width:140px;">
              <input type="text" name="new_note_{{i}}" placeholder="nota (obligatoria si Otros)" style="flex:1;">
            </div>
          </div>
          {% endfor %}
        </div>
      </div>

      <div class="box">
        
        <h3>Cierre</h3>
        <div class="row">
          <div>
            <label>Retirado (Efectivo total)</label><br>
            <input type="number" name="withdrawn" min="0" value="{{close.withdrawn_cash}}" required style="width:100%;">
          </div>
          <div>
            <label>Caja final</label><br>
            <input type="number" name="ending_real" min="0" value="{{close.ending_cash}}" required style="width:100%;">
          </div>
          <div>
            <label class="muted">Estado</label><br>
            <input type="text" value="OK" readonly style="width:100%;">
          </div>
          <div>
            <label class="muted">Observacion</label><br>
            <input type="text" value="{{close.note or ''}}" readonly style="width:100%;">
          </div>
        </div>
        <p class="muted">La diferencia se mantiene en 0 (sin validacion OK/NO OK por ahora).</p>

      </div>

      <div class="box">
        <h3>Auditoria</h3>
        <div class="row2">
          <div><b>Motivo de edicion (obligatorio)</b></div>
          <div><input type="text" name="reason" required placeholder="Ej: se cargo venta luego del cierre" style="width:100%;"></div>
        </div>
      </div>

      <button class="btn">Guardar edicion</button>
    </form>

    </body>
    </html>
    """


    def is_placeholder_shift(s: "Shift") -> bool:
        """Devuelve True si el turno existe en BD pero en la practica esta 'vacio'
        (tipico de importaciones/placeholder): todo en 0, sin egresos, y (si existe) cierre vacio.
        En ese caso, en el panel principal lo tratamos como NO ABIERTO para mostrar el boton 'Abrir'.
        """
        if not s:
            return True

        # Solo consideramos placeholder a turnos CERRADOS en cero (evita ocultar un turno abierto recien iniciado).
        if (s.status or "").upper() != "CLOSED":
            return False

        nums = [
            int(s.opening_cash or 0),
            int(s.sales_cash or 0),
            int(s.sales_mp or 0),
            int(s.sales_pya or 0),
            int(s.sales_rappi or 0),
            int(getattr(s, "sales_apps", 0) or 0),
        ]
        if any(v != 0 for v in nums):
            return False

        # Si tiene egresos, NO es placeholder
        if CashExpense.query.filter_by(shift_id=s.id).first():
            return False

        c = ShiftClose.query.filter_by(shift_id=s.id).first()
        if c:
            close_nums = [
                int(c.withdrawn_cash or 0),
                int(c.ending_calc or 0),
                int(c.ending_cash or 0),
                int(c.difference or 0),
            ]
            if any(v != 0 for v in close_nums):
                return False
            if (c.note or "").strip():
                return False

        return True


    @app.route("/caja")
    @login_required
    def caja_index():
        day_obj, d = parse_day_param(request.args.get("d"))

        start_date, end_date = parse_range_params(request.args.get("start"), request.args.get("end"))
        start = start_date.isoformat()
        end = end_date.isoformat()
        turn = request.args.get("turn") or "ALL"
        resp = request.args.get("resp") or "ALL"

        shifts = {s.turn: s for s in Shift.query.filter_by(day=day_obj).all()}

        # Si por importacion o error existe un turno 'cerrado' totalmente en cero, lo tratamos como NO ABIERTO.
        # (Asi el usuario ve el boton 'Abrir' en vez de 'Ver/Editar'.)
        for _t, _s in list(shifts.items()):
            if _s and is_placeholder_shift(_s):
                shifts[_t] = None

        closes = {}
        can_edit_map = {}
        u = current_user()

        for s in shifts.values():
            if not s:
                continue
            if s.status == "CLOSED":
                c = ShiftClose.query.filter_by(shift_id=s.id).first()
                closes[s.id] = c
                can_edit_map[s.id] = bool(c) and can_edit_close(u, c)

        ventas_netas_map = {}
        efectivo_disp_map = {}

        for s in shifts.values():
            if not s:
                continue
            if s.status == "CLOSED":
                c = closes.get(s.id)
                cash_final = int(getattr(c, 'ending_cash', 0) or 0) if c else 0
                eg = expenses_total(s.id, CashExpense)
                ventas_netas_map[s.id] = int(calc_ingreso_neto(s, egresos=eg, cash_final=cash_final))
                c = closes.get(s.id)
                efectivo_disp_map[s.id] = int(s.sales_cash or 0)

        efectivo_total_dia = None
        ingreso_neto_total_dia = None
        if shifts.get("MORNING") and shifts.get("AFTERNOON"):
            if shifts["MORNING"].status == "CLOSED" and shifts["AFTERNOON"].status == "CLOSED":
                efectivo_total_dia = (
                    efectivo_disp_map.get(shifts["MORNING"].id, 0) +
                    efectivo_disp_map.get(shifts["AFTERNOON"].id, 0)
                )
                ingreso_neto_total_dia = (ventas_netas_map.get(shifts["MORNING"].id, 0) + ventas_netas_map.get(shifts["AFTERNOON"].id, 0))

        def is_valid_responsible(name: str) -> bool:
            name = (name or "").strip()
            if not name:
                return False
            if name in EMPLOYEES or name in ADMINS:
                return True
            return False

        responsibles = sorted({
            s.responsible for s in Shift.query.filter(Shift.status == "CLOSED").all()
            if is_valid_responsible(s.responsible)
        })

        summary = get_caja_summary(
            limit=400,
            start=start_date,
            end=end_date,
            turn=turn,
            responsible=resp
        )

        # Agrupar visualmente por fecha en "Resumen de cierres"
        grouped_summary = []
        i = 0
        while i < len(summary):
            day_key = summary[i]["day"]
            j = i
            day_total_neto = 0
            while j < len(summary) and summary[j]["day"] == day_key:
                day_total_neto += int(summary[j].get("ventas_neto", 0) or 0)
                j += 1

            rowspan = j - i
            for k in range(i, j):
                row = dict(summary[k])
                row["day_rowspan"] = rowspan
                row["show_day"] = (k == i)
                row["show_day_total"] = (k == i)
                row["day_total_neto"] = day_total_neto
                grouped_summary.append(row)
            i = j

        summary = grouped_summary

        weekday_name = weekday_es(day_obj)

        return render_template_string(
            CAJA_INDEX_HTML,
            base_css=BASE_CSS,
            d=d,
            turns=TURNS,
            shifts=shifts,
            closes=closes,
            can_edit_map=can_edit_map,
            ventas_netas_map=ventas_netas_map,
            efectivo_disp_map=efectivo_disp_map,
            efectivo_total_dia=efectivo_total_dia,
            ingreso_neto_total_dia=ingreso_neto_total_dia,
            summary=summary,
            db_path=DB_PATH,
            start=start,
            end=end,
            turn=turn,
            resp=resp,
            responsibles=responsibles,
            weekday_name=weekday_name
        )

    @app.route("/caja/open/<turn>", methods=["GET","POST"])
    @login_required
    def open_shift(turn):
        day_obj, d = parse_day_param(request.args.get("d") or request.form.get("d"))

        existing = Shift.query.filter_by(day=day_obj, turn=turn).first()
        if existing:
            # Si existe como placeholder (cerrado y todo en 0), lo limpiamos para permitir "Abrir" correctamente.
            if is_placeholder_shift(existing):
                try:
                    ShiftClose.query.filter_by(shift_id=existing.id).delete()
                    CashExpense.query.filter_by(shift_id=existing.id).delete()
                    db.session.delete(existing)
                    db.session.commit()
                    backup_caja_local_y_drive()
                except Exception:
                    db.session.rollback()
                existing = None
            else:
                return redirect(url_for("shift", id=existing.id, d=d))

        locked_opening = get_locked_opening_cash(day_obj, turn)
        u = current_user()
        default_responsible = u.username if u else ""

        if request.method == "POST":
            responsible = (request.form.get("responsible") or "").strip()
            opening_cash = safe_int(request.form.get("opening_cash"))

            if opening_cash is None:
                return redirect(url_for("open_shift", turn=turn, d=d))

            # Responsible can be omitted (or was imported wrong). Default to current user / Bernardo.
            if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
                responsible = default_responsible or "Bernardo"

            if locked_opening is not None and int(opening_cash) != int(locked_opening):
                return redirect(url_for("open_shift", turn=turn, d=d))

            s = Shift(day=day_obj, turn=turn, responsible=responsible, opening_cash=int(opening_cash))
            db.session.add(s)
            db.session.commit()
            backup_caja_local_y_drive()
            return redirect(url_for("shift", id=s.id, d=d))

        return render_template_string(
            OPEN_HTML,
            base_css=BASE_CSS,
            d=d,
            day=day_obj.isoformat(),
            turn_name=TURN_NAMES.get(turn, turn),
            locked_opening=locked_opening,
            default_responsible=default_responsible
        )

    @app.route("/caja/shift/<int:id>")
    @login_required
    def shift(id):
        d = request.args.get("d") or date.today().isoformat()
        s = Shift.query.get_or_404(id)

        expenses = CashExpense.query.filter_by(shift_id=id).order_by(CashExpense.created_at.asc()).all()
        close_row = ShiftClose.query.filter_by(shift_id=id).first()

        egresos_total = expenses_total(id, CashExpense)
        cash_final = int(getattr(close_row, 'ending_cash', 0) or 0) if close_row else 0
        ingreso_bruto = calc_ingreso_bruto(s, egresos_total, cash_final=cash_final)
        ingreso_neto = calc_ingreso_neto(s, egresos=egresos_total, cash_final=cash_final)

        efectivo_disponible = int(s.sales_cash or 0)

        u = current_user()
        can_edit = bool(close_row) and can_edit_close(u, close_row)

        delivery_payload = build_delivery_payload(s)

        return render_template_string(
            SHIFT_HTML,
            base_css=BASE_CSS,
            s=s,
            expenses=expenses,
            close=close_row,
            categories=CATEGORIES,
            turn_name=TURN_NAMES.get(s.turn, s.turn),
            d=d,
            egresos_total=egresos_total,
            ingreso_neto=ingreso_neto,
            ingreso_bruto=ingreso_bruto,
            efectivo_disponible=efectivo_disponible,
            can_edit=can_edit,
            delivery_payload_json=json.dumps(delivery_payload),
            delivery_shift_presets_json=json.dumps(DELIVERY_SHIFT_PRESETS)
        )

    @app.route("/caja/shift/<int:id>/delivery_draft", methods=["POST"])
    @login_required
    def save_delivery_draft(id):
        s = Shift.query.get_or_404(id)
        if s.status != "OPEN":
            return {"ok": False, "error": "Turno cerrado"}, 400

        payload = request.get_json(silent=True) or {}
        hour_shift = (payload.get("hour_shift") or "MORNING").strip().upper()
        if hour_shift not in DELIVERY_SHIFT_PRESETS:
            hour_shift = "MORNING"

        hour_in = (payload.get("hour_in") or "").strip()
        hour_out = (payload.get("hour_out") or "").strip()
        if hour_in and not valid_time_str(hour_in):
            return {"ok": False, "error": "Hora de entrada invalida"}, 400
        if hour_out and not valid_time_str(hour_out):
            return {"ok": False, "error": "Hora de salida invalida"}, 400

        rates = payload.get("rates") if isinstance(payload.get("rates"), list) else [1500, 2000, 2500, 3000, 2500]
        qtys = payload.get("qtys") if isinstance(payload.get("qtys"), list) else [0, 0, 0, 0, 0]
        rates = (rates + [0, 0, 0, 0, 0])[:5]
        qtys = (qtys + [0, 0, 0, 0, 0])[:5]

        safe_rates = []
        for v in rates:
            try:
                safe_rates.append(int(float(v or 0)))
            except Exception:
                safe_rates.append(0)

        hours_qty = delivery_hours_decimal(hour_in, hour_out)
        safe_qtys = []
        for i, v in enumerate(qtys):
            if i == 4:
                safe_qtys.append(hours_qty)
            else:
                try:
                    safe_qtys.append(int(float(v or 0)))
                except Exception:
                    safe_qtys.append(0)

        try:
            consume_amount = int(float(payload.get("consume_amount") or 0))
        except Exception:
            consume_amount = 0
        consume_note = str(payload.get("consume_note") or "")[:200]

        clean_payload = {
            "rates": safe_rates,
            "qtys": safe_qtys,
            "consume_amount": consume_amount,
            "consume_note": consume_note,
            "hour_shift": hour_shift,
            "hour_in": hour_in,
            "hour_out": hour_out,
        }

        s.hour_shift = hour_shift
        s.hour_in = hour_in or None
        s.hour_out = hour_out or None
        s.delivery_data_json = json.dumps(clean_payload, ensure_ascii=False)
        db.session.commit()
        backup_caja_local_y_drive()
        return {"ok": True, "hours": safe_qtys[4]}

    @app.route("/caja/shift/<int:id>/sales", methods=["POST"])
    @login_required
    def sales(id):
        d = request.args.get("d") or date.today().isoformat()
        s = Shift.query.get_or_404(id)
        if s.status != "OPEN":
            return redirect(url_for("shift", id=id, d=d))

        s.sales_cash = to_int(request.form.get("sales_cash"))
        s.sales_mp = to_int(request.form.get("sales_mp"))
        s.sales_pya = to_int(request.form.get("sales_pya"))
        s.sales_rappi = to_int(request.form.get("sales_rappi"))

        db.session.commit()
        backup_caja_local_y_drive()
        return redirect(url_for("shift", id=id, d=d))

    @app.route("/caja/shift/<int:id>/expense", methods=["POST"])
    @login_required
    def expense(id):
        d = request.args.get("d") or date.today().isoformat()
        s = Shift.query.get_or_404(id)
        if s.status != "OPEN":
            return redirect(url_for("shift", id=id, d=d))

        category = (request.form.get("category") or "").strip()
        amount = safe_int(request.form.get("amount"))
        note = (request.form.get("note") or "").strip()

        if category not in CATEGORIES or amount is None or amount <= 0:
            return redirect(url_for("shift", id=id, d=d))

        if category.startswith("Otros") and not note:
            return redirect(url_for("shift", id=id, d=d))

        e = CashExpense(shift_id=id, category=category, amount=int(amount), note=note or None)
        db.session.add(e)
        db.session.commit()
        backup_caja_local_y_drive()
        return redirect(url_for("shift", id=id, d=d))


    @app.route("/caja/shift/<int:id>/close", methods=["POST"])
    @login_required
    def close_shift(id):
        d = request.args.get("d") or date.today().isoformat()
        s = Shift.query.get_or_404(id)
        if s.status != "OPEN":
            return redirect(url_for("shift", id=id, d=d))

        # Nuevo modelo:
        # - Retirado (efectivo total): se carga manualmente -> lo guardamos en Shift.sales_cash
        # - Caja final: se carga manualmente -> lo guardamos en ShiftClose.ending_cash
        retirado = safe_int(request.form.get("sales_cash"))
        caja_final = safe_int(request.form.get("cash_final"))

        if retirado is None or retirado < 0:
            return redirect(url_for("shift", id=id, d=d))
        if caja_final is None or caja_final < 0:
            return redirect(url_for("shift", id=id, d=d))

        # Guardar ventas (mismos nombres para no romper import/export)
        s.sales_cash = int(retirado)
        s.sales_mp = to_int(request.form.get("sales_mp"))
        s.sales_pya = to_int(request.form.get("sales_pya"))
        s.sales_rappi = to_int(request.form.get("sales_rappi"))

        # Cierre (sin estado OK/NO OK por ahora)
        ending_calc = calc_ending_calc(int(caja_final), int(retirado))
        ending_real = int(caja_final)
        difference = 0

        c = ShiftClose(
            shift_id=id,
            withdrawn_cash=int(retirado),
            ending_calc=int(ending_calc),
            ending_cash=int(ending_real),
            difference=int(difference),
            note=None,
            close_ok=1,
            edit_count=0
        )
        s.status = "CLOSED"
        s.closed_at = datetime.utcnow()
        db.session.add(c)
        db.session.commit()
        backup_caja_local_y_drive()
        return redirect(url_for("caja_index", d=d))




    @app.route("/caja/shift/<int:id>/edit_all", methods=["GET","POST"])
    @login_required
    def edit_all(id):
        d = request.args.get("d") or date.today().isoformat()
        s = Shift.query.get_or_404(id)
        close_row = ShiftClose.query.filter_by(shift_id=id).first()
        if not close_row or s.status != "CLOSED":
            abort(404)

        u = current_user()
        if not can_edit_close(u, close_row):
            abort(403)

        expenses = CashExpense.query.filter_by(shift_id=id).order_by(CashExpense.created_at.asc()).all()

        if request.method == "POST":
            reason = (request.form.get("reason") or "").strip()
            if not reason:
                return redirect(url_for("edit_all", id=id, d=d))

            responsible = (request.form.get("responsible") or "").strip()
            opening_cash = safe_int(request.form.get("opening_cash"))
            if opening_cash is None:
                return redirect(url_for("edit_all", id=id, d=d))

            if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
                responsible = (u.username if u else None) or "Bernardo"

            s.responsible = responsible
            s.opening_cash = int(opening_cash)

            s.sales_cash = to_int(request.form.get("sales_cash"))
            s.sales_mp = to_int(request.form.get("sales_mp"))
            s.sales_pya = to_int(request.form.get("sales_pya"))
            s.sales_rappi = to_int(request.form.get("sales_rappi"))

            for e in expenses:
                if request.form.get(f"del_{e.id}") == "on":
                    db.session.delete(e)
                    continue

                cat = (request.form.get(f"cat_{e.id}") or "").strip()
                amt = safe_int(request.form.get(f"amt_{e.id}"))
                note = (request.form.get(f"note_{e.id}") or "").strip()

                if cat not in CATEGORIES:
                    return redirect(url_for("edit_all", id=id, d=d))
                if amt is None or amt < 0:
                    return redirect(url_for("edit_all", id=id, d=d))
                if cat.startswith("Otros") and not note:
                    return redirect(url_for("edit_all", id=id, d=d))

                e.category = cat
                e.amount = int(amt)
                e.note = note or None

            # Agregar multiples egresos nuevos (hasta 5)
            for i in range(1, 6):
                new_cat = (request.form.get(f"new_category_{i}") or "").strip()
                new_amt = safe_int(request.form.get(f"new_amount_{i}"))
                new_note = (request.form.get(f"new_note_{i}") or "").strip()
                if not new_cat:
                    continue
                if new_cat not in CATEGORIES:
                    return redirect(url_for("edit_all", id=id, d=d))
                if new_amt is None or new_amt <= 0:
                    return redirect(url_for("edit_all", id=id, d=d))
                if new_cat.startswith("Otros") and not new_note:
                    return redirect(url_for("edit_all", id=id, d=d))
                db.session.add(CashExpense(
                    shift_id=id,
                    category=new_cat,
                    amount=int(new_amt),
                    note=new_note or None
                ))

            db.session.flush()

            withdrawn = safe_int(request.form.get("withdrawn"))
            ending_real = safe_int(request.form.get("ending_real"))

            if withdrawn is None or withdrawn < 0 or ending_real is None or ending_real < 0:
                return redirect(url_for("edit_all", id=id, d=d))

            # Nuevo modelo: caja final se carga manualmente
            ending_calc = calc_ending_calc(int(ending_real), int(withdrawn))
            close_ok = 1
            note_to_save = None
            difference = 0

            close_row.withdrawn_cash = int(withdrawn)
            close_row.ending_calc = int(ending_calc)
            close_row.ending_cash = int(ending_real)
            close_row.difference = int(difference)
            close_row.close_ok = close_ok
            close_row.note = note_to_save

            close_row.edit_count = int(close_row.edit_count or 0) + 1
            close_row.edited_by = u.username
            close_row.edited_at = datetime.utcnow()
            close_row.edit_reason = reason

            db.session.commit()
            backup_caja_local_y_drive()
            return redirect(url_for("shift", id=id, d=d))

        return render_template_string(
            EDIT_ALL_HTML,
            base_css=BASE_CSS,
            s=s,
            close=close_row,
            expenses=expenses,
            categories=CATEGORIES,
            user=u,
            d=d,
            turn_name=TURN_NAMES.get(s.turn, s.turn)
        )

    @app.route("/caja/export/excel")
    @login_required
    def export_excel():
        data = get_caja_summary(
            db,
            Shift,
            ShiftClose,
            CashExpense,
            limit=10000,
            weekday_es=weekday_es,
            responsible_name=responsible_name,
            turn_names=TURN_NAMES,
            expenses_total=expenses_total,
            calc_ingreso_bruto=calc_ingreso_bruto,
            calc_ingreso_neto=calc_ingreso_neto,
            cash_bruto=cash_bruto,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Cierres"

        ws.append([
            "Fecha","Turno","Responsable",
            "Caja Inicial","Efectivo bruto","Ventas Mercado Pago","Ventas Pedidos Ya","Ventas Rappi",
            "Egresos","Ventas Totales","Ventas Netas","Retirado","Caja Final (Real)","Diferencia"
        ])

        for r in data:
            ws.append([
                r["day"], r["turn_name"], r["responsible"],
                r["opening_cash"], r["sales_cash"], r["sales_mp"], r["sales_pya"], r["sales_rappi"],
                r["expenses"], r["ventas_bruto"], r["ventas_neto"],
                r["withdrawn"], r["ending_real"], r["difference"]
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"cierres_caja_{date.today().isoformat()}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/caja/export/json")
    @login_required
    def export_caja_json():
        # Backup JSON de Caja: shifts + expenses + close
        start_raw = request.args.get("start")
        end_raw = request.args.get("end")
        start_date, end_date = parse_range_params(start_raw, end_raw)

        shifts = Shift.query.order_by(Shift.day.asc(), Shift.turn.asc()).all()
        expenses = CashExpense.query.order_by(CashExpense.created_at.asc()).all()
        closes = ShiftClose.query.order_by(ShiftClose.created_at.asc()).all()

        # si mandan start/end por query, filtramos (por day del shift)
        if start_raw or end_raw:
            shift_ids_in_range = [
                s.id for s in shifts
                if (s.day >= start_date and s.day <= end_date)
            ]
            shifts = [s for s in shifts if s.id in shift_ids_in_range]
            expenses = [e for e in expenses if e.shift_id in shift_ids_in_range]
            closes = [c for c in closes if c.shift_id in shift_ids_in_range]

        shift_lookup = {s.id: s for s in shifts}
        close_lookup = {c.shift_id: c for c in closes}

        shifts_payload = []
        for s in shifts:
            c = close_lookup.get(s.id)
            cash_final = int(c.ending_cash or 0) if c else 0
            shifts_payload.append({
                "day": s.day.isoformat(),
                "turn": s.turn,
                "responsible": s.responsible,
                "opening_cash": int(s.opening_cash or 0),
                "retirado": int(s.sales_cash or 0),
                "cash_final": cash_final,
                "sales_cash": int(cash_bruto(s, cash_final=cash_final)),
                "sales_mp": int(s.sales_mp or 0),
                "sales_pya": int(getattr(s, "sales_pya", 0) or 0),
                "sales_rappi": int(getattr(s, "sales_rappi", 0) or 0),
                "delivery_data_json": s.delivery_data_json,
                "hour_shift": s.hour_shift,
                "hour_in": s.hour_in,
                "hour_out": s.hour_out,
                "status": s.status,
                "closed_at": s.closed_at.isoformat() if s.closed_at else None,
            })

        expenses_payload = []
        for e in expenses:
            sh = shift_lookup.get(e.shift_id)
            expenses_payload.append({
                "shift_day": sh.day.isoformat() if sh else None,
                "shift_turn": sh.turn if sh else None,
                "category": e.category,
                "amount": int(e.amount or 0),
                "note": e.note,
                "created_at": e.created_at.isoformat() if e.created_at else None,
            })

        closes_payload = []
        for c in closes:
            sh = shift_lookup.get(c.shift_id)
            closes_payload.append({
                "shift_day": sh.day.isoformat() if sh else None,
                "shift_turn": sh.turn if sh else None,
                "withdrawn_cash": int(c.withdrawn_cash or 0),
                "ending_calc": int(c.ending_calc or 0),
                "ending_cash": int(c.ending_cash or 0),
                "difference": int(c.difference or 0),
                "note": c.note,
                "close_ok": int(c.close_ok or 1),
                "created_at": c.created_at.isoformat() if c.created_at else None,
                "edited_by": c.edited_by,
                "edited_at": c.edited_at.isoformat() if c.edited_at else None,
                "edit_reason": c.edit_reason,
                "edit_count": int(c.edit_count or 0),
            })

        payload = {
            "type": "caja_backup",
            "version": 1,
            "exported_at": datetime.utcnow().isoformat(),
            "range": {
                "start": start_date.isoformat() if (start_raw or end_raw) else None,
                "end": end_date.isoformat() if (start_raw or end_raw) else None
            },
            "shifts": shifts_payload,
            "expenses": expenses_payload,
            "closes": closes_payload,
        }

        bio = BytesIO()
        bio.write(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
        bio.seek(0)
        filename = f"caja_backup_{date.today().isoformat()}.json"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/json")

    # ==============================
    # IMPORT CAJA (Opcion A)
    # ==============================
    IMPORT_CAJA_HTML = """
    <!doctype html>
    <html lang="es">
    <head><meta charset="utf-8"><title>Import Caja</title>{{ base_css|safe }}</head>
    <body style="max-width:950px;">
      <a href="{{ url_for('caja_index') }}"><- Volver</a>
      <h2>Importar datos (Excel)</h2>

      <div class="box" style="margin-bottom:12px; border-color:#ffbf66;background:#fff0d6;">
        <b>Formato:</b> hojas <code>shifts</code> y <code>expenses</code>.<br>
        Se importan turnos y egresos, y el turno queda <b>CERRADO</b>.<br>
        <b>Regla aplicada:</b> Retirado = Efectivo neto (Efectivo bruto - Caja inicial).
      </div>

      {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
      {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

      <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
        <div class="box">
          <label><b>Archivo Excel</b></label><br>
          <input type="file" name="file" accept=".xlsx" required>
          <br><br>
          <label><b>Modo</b></label><br>
          <select name="mode">
            <option value="skip" selected>Si existe el turno, NO tocarlo (skip)</option>
            <option value="replace">Si existe el turno, REEMPLAZAR (replace)</option>
          </select>
          <div class="muted" style="margin-top:6px;">Tip: para "pisar todo", usa modo <b>replace</b>.</div>
        </div>
        <br>
        <button class="btn">Importar</button>
      </form>
    </body>
    </html>
    """

    def _excel_to_date(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v).strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s[:10])
        except:
            pass
        try:
            dd, mm, yy = s.split("/")
            return date(int(yy), int(mm), int(dd))
        except:
            return None

    @app.route("/import/caja", methods=["GET","POST"])
    @login_required
    def import_caja():
        err = None
        msg = None

        if request.method == "POST":
            mode = (request.form.get("mode") or "skip").strip()
            f = request.files.get("file")
            if not f:
                err = "No se recibio archivo."
            else:
                try:
                    wb = load_workbook(f, data_only=True)
                    if "shifts" not in wb.sheetnames:
                        raise ValueError("Error: Falta la hoja 'shifts'.")
                    if "expenses" not in wb.sheetnames:
                        raise ValueError("Error: Falta la hoja 'expenses'.")

                    ws_s = wb["shifts"]
                    ws_e = wb["expenses"]

                    rows_s = list(ws_s.iter_rows(values_only=True))
                    if len(rows_s) < 2:
                        raise ValueError("Error: La hoja 'shifts' no tiene datos.")

                    headers_s = [str(x).strip() if x is not None else "" for x in rows_s[0]]
                    need_s = ["day","turn","responsible","opening_cash","sales_cash","sales_mp","sales_pya","sales_rappi"]
                    for n in need_s:
                        if n not in headers_s:
                            raise ValueError(f"Error: Falta columna '{n}' en shifts.")
                    idxs = {k:i for i,k in enumerate(headers_s)}

                    rows_e = list(ws_e.iter_rows(values_only=True))
                    headers_e = [str(x).strip() if x is not None else "" for x in rows_e[0]] if rows_e else []
                    need_e = ["day","turn","category","amount","note"]
                    for n in need_e:
                        if n not in headers_e:
                            raise ValueError(f"Error: Falta columna '{n}' en expenses.")
                    idxe = {k:i for i,k in enumerate(headers_e)}

                    imported = 0
                    replaced = 0

                    exp_map = {}
                    for r in rows_e[1:]:
                        if not r or all(v is None or str(v).strip()=="" for v in r):
                            continue
                        dd = _excel_to_date(r[idxe["day"]])
                        tt = (str(r[idxe["turn"]] or "").strip().upper())
                        if tt not in ("MORNING","AFTERNOON"):
                            continue
                        cat = (str(r[idxe["category"]] or "").strip())
                        amt = int(r[idxe["amount"]] or 0)
                        note = (str(r[idxe["note"]] or "").strip() or None)
                        exp_map.setdefault((dd, tt), []).append((cat, amt, note))

                    for r in rows_s[1:]:
                        if not r or all(v is None or str(v).strip()=="" for v in r):
                            continue

                        dd = _excel_to_date(r[idxs["day"]])
                        tt = (str(r[idxs["turn"]] or "").strip().upper())
                        if dd is None:
                            continue
                        if tt not in ("MORNING","AFTERNOON"):
                            continue

                        responsible = str(r[idxs["responsible"]] or "").strip() or ""
                        opening_cash = int(r[idxs["opening_cash"]] or 0)
                        # Some historical imports had the "responsable" column shifted with opening cash.
                        if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", responsible):
                            if opening_cash == 0:
                                try:
                                    opening_cash = int(float(str(r[idxs["responsible"]] or 0).replace(",", ".")))
                                except Exception:
                                    pass
                            responsible = "Bernardo"
                        sales_cash = int(r[idxs["sales_cash"]] or 0)
                        sales_mp = int(r[idxs["sales_mp"]] or 0)
                        sales_pya = int(r[idxs["sales_pya"]] or 0)
                        sales_rappi = int(r[idxs["sales_rappi"]] or 0)

                        existing = Shift.query.filter_by(day=dd, turn=tt).first()
                        if existing:
                            if mode == "skip":
                                continue
                            c = ShiftClose.query.filter_by(shift_id=existing.id).first()
                            if c:
                                db.session.delete(c)
                            CashExpense.query.filter_by(shift_id=existing.id).delete()
                            db.session.delete(existing)
                            db.session.flush()
                            replaced += 1

                        s = Shift(
                            day=dd, turn=tt, responsible=responsible,
                            opening_cash=opening_cash,
                            sales_cash=sales_cash, sales_mp=sales_mp,
                            sales_pya=sales_pya, sales_rappi=sales_rappi,
                            delivery_data_json=None,
                            hour_shift=None,
                            hour_in=None,
                            hour_out=None,
                            status="CLOSED", closed_at=datetime.utcnow()
                        )
                        db.session.add(s)
                        db.session.flush()

                        for (cat, amt, note) in exp_map.get((dd, tt), []):
                            if cat not in CATEGORIES:
                                cat = "Delivery / Cadete (efectivo)"
                            if cat.startswith("Otros") and not note:
                                note = "import"
                            if amt and amt > 0:
                                db.session.add(CashExpense(shift_id=s.id, category=cat, amount=int(amt), note=note))

                        withdrawn = max(0, int(sales_cash) - int(opening_cash))
                        ending_calc = calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn)
                        ending_real = ending_calc
                        diff = 0

                        db.session.add(ShiftClose(
                            shift_id=s.id,
                            withdrawn_cash=withdrawn,
                            ending_calc=ending_calc,
                            ending_cash=ending_real,
                            difference=diff,
                            note=None,
                            close_ok=1,
                            edit_count=0
                        ))

                        imported += 1

                    db.session.commit()
                    backup_caja_local_y_drive()
                    msg = f"Import OK. Turnos importados: {imported}. Reemplazados: {replaced}."

                except Exception as ex:
                    db.session.rollback()
                    err = str(ex)

        return render_template_string(IMPORT_CAJA_HTML, base_css=BASE_CSS, err=err, msg=msg)

    IMPORT_CAJA_JSON_HTML = """
    <!doctype html>
    <html lang="es">
    <head><meta charset="utf-8"><title>Import Caja (JSON)</title>{{ base_css|safe }}</head>
    <body style="max-width:950px;">
      <a href="{{ url_for('caja_index') }}"><- Volver</a>
      <h2>Importar Backup JSON - Caja</h2>

      <div class="box" style="margin-bottom:12px;">
        <b>Formato:</b> archivo JSON exportado por "Exportar JSON".<br>
        <b>Modo skip:</b> si el turno (dia+turno) ya existe, no lo toca.<br>
        <b>Modo replace:</b> borra ese turno (incluye egresos + cierre) y lo recrea.
      </div>

      {% if msg %}<div class="box" style="border-color:#7ad2a4;background:#dff5e6;"><b>{{msg}}</b></div>{% endif %}
      {% if err %}<div class="box" style="border-color:#ffb866;background:#ffe6d6;"><b>Error:</b> {{err}}</div>{% endif %}

      <form method="post" enctype="multipart/form-data" style="margin-top:12px;">
        <div class="box">
          <label><b>Archivo JSON</b></label><br>
          <input type="file" name="file" accept=".json" required>
          <br><br>
          <label><b>Modo</b></label><br>
          <select name="mode">
            <option value="skip" selected>Si existe el turno, NO tocarlo (skip)</option>
            <option value="replace">Si existe el turno, REEMPLAZAR (replace)</option>
          </select>
        </div>
        <br>
        <button class="btn">Importar JSON</button>
      </form>
    </body>
    </html>
    """

    @app.route("/import/caja/json", methods=["GET","POST"])
    @login_required
    def import_caja_json():
        err = None
        msg = None

        if request.method == "POST":
            mode = (request.form.get("mode") or "skip").strip()
            f = request.files.get("file")
            if not f:
                err = "No se recibio archivo."
            else:
                try:
                    payload = json.load(f)

                    if not isinstance(payload, dict) or payload.get("type") != "caja_backup":
                        raise ValueError("JSON invalido: no parece un backup de Caja.")

                    shifts = payload.get("shifts") or []
                    expenses = payload.get("expenses") or []
                    closes = payload.get("closes") or []

                    # index de expenses/closes por (day, turn)
                    exp_map = {}
                    for e in expenses:
                        dd = (e.get("shift_day") or "").strip()
                        tt = (e.get("shift_turn") or "").strip().upper()
                        if not dd or tt not in ("MORNING","AFTERNOON"):
                            continue
                        exp_map.setdefault((dd, tt), []).append(e)

                    close_map = {}
                    for c in closes:
                        dd = (c.get("shift_day") or "").strip()
                        tt = (c.get("shift_turn") or "").strip().upper()
                        if not dd or tt not in ("MORNING","AFTERNOON"):
                            continue
                        close_map[(dd, tt)] = c

                    imported = 0
                    replaced = 0
                    skipped = 0

                    for s in shifts:
                        dd = (s.get("day") or "").strip()
                        tt = (s.get("turn") or "").strip().upper()
                        if not dd or tt not in ("MORNING","AFTERNOON"):
                            continue

                        day_obj = date.fromisoformat(dd)
                        responsible = (s.get("responsible") or "").strip()
                        opening_cash = int(s.get("opening_cash") or 0)
                        if (not responsible) or re.fullmatch(r"[0-9]+([\.,][0-9]+)?", str(responsible)):
                            if opening_cash == 0:
                                try:
                                    opening_cash = int(float(str(s.get("responsible") or 0).replace(",", ".")))
                                except Exception:
                                    pass
                            responsible = "Bernardo"
                        sales_cash = int(s.get("sales_cash") or 0)
                        sales_mp = int(s.get("sales_mp") or 0)
                        sales_pya = int(s.get("sales_pya") or 0)
                        sales_rappi = int(s.get("sales_rappi") or 0)

                        existing = Shift.query.filter_by(day=day_obj, turn=tt).first()
                        if existing:
                            if mode == "skip":
                                skipped += 1
                                continue

                            # replace: borrar cierre + egresos + shift
                            c_old = ShiftClose.query.filter_by(shift_id=existing.id).first()
                            if c_old:
                                db.session.delete(c_old)
                            CashExpense.query.filter_by(shift_id=existing.id).delete()
                            db.session.delete(existing)
                            db.session.flush()
                            replaced += 1

                        # crear shift
                        shift_status = (s.get("status") or "CLOSED").strip().upper()
                        if shift_status not in ("OPEN","CLOSED"):
                            shift_status = "CLOSED"

                        new_shift = Shift(
                            day=day_obj,
                            turn=tt,
                            responsible=responsible,
                            opening_cash=opening_cash,
                            sales_cash=sales_cash,
                            sales_mp=sales_mp,
                            sales_pya=sales_pya,
                            sales_rappi=sales_rappi,
                            delivery_data_json=s.get("delivery_data_json"),
                            hour_shift=s.get("hour_shift"),
                            hour_in=s.get("hour_in"),
                            hour_out=s.get("hour_out"),
                            status=shift_status,
                            closed_at=datetime.utcnow() if shift_status == "CLOSED" else None
                        )
                        db.session.add(new_shift)
                        db.session.flush()

                        # egresos
                        for e in exp_map.get((dd, tt), []):
                            cat = (e.get("category") or "").strip() or "Mantenimiento / Varios"
                            amt = int(e.get("amount") or 0)
                            note = (e.get("note") or None)

                            if cat not in CATEGORIES:
                                cat = "Mantenimiento / Varios"
                            if cat.startswith("Otros") and not note:
                                note = "import json"

                            if amt > 0:
                                db.session.add(CashExpense(
                                    shift_id=new_shift.id,
                                    category=cat,
                                    amount=amt,
                                    note=note
                                ))

                        # cierre (si existe en JSON). Si no, lo calculamos.
                        cdata = close_map.get((dd, tt))
                        if shift_status == "CLOSED":
                            if cdata:
                                withdrawn = int(cdata.get("withdrawn_cash") or 0)
                                ending_calc = int(cdata.get("ending_calc") or calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn))
                                ending_real = int(cdata.get("ending_cash") or ending_calc)
                                diff = int(cdata.get("difference") or (ending_real - ending_calc))
                                close_ok = int(cdata.get("close_ok") or 1)
                                note = cdata.get("note") or None

                                db.session.add(ShiftClose(
                                    shift_id=new_shift.id,
                                    withdrawn_cash=withdrawn,
                                    ending_calc=ending_calc,
                                    ending_cash=ending_real,
                                    difference=diff,
                                    note=note,
                                    close_ok=close_ok,
                                    edit_count=int(cdata.get("edit_count") or 0),
                                    edited_by=cdata.get("edited_by"),
                                    edited_at=datetime.fromisoformat(cdata["edited_at"]) if cdata.get("edited_at") else None,
                                    edit_reason=cdata.get("edit_reason"),
                                ))
                            else:
                                withdrawn = max(0, int(sales_cash) - int(opening_cash))
                                ending_calc = calc_ending_calc(int(sales_cash) - int(withdrawn), withdrawn)
                                ending_real = ending_calc
                                db.session.add(ShiftClose(
                                    shift_id=new_shift.id,
                                    withdrawn_cash=withdrawn,
                                    ending_calc=ending_calc,
                                    ending_cash=ending_real,
                                    difference=0,
                                    note=None,
                                    close_ok=1,
                                    edit_count=0
                                ))

                        imported += 1

                    db.session.commit()
                    backup_caja_local_y_drive()
                    msg = f"Import JSON OK. Importados: {imported}. Reemplazados: {replaced}. Omitidos (skip): {skipped}."
                except Exception as ex:
                    db.session.rollback()
                    err = str(ex)

        return render_template_string(IMPORT_CAJA_JSON_HTML, base_css=BASE_CSS, err=err, msg=msg)

    # ============================================================
    # =====================  ASISTENCIA (UI)  =====================
    # ============================================================

